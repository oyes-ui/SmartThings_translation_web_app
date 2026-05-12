# -*- coding: utf-8 -*-
"""
backend/checker_service.py
Refactored for Web API usage (FastAPI + SSE)
"""

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
import os
from google import genai
from google.genai import types
from datetime import datetime
from dotenv import load_dotenv
import asyncio
import json
import re
import pandas as pd
import io
import urllib.parse
from model_handler import ModelHandler
from prompt_builder import PromptBuilder

# RAG 연동 (DB가 없우면 graceful fallback)
try:
    from rag_retriever import get_retriever as _get_rag_retriever
except ImportError:
    _get_rag_retriever = None

# API Keys from Environment
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

# ----------------- Case Hard-rule Applicable Languages -----------------
CASE_APPLICABLE_LANG_PREFIXES = {
    "English", "German", "French", "Spanish", "Portuguese",
    "Italian", "Dutch", "Swedish", "Polish", "Turkish",
    "Indonesian", "Vietnamese", "Russian"
}

def _is_case_sensitive_language(lang_name: str) -> bool:
    if not lang_name:
        return False
    return any(lang_name.startswith(pref) for pref in CASE_APPLICABLE_LANG_PREFIXES)

class TranslationChecker:
    """
    Excel Translation Checker (Gemini Single Model) - Service Version
    """

    def __init__(
        self,
        model_name: str = "gpt-5-mini",
        max_concurrency: int = 10,
        short_text_whitelist=None,
        skip_llm_when_glossary_mismatch: bool = False,
        no_backtranslation: bool = False
    ):
        # API Setup
        self.model_name = model_name
        
        # Concurrency
        if max_concurrency < 1:
            max_concurrency = 1
        self.max_concurrency = max_concurrency
        self._sem = asyncio.Semaphore(self.max_concurrency)
            
        self.no_backtranslation = no_backtranslation
        
        # Concurrency
        if max_concurrency < 1:
            max_concurrency = 1
        self.max_concurrency = max_concurrency
        self._sem = asyncio.Semaphore(self.max_concurrency)

        # Whitelist
        default_whitelist = {"ok", "on", "off", "ai", "5g", "go", "up", "usb", "nfc"}
        if short_text_whitelist:
            if isinstance(short_text_whitelist, str):
                extra = {x.strip().lower() for x in short_text_whitelist.split(",") if x.strip()}
            else:
                extra = {str(x).strip().lower() for x in short_text_whitelist}
            default_whitelist |= extra
        self.short_text_whitelist = default_whitelist

        self.skip_llm_when_glossary_mismatch = skip_llm_when_glossary_mismatch

        # Glossary Structure: { source_term: { 'targets': {code: term, ...}, 'rule': ... } }
        self.glossary = {}
        self.glossary_re = None # Pre-compiled regex for fast lookup
        self.glossary_map = {}  # {lower_case_term: original_case_term}
        self.glossary_headers = []
        self.source_lang_code = None

        # Integrated components
        self.model_handler = ModelHandler()
        self.prompt_builder = PromptBuilder()

        # RAG Retriever (오프라인 전용, DB 없으면 None)
        self.rag_retriever = None
        if _get_rag_retriever:
            try:
                self.rag_retriever = _get_rag_retriever()
                if not self.rag_retriever.is_available():
                    self.rag_retriever = None
            except Exception:
                self.rag_retriever = None
    

    async def load_glossary_from_file(self, file_path: str, source_lang_code: str):
        """
        3행 구조의 용어집 CSV를 로드합니다.
        1행: 한국어, 영어_미국 등 (JSON 'code'와 매칭)
        2행: ko_KR, en_US 등
        3행: Lng 행
        """
        try:
            if not os.path.exists(file_path):
                return f"Glossary file not found: {file_path}"

            # 1. 헤더 3개 행을 먼저 읽어서 컬럼 인덱스 파악
            import pandas as pd
            df_headers = pd.read_csv(file_path, header=None, nrows=3, encoding='utf-8-sig').fillna("")
            
            num_cols = df_headers.shape[1]
            source_col_idx = -1
            rule_col_idx = -1
            # 소스 언어 컬럼 찾기 (1~3행 모두 검색)
            search_terms = []
            if source_lang_code:
                search_terms.append(source_lang_code.lower())
                # 한국어/영어 등 주요 언어 매칭 보강
                lang_alias = {
                    "korean": ["한국어", "ko_kr", "ko-kr", "kr", "ko"],
                    "한국어": ["korean", "ko_kr", "ko-kr", "kr", "ko"],
                    "english": ["영어", "en_us", "en-us", "us", "en_gb", "en-gb", "uk", "en"],
                    "영어": ["english", "en_us", "en-us", "us", "en_gb", "en-gb", "uk", "en"]
                }
                if source_lang_code.lower() in lang_alias:
                    search_terms.extend(lang_alias[source_lang_code.lower()])

            if search_terms:
                for c in range(num_cols):
                    for r in range(3):
                        val = str(df_headers.iloc[r, c]).strip().lower()
                        if val and any(term == val or term in val or val in term for term in search_terms):
                            source_col_idx = c
                            break
                    if source_col_idx != -1: break
            
            if source_col_idx == -1:
                # Fallback: 만약 못 찾으면 0번 컬럼(A열)을 원문으로 가정 (대부분의 용어집 구조)
                source_col_idx = 0
                print(f"ℹ Glossary: '{source_lang_code}' 컬럼을 찾지 못해 0번 컬럼을 기본값으로 사용합니다.")

            # 설명/규칙/비고 컬럼 찾기
            for c in range(num_cols):
                for r in range(3):
                    val = str(df_headers.iloc[r, c]).strip().lower()
                    if any(kw in val for kw in ["설명", "규칙", "rule", "비고", "note", "remark", "desc"]):
                        rule_col_idx = c
                        break
                if rule_col_idx != -1: break
            
            if rule_col_idx == -1:
                print("⚠ Glossary: '규칙/비고' 컬럼을 찾지 못했습니다. 규칙이 적용되지 않을 수 있습니다.")

            # 2. 데이터 로드 (4행부터)
            df_data = pd.read_csv(file_path, header=None, skiprows=3, encoding='utf-8-sig').fillna("")
            
            self.source_lang_code = source_lang_code

            # 소스 언어가 한국어인 경우 한국어 컬럼도 찾아서 이중 키 등록에 사용
            is_korean_source = any(
                t in source_lang_code.lower() for t in ["korean", "한국어", "ko_kr", "ko-kr"]
            )
            kr_col_idx = -1
            if is_korean_source:
                for c in range(num_cols):
                    for r in range(3):
                        val = str(df_headers.iloc[r, c]).strip().lower()
                        if val in ["한국어", "ko_kr", "ko-kr", "korean"]:
                            kr_col_idx = c
                            break
                    if kr_col_idx != -1:
                        break

            count = 0
            for _, row in df_data.iterrows():
                source_term = str(row[source_col_idx]).strip()
                if not source_term or source_term.lower() == 'lng':
                    continue

                rule = str(row[rule_col_idx]).strip() if rule_col_idx != -1 else ""

                targets = {}
                for c in range(num_cols):
                    if c == source_col_idx or c == rule_col_idx:
                        continue
                    
                    # 타겟 키는 1행(영어_미국 등) 또는 2행(en_US 등)에서 가져옴
                    header_key = str(df_headers.iloc[0, c]).strip() or str(df_headers.iloc[1, c]).strip()
                    if not header_key or header_key.lower() in ['key', 'lng']:
                        continue
                        
                    val = str(row[c]).strip()
                    if val:
                        targets[header_key] = val
                        # 2행의 코드값도 키로 추가 (en_US 등)
                        code_key = str(df_headers.iloc[1, c]).strip()
                        if code_key and code_key != header_key:
                            targets[code_key] = val
                
                if targets:
                    # 영문 key 컬럼 값으로 등록 (항상)
                    self.glossary[source_term] = {"targets": targets, "rule": rule}
                    count += 1

                    # 🔑 소스 언어가 한국어면 한국어 컬럼 값도 소스 키로 이중 등록
                    # 예) 'SmartThings' 키와 함께 '스마트싱스'도 키로 등록 → 한국어 원문 검수 시 매칭 가능
                    if is_korean_source and kr_col_idx != -1:
                        kr_term = str(row[kr_col_idx]).strip()
                        if kr_term and kr_term.lower() not in ("lng", "") and kr_term != source_term:
                            if kr_term not in self.glossary:
                                self.glossary[kr_term] = {"targets": targets, "rule": rule}

            self._compile_glossary_re()
            kr_note = " (한국어 원문 키 이중 등록 적용)" if is_korean_source and kr_col_idx != -1 else ""
            return f"✓ 용어집 로드 성공: {len(self.glossary)}개 항목 (매칭 기준: {source_lang_code}){kr_note}"

        except Exception as e:
            return f"Glossary load failed: {str(e)}"

    def _compile_glossary_re(self):
        """
        용어집의 소스어들을 하나의 정규표현식으로 컴파일하여 검색 성능을 최적화합니다.
        길이가 긴 단어부터 매칭되도록 정렬하여 부분 일치 오류를 방지합니다.
        """
        if not self.glossary:
            self.glossary_re = None
            return

        # 단어 길이가 긴 순서대로 정렬 (긴 단어가 우선 매칭되도록)
        sorted_terms = sorted(self.glossary.keys(), key=len, reverse=True)
        self.glossary_map = {t.lower(): t for t in sorted_terms}
        patterns = []
        
        for term in sorted_terms:
            escaped = re.escape(term)
            # 영어/숫자로 시작/종료되는 경우 단어 경계(\b)와 유사한 로직 적용 
            # (단, 한국어 조사가 붙는 경우를 고려하여 앞뒤가 영문/숫자인 경우만 경계 체크)
            pattern = escaped
            if term[0].isalnum():
                pattern = r'(?<![a-zA-Z0-9])' + pattern
            if term[-1].isalnum():
                pattern = pattern + r'(?![a-zA-Z0-9])'
            patterns.append(f"({pattern})")
            
        # 모든 패턴을 OR(|)로 결합
        try:
            self.glossary_re = re.compile('|'.join(patterns), re.IGNORECASE)
        except Exception as e:
            print(f"Regex compilation failed: {e}")
            self.glossary_re = None

    def _get_relevant_glossary_terms(self, source_text: str):
        """
        원문에서 용어집에 포함된 단어들을 추출합니다.
        """
        if not self.glossary or not self.glossary_re or not source_text:
            return []
            
        found_terms = set()
        for match in self.glossary_re.finditer(source_text):
            # 매칭된 텍스트를 소문자로 변환하여 원래 용어집 키 찾기
            matched_text = match.group(0).lower()
            original_term = self.glossary_map.get(matched_text)
            if original_term:
                found_terms.add(original_term)
        return list(found_terms)

    # ----------------- Excel Loader -----------------
    def load_excel_files(self, source_path, target_path, selected_sheets=None):
        pass

    def load_excel_data(self, source_path, target_path, cell_range="A:Z", selected_sheets=None, log_func=None, source_sheet_name=None):
        all_data = []
        if log_func: log_func(f"데이터 추출 범위: {cell_range}")
        try:
            wb_source = openpyxl.load_workbook(source_path, data_only=True)
            wb_target = openpyxl.load_workbook(target_path, data_only=True)
        except Exception as e:
            raise Exception(f"Excel load error: {e}")

        source_order = wb_source.sheetnames
        target_set = set(wb_target.sheetnames)
        common_sheets = [s for s in source_order if s in target_set]

        if not common_sheets:
            raise Exception("No matching sheets found between Source and Target files.")

        if selected_sheets:
            target_sheets = [s for s in selected_sheets if s in common_sheets]
            if not target_sheets:
                raise Exception(f"Selected sheets {selected_sheets} not found in common sheets.")
        else:
            target_sheets = common_sheets

        processed_sheets = []

        for sheet_name in target_sheets:
            # Determine source sheet for this item
            src_sheet = source_sheet_name if source_sheet_name else sheet_name
            
            if src_sheet not in wb_source.sheetnames or sheet_name not in wb_target.sheetnames:
                if log_func: log_func(f"⚠ [{sheet_name}] 시트가 원본 파일({src_sheet}) 또는 대상 파일에 없습니다. 건너뜁니다.")
                continue
            
            ws_source = wb_source[src_sheet]
            ws_target = wb_target[sheet_name]
            
            extracted_count = 0
            extracted_coords = set()
            # 쉼표로 구분된 범위를 나누고, 개별 셀(c16)이나 범위(c10:c20) 모두 지원
            range_parts = [r.strip() for r in cell_range.split(',') if r.strip()]
            
            for current_range in range_parts:
                try:
                    # Handle full range like "A:Z" or specific "A1:C10"
                    src_rows = ws_source[current_range]
                    tgt_rows = ws_target[current_range]
                    
                    # Normalize to list of rows if single cell
                    if not isinstance(src_rows, tuple) and not isinstance(src_rows, list): 
                        # it might be a single cell if range is "A1"
                        src_rows = ((src_rows,),)
                        tgt_rows = ((tgt_rows,),)

                    for source_row, target_row in zip(src_rows, tgt_rows):
                        for s_cell, t_cell in zip(source_row, target_row):
                            if s_cell.coordinate in extracted_coords: 
                                continue
                            extracted_coords.add(s_cell.coordinate)
                            
                            s_val = str(s_cell.value).strip() if s_cell.value is not None else ""
                            t_val = str(t_cell.value).strip() if t_cell.value is not None else ""
                            
                            if s_val and t_val and s_val.lower() != "x" and t_val.lower() != "x":
                                # Extract row_key from column B of the same row
                                row_idx = s_cell.row
                                row_key = str(ws_source[f"B{row_idx}"].value).strip() if ws_source[f"B{row_idx}"].value else ""
                                
                                all_data.append({
                                    "cell_ref": s_cell.coordinate,
                                    "sheet_name": sheet_name,
                                    "source": s_val,
                                    "target": t_val,
                                    "row_key": row_key,
                                })
                                extracted_count += 1
                except Exception as e:
                    print(f"Error accessing range {current_range} in {sheet_name}: {e}")
                    continue
            
            if extracted_count > 0:
                processed_sheets.append(sheet_name)
                if log_func: log_func(f"✓ [{sheet_name}] {extracted_count}개 항목 추출 완료")
            else:
                if log_func: log_func(f"⚠ [{sheet_name}] 해당 범위({cell_range})에 유효한 데이터가 없습니다.")
        
        return all_data, processed_sheets

    # ----------------- Helpers -----------------
    async def _with_semaphore(self, coro):
        async with self._sem:
            return await coro

    def _build_glossary_lines_for_code(self, target_lang_code: str, source_text: str = None):
        if not self.glossary:
            return "용어집 없음"
        
        # 원문이 주어지면 관련 용어만 필터링, 아니면 전체 (하위 호환성)
        if source_text:
            relevant_terms = self._get_relevant_glossary_terms(source_text)
        else:
            relevant_terms = list(self.glossary.keys())
            
        if not relevant_terms:
            return "관련 용어 없음"

        out = []
        for term in relevant_terms:
            meta = self.glossary[term]
            tgt = meta["targets"].get(target_lang_code)
            if not tgt:
                continue
            rule = meta.get("rule")
            rule_info = f" (규칙: {rule})" if rule else ""
            out.append(f"- 원어: {term} → 대상어({target_lang_code}): {tgt}{rule_info}")
        
        return "\n".join(out) if out else f"용어집에 '{target_lang_code}' 타겟 항목 없음"

    def _get_glossary_context_as_dict(self, target_lang_code: str, source_text: str | None = None, skip_deactivated: bool = False):
        """
        용어집 데이터를 JSON/Dict 형태로 반환합니다. (프롬프트 최적화용)
        """
        if not self.glossary:
            return {}
        
        # 정규표현식이 컴파일되지 않은 경우 컴파일 시도
        if not self.glossary_re:
            self._compile_glossary_re()
            
        relevant_terms = self._get_relevant_glossary_terms(source_text) if source_text else list(self.glossary.keys())
        
        context = {}
        for term in relevant_terms:
            tgt = self.glossary[term]["targets"].get(target_lang_code)
            rule = self.glossary[term].get("rule", "").lower()
            clean_rule = rule.replace(" ", "")
            
            if skip_deactivated and ("비활성화" in clean_rule or "deactivate" in clean_rule or "disable" in clean_rule):
                continue
            
            if tgt:
                if not self.prompt_builder.should_wrap_glossary("", rule):
                    context[term] = f"{tgt} (EXCEPTION: Do NOT wrap '{tgt}' in brackets)"
                else:
                    context[term] = tgt
        return context

    def _precheck_glossary_mismatch(self, source_text: str, target_text: str, target_lang_code: str):
        if not self.glossary or not target_lang_code:
            return []
        
        relevant_terms = self._get_relevant_glossary_terms(source_text)
        if not relevant_terms:
            return []

        mismatches = []
        tgt_lower = target_text.lower()
        
        for s_term in relevant_terms:
            meta = self.glossary[s_term]
            
            rule = meta.get("rule", "").lower()
            clean_rule = rule.replace(" ", "")
            if "비활성화" in clean_rule or "deactivate" in clean_rule or "disable" in clean_rule:
                continue

            t_term = meta["targets"].get(target_lang_code)
            if not t_term: continue
            
            if t_term.lower() not in tgt_lower:
                mismatches.append(f"[미적용] '{s_term}' → '{t_term}' 미적용")
        return mismatches

    def _check_glossary_casing(self, source_text: str, target_text: str, target_lang_code: str):
        if not self.glossary or not target_lang_code or not source_text or not target_text:
            return []
            
        relevant_terms = self._get_relevant_glossary_terms(source_text)
        if not relevant_terms:
            return []

        issues = []
        tgt_lower = target_text.lower()
        
        for s_term in relevant_terms:
            meta = self.glossary[s_term]

            rule = meta.get("rule", "").lower()
            clean_rule = rule.replace(" ", "")
            if "비활성화" in clean_rule or "deactivate" in clean_rule or "disable" in clean_rule:
                continue

            t_term = meta["targets"].get(target_lang_code)
            if not t_term: continue
            
            if t_term.lower() in tgt_lower:
                idx = tgt_lower.find(t_term.lower())
                if idx == -1: continue
                actual = target_text[idx:idx + len(t_term)]
                if actual.lower() == t_term.lower() and actual != t_term:
                    issues.append(f"[대소문자] 용어집 '{t_term}'의 표기가 '{actual}'로 사용됨")
        return issues

    def _check_glossary_brackets(self, source_text: str, target_text: str, target_lang_code: str, target_lang: str, row_key: str = ""):
        if not self.glossary or not target_lang_code or not source_text or not target_text:
            return []
            
        relevant_terms = self._get_relevant_glossary_terms(source_text)
        if not relevant_terms:
            return []

        issues = []
        brackets = self.prompt_builder.get_brackets(target_lang)
        b_left, b_right = brackets[0], brackets[1]

        for s_term in relevant_terms:
            meta = self.glossary[s_term]
            
            rule = meta.get("rule", "").lower()
            clean_rule = rule.replace(" ", "")
            if "비활성화" in clean_rule or "deactivate" in clean_rule or "disable" in clean_rule:
                continue
                
            t_meta = meta["targets"]
            target_val = t_meta.get(target_lang_code)
            if not target_val:
                for k, v in t_meta.items():
                    if k.lower() == target_lang_code.lower():
                        target_val = v
                        break
            if not target_val: 
                continue
                
            clean_val = re.sub(r'\(.*?\)', '', target_val).strip()
            if not clean_val:
                continue

            should_exclude_bracket = not self.prompt_builder.should_wrap_glossary(row_key, rule)
            
            # Check all occurrences in target_text
            pattern = re.escape(clean_val)
            matches = list(re.finditer(pattern, target_text, re.IGNORECASE))
            
            if not matches:
                continue
                
            for m in matches:
                idx = m.start()
                has_left = (idx > 0 and target_text[idx-1] == b_left)
                idx_end = m.end()
                has_right = (idx_end < len(target_text) and target_text[idx_end] == b_right)
                has_brackets = has_left and has_right
                
                if should_exclude_bracket and has_brackets:
                    issues.append(f"[괄호 오류] '{clean_val}'는 규칙상 괄호 제외 대상이나 괄호({brackets})가 사용됨")
                    break # Report once per term per cell
                elif not should_exclude_bracket and not has_brackets:
                    issues.append(f"[괄호 오류] '{clean_val}'에 괄호({brackets})가 누락되었을 수 있습니다 (내비게이션 경로 제외)")
                    break # Report once per term per cell
                    
        return issues

    def _analyze_sentence_case(self, target_text: str, target_lang: str):
        if not target_text or not _is_case_sensitive_language(target_lang):
            return None, None
        
        text = target_text.strip()
        sentences = re.split(r'(?<=[\.!?])\s+', text)
        sentences = [s for s in sentences if s.strip()]
        if not sentences:
            return None, None
            
        report_lines = []
        fixed_sentences = []
        
        for idx, sent in enumerate(sentences, start=1):
            s = sent
            first_alpha_index = None
            for i, ch in enumerate(s):
                if ch.isalpha():
                    first_alpha_index = i
                    break
            
            if first_alpha_index is None:
                fixed_sentences.append(s)
                continue
                
            first_char = s[first_alpha_index]
            rest = s[first_alpha_index+1:]
            
            is_sentence_case = first_char.isupper() and rest == rest.lower()
            extra_caps = sum(1 for ch in rest if ch.isalpha() and ch.isupper())
            
            status = "문장형(첫 글자만 대문자)" if is_sentence_case else "문장형 아님"
            report_lines.append(f"- 문장 {idx}: {status}, 추가 대문자 수: {extra_caps}개")
            
            fixed_rest = rest.lower()
            fixed_sent = s[:first_alpha_index] + first_char.upper() + fixed_rest
            fixed_sentences.append(fixed_sent)
            
        simple_fixed_text = " ".join(fixed_sentences)
        report = "\n".join(report_lines)
        if simple_fixed_text == target_text:
            simple_fixed_text = None
        return report, simple_fixed_text
    async def _run_llm_translation(self, text, target_lang, model_name="gemini-2.5-flash", bx_style_on=False, glossary_context=None, rag_context=None, row_key="", source_lang="English", rag_identity_match=True):
        """
        내부 전용 번역 메서드: JSON 프롬프트 생성 및 LLM 호출을 담당합니다.
        """
        # 입력 데이터 구조화
        input_data = {
            "context_key": row_key,
            "source_text": text,
            "target_language": target_lang,
            "glossary": glossary_context if isinstance(glossary_context, dict) else {},
            "formatting": self.prompt_builder.build_input_formatting(target_lang)
        }

        # RAG 예시 주입: 유사 번역 사례 최대 2건을 시스템 프롬프트에 첨부
        if self.rag_retriever and not rag_context:
            try:
                rag_context = self.rag_retriever.format_for_prompt(
                    text,
                    target_lang,
                    source_lang=source_lang,
                    n_results=2,
                    identity_match_enabled=rag_identity_match
                )
                if rag_context:
                    pass # PromptBuilder handles injection via build_translation_prompt
            except Exception:
                pass  # RAG 실패해도 번역은 정상 진행

        system_prompt = self.prompt_builder.build_translation_prompt(
            target_lang=target_lang,
            source_lang=source_lang,
            bx_style_on=bx_style_on,
            rag_context=rag_context,
            row_key=row_key,
            glossary_context=glossary_context
        )

        try:
            prompt = json.dumps(input_data, ensure_ascii=False, indent=2)
            response_data = await self.model_handler.generate_content(
                prompt, 
                model_name=model_name, 
                system_instruction=system_prompt,
                response_json=True
            )

            if isinstance(response_data, dict):
                return response_data.get("translation", str(response_data))
            return str(response_data)
        except Exception as e:
            return f"[번역 오류] {str(e)}"


    async def _run_llm_audit(self, source_text, translated_text, target_lang, model_name="gpt-5-mini"):
        """
        내부 전용 감사 메서드: JSON 구조화된 프롬프트를 사용하여 품질 검수합니다.
        """
        input_data = {
            "source_text": source_text,
            "translated_text": translated_text,
            "language": target_lang
        }

        system_instruction = "당신은 원문과 번역문을 대조하여 언어적 정확성과 품질을 평가하는 전문 검수자입니다. 분석 결과를 한국어 의견으로 반환하세요."
        
        try:
            prompt = json.dumps(input_data, ensure_ascii=False, indent=2)
            response = await self.model_handler.generate_content(
                prompt,
                model_name=model_name,
                system_instruction=system_instruction + "\nReturn a JSON object with 'reasoning' and 'is_accurate' (bool) keys.",
                response_json=True
            )
            
            if isinstance(response, dict):
                return response.get("reasoning", str(response))
            return str(response)
        except Exception as e:
            return f"[감사 오류] {str(e)}"


    # ----------------- LLM Calls -----------------
    async def check_with_llm_qa(self, source_text, target_text, source_lang, target_lang, target_lang_code, row_key: str = ""):
        glossary_dict = self._get_glossary_context_as_dict(target_lang_code, source_text=source_text)
        
        # QA용 구조화된 프롬프트 데이터
        input_data = {
            "source": {"lang": source_lang, "text": source_text},
            "translation": {"lang": f"{target_lang}/{target_lang_code}", "text": target_text},
            "glossary": glossary_dict,
            "context_key": row_key
        }

        prompt = f"""당신은 전문 번역 검수 전문가입니다. 아래 JSON 데이터를 분석하여 상세한 검수 결과를 반환하세요.

[Input Data]
{json.dumps(input_data, ensure_ascii=False, indent=2)}
"""
        system_instruction = self.prompt_builder.build_audit_prompt(
            source_lang=source_lang,
            target_lang=target_lang,
            target_lang_code=target_lang_code,
            row_key=row_key,
            glossary_context=glossary_dict
        )
        try:
            # ModelHandler가 JSON 모드를 지원하므로 딕셔너리로 바로 받을 수 있음
            response = await self.model_handler.generate_content(
                prompt, 
                model_name=self.model_name, 
                system_instruction=system_instruction,
                response_json=True
            )
            
            if isinstance(response, dict) and response.get("error") == "parsing_failed":
                 err_msg = f"[파싱/서버 오류]: Gemini가 유효한 JSON을 반환하지 않았거나 서버가 응답하지 않습니다. (원문: {str(response.get('original_text', ''))[:100]}...)"
                 eval_json = json.dumps({"evaluation": [{"category": "에러", "comment": err_msg}], "is_excellent": False, "suggested_fix": ""}, ensure_ascii=False)
                 return (err_msg, eval_json)

            if isinstance(response, str):
                 eval_text = f"[QA 분석 결과]\n{response}"
                 eval_json = json.dumps({"evaluation": [{"category": "결과", "comment": response}], "is_excellent": False, "suggested_fix": ""}, ensure_ascii=False)
                 return (eval_text, eval_json)
            
            # 1. Build beautiful human-readable text
            eval_list = response.get("evaluation", [])
            lines = []
            for item in eval_list:
                cat = item.get("category", "")
                com = item.get("comment", "")
                if cat and com:
                   lines.append(f"- {cat}: {com}")
            
            eval_text = "\n".join(lines) if lines else "검수 결과 없음."
            if response.get("is_excellent"):
                eval_text += "\n\n최종 평가: 우수, 주요 문제 없음."
            if response.get("suggested_fix"):
                eval_text += f"\n\n[수정안 제안]:\n{response['suggested_fix']}"

            # 2. Extract JSON payload
            eval_json = json.dumps({
                "evaluation": eval_list,
                "is_excellent": response.get("is_excellent", False),
                "suggested_fix": response.get("suggested_fix", "")
            }, ensure_ascii=False)
            
            return (eval_text, eval_json)
        except Exception as e:
            err_text = f"[QA 오류]: {str(e)}"
            err_json = json.dumps({"evaluation": [{"category": "에러", "comment": str(e)}], "is_excellent": False, "suggested_fix": ""}, ensure_ascii=False)
            return (err_text, err_json)






    async def get_back_translation(self, target_text, target_lang, source_lang):
        prompt = (
            f"다음 {target_lang} 텍스트를 {source_lang}으로 다시 번역해주세요. "
            f"오직 번역된 텍스트만 제공해야 합니다. 다른 설명이나 텍스트는 포함하지 마세요.\n\n{target_text}"
        )
        try:
            return await self.model_handler.generate_content(prompt, model_name=self.model_name)
        except Exception as e:
            return f"[역번역 오류]: {e}"

    def _apply_rich_text(self, text: str, keywords: list, base_font=None):
        """
        텍스트 내의 키워드를 파란색으로 하이라이트하되, 나머지 텍스트는 base_font의 스타일을 유지합니다.
        """
        if not text:
            return ""
        if not keywords:
            return text # 키워드가 없으면 일반 문자열 반환 (교체하지 않음으로써 기존 스타일 보존)
            
        sorted_keywords = sorted([k.strip() for k in keywords if k.strip()], key=len, reverse=True)
        if not sorted_keywords:
            return text

        # 정규표현식으로 키워드 위치 찾기
        pattern = '|'.join(re.escape(k) for k in sorted_keywords)
        matches = list(re.finditer(pattern, text, flags=re.IGNORECASE))
        if not matches:
            return text

        # 베이스 폰트 정보 추출 (없으면 기본값)
        # InlineFont는 Font와 달리 'name' 대신 'rFont' 사용
        font_params = {}
        if base_font:
            font_params = {
                "rFont": base_font.name,
                "sz": base_font.sz,
                "b": base_font.b,
                "i": base_font.i,
                "u": base_font.u,
                "strike": base_font.strike,
                "family": base_font.family,
                "charset": base_font.charset,
                "outline": base_font.outline,
                "shadow": base_font.shadow,
                "condense": base_font.condense,
                "extend": base_font.extend,
                "vertAlign": base_font.vertAlign,
                "scheme": base_font.scheme,
            }

        def get_font(is_keyword=False):
            params = font_params.copy()
            if is_keyword:
                params["color"] = "0000FF"  # 키워드만 파란색 명시, 나머지는 셀 기본 색상 상속
            # is_keyword=False 시 color 미설정 → Excel 셀 레벨 스타일에서 색상 상속
            return InlineFont(**params)


        parts = []
        last_end = 0
        for m in matches:
            start, end = m.span()
            if start > last_end:
                parts.append((text[last_end:start], get_font(False)))
            parts.append((text[start:end], get_font(True)))
            last_end = end
            
        if last_end < len(text):
            parts.append((text[last_end:], get_font(False)))

        rt = CellRichText()
        for segment_text, segment_font in parts:
            rt.append(TextBlock(text=segment_text, font=segment_font))
        return rt

    # ----------------- Process Single Item -----------------
    async def process_item(self, item, source_lang, default_target_lang, sheet_lang_map, default_target_lang_code, rag_identity_match=True):
        cell_ref = item["cell_ref"]
        sheet_name = item["sheet_name"]
        source = item["source"]
        target = item["target"]
        
        tgt_lang = sheet_lang_map.get(sheet_name, {}).get("lang", default_target_lang)
        tgt_code = sheet_lang_map.get(sheet_name, {}).get("code", default_target_lang_code)
        
        # Skip detection
        is_placeholder = (
            source.strip().lower() == "x" or
            target.strip().lower() == "x" or
            (len(source) <= 2 and len(target) <= 2 and source.lower() == target.lower()) or
            (not source.strip() and not target.strip()) or
            (source.strip().lower() == target.strip().lower() and len(source.strip()) < 10)
        )
        if source.strip().lower() in self.short_text_whitelist or target.strip().lower() in self.short_text_whitelist:
            is_placeholder = False
        
        pre_mismatch = self._precheck_glossary_mismatch(source, target, tgt_code)
        skip_llm = self.skip_llm_when_glossary_mismatch and bool(pre_mismatch)
        
        row_key = item.get("row_key", "")
        
        case_report, simple_case_fix = self._analyze_sentence_case(target, tgt_lang)
        glossary_case_issues = self._check_glossary_casing(source, target, tgt_code)
        glossary_bracket_issues = self._check_glossary_brackets(source, target, tgt_code, tgt_lang, row_key=row_key)
        
        # Construct Partial Report Sections
        case_section = "대소문자 하드룰(문장형) 점검:\n" + case_report if case_report else "별도 지적 사항 없음."
        if simple_case_fix:
            case_section += f"\n\n[단순 규칙 기반 문장형 변환안]:\n{simple_case_fix}"

        glossary_parts = []
        if pre_mismatch: glossary_parts.append("용어집 사전 감지:\n- " + "\n- ".join(pre_mismatch))
        if glossary_case_issues: glossary_parts.append("용어집 대소문자 표기 점검:\n" + "\n".join(f"- {msg}" for msg in glossary_case_issues))
        if glossary_bracket_issues: glossary_parts.append("용어집 괄호 규정 점검:\n" + "\n".join(f"- {msg}" for msg in glossary_bracket_issues))
        glossary_section = "\n\n".join(glossary_parts) if glossary_parts else "별도 지적 사항 없음."

        # RAG consistency check (Post-translation/audit, does not use LLM)
        rag_text = "별도 설정 없음."
        rag_json = "[]"
        
        # Skip RAG for placeholder or very short/empty text
        if is_placeholder and not pre_mismatch:
            rag_text = "[건너뜀: 짧은/무의미]"
        elif getattr(self, "rag_retriever", None) and self.rag_retriever.is_available():
            try:
                results = self.rag_retriever.retrieve(
                    source, 
                    sheet_name, 
                    source_lang=source_lang, 
                    n_results=2, 
                    exclude_same_source=False,
                    identity_match_enabled=rag_identity_match
                )
                if results:
                    # 1. Human-readable text
                    text_lines = []
                    for i, r in enumerate(results, 1):
                        match_info = f"{r['match_type'].upper()}"
                        if r['match_type'] == "semantic":
                            match_info += f" ({r['similarity_score']*100:.1f}%)"
                        
                        text_lines.append(f"[사례 {i}] {match_info} | {r.get('story_id','')} | {r['section_code']}\n- 번역: {r['target']}")
                    rag_text = "\n\n".join(text_lines)
                    
                    # 2. JSON payload
                    rag_data = []
                    for r in results:
                        rag_data.append({
                            "type": r['match_type'],
                            "score": round(r['similarity_score'] * 100, 1),
                            "story_id": r.get('story_id', ''),
                            "section": r.get('section_code', ''),
                            "source": r.get('source', ''),
                            "target": r.get('target', '')
                        })
                    rag_json = json.dumps(rag_data, ensure_ascii=False)
                else:
                    rag_text = "유사 과거 사례 없음."
                    rag_json = "[]"
            except Exception as e:
                rag_text = f"RAG 조회 오류: {str(e)}"
                rag_json = json.dumps([{"error": str(e)}], ensure_ascii=False)

        # Result container
        res = {
            "sheet_name": sheet_name,
            "cell_ref": cell_ref,
            "source": source,
            "target": target,
            "case_section": case_section,
            "glossary_section": glossary_section,
            "rag_text": rag_text,
            "rag_json": rag_json,
            "back_translation": "",
            "ai_text": "",
            "ai_json": ""
        }

        # Skip Logic
        if is_placeholder and not pre_mismatch:
            res["back_translation"] = "[건너뜀: 짧은/무의미]"
            res["ai_text"] = "[건너뜀: 짧은/무의미]"
            res["ai_json"] = "[]"
            return res

        # LLM Logic
        if skip_llm:
            res["back_translation"] = "[사전 감지로 LLM 호출 생략]"
            res["ai_text"] = "※ 용어집 사전 감지 결과를 우선 검토하세요."
            res["ai_json"] = "[]"
        else:
            # LLM Calls
            qa_task = self._with_semaphore(
                self.check_with_llm_qa(source, target, source_lang, tgt_lang, tgt_code, row_key=row_key)
            )
            
            if self.no_backtranslation:
                ai_tuple = await qa_task
                res["ai_text"], res["ai_json"] = ai_tuple
                res["back_translation"] = "[역번역 비활성화됨]"
            else:
                bt_task = self._with_semaphore(
                    self.get_back_translation(target, tgt_lang, source_lang)
                )
                ai_tuple, bt = await asyncio.gather(qa_task, bt_task)
                res["ai_text"], res["ai_json"] = ai_tuple
                res["back_translation"] = bt
        
        return res

    # ----------------- Generator Method for SSE -----------------
    async def run_inspection_async_generator(
        self,
        source_file_path,
        target_file_path,
        cell_range,
        source_lang,
        target_lang,
        target_lang_code,
        sheet_lang_map,
        glossary_url=None,
        selected_sheets: list = None,
        glossary_file_path: str = None,
        source_sheet_name: str = None,
        rag_identity_match: bool = True
    ):
        """
        Yields events:
        {"type": "log", "message": "..."}
        {"type": "progress", "current": n, "total": m, "percent": ...}
        {"type": "result_chunk", "data": formatted_string}
        {"type": "complete", "total": m, "output_data": all_text}
        """
        yield {"type": "log", "message": "용어집 로드 시작..."}
        
        # If source_sheet_name is provided, use it to infer source_lang
        if source_sheet_name and sheet_lang_map:
            source_info = sheet_lang_map.get(source_sheet_name)
            if source_info:
                source_lang = source_info.get('lang', source_lang)
        
        # Load Glossary
        if glossary_file_path:
             # Try to get more descriptive source lang for matching
             src_lookup = source_lang
             if source_sheet_name and sheet_lang_map:
                 s_info = sheet_lang_map.get(source_sheet_name)
                 if s_info:
                     src_lookup = s_info.get('code', s_info.get('lang', source_lang))

             yield {"type": "log", "message": f"용어집 파일 로드 중: {os.path.basename(glossary_file_path)} (기준: {src_lookup})"}
             msg = await self.load_glossary_from_file(glossary_file_path, src_lookup)
             yield {"type": "log", "message": msg}
        
        yield {"type": "log", "message": "엑셀 파일 및 시트 분석 중..."}
        try:
            # Use explicit selected_sheets if provided, otherwise fallback to map keys
            sel_sheets = selected_sheets if selected_sheets else (list(sheet_lang_map.keys()) if sheet_lang_map else None)
            
            log_messages = []
            def collect_log(m): log_messages.append(m)

            all_data, processed_sheets = self.load_excel_data(
                source_file_path, target_file_path, cell_range=cell_range, selected_sheets=sel_sheets, log_func=collect_log, source_sheet_name=source_sheet_name
            )
            for m in log_messages:
                yield {"type": "log", "message": m}
        except Exception as e:
            yield {"type": "error", "message": str(e)}
            return

        total_items = len(all_data)
        yield {"type": "log", "message": f"검수 대상: {len(processed_sheets)}개 시트, 총 {total_items}개 항목"}
        yield {"type": "progress", "current": 0, "total": total_items, "percent": 0}

        if total_items == 0:
            yield {"type": "complete", "total": 0, "output_data": "검수할 데이터가 없습니다."}
            return

        # Wrapper to track index
        async def process_with_index(index, item):
            res = await self.process_item(item, source_lang, target_lang, sheet_lang_map, target_lang_code, rag_identity_match=rag_identity_match)
            return index, res

        # Create tasks with index
        tasks = [
            process_with_index(i, item)
            for i, item in enumerate(all_data)
        ]

        completed_count = 0
        # Initialize list to store results in order
        ordered_results = [None] * total_items
        
        # Use asyncio.as_completed to yield progress, but store by index
        for future in asyncio.as_completed(tasks):
            try:
                index, res = await future
                completed_count += 1
                
                # Format result
                fmt_result = (
                    f"==========================================================================================\n"
                    f"[시트] {res['sheet_name']} | [셀] {res['cell_ref']}\n"
                    f"------------------------------------------------------------------------------------------\n\n"
                    f"[상세 - 원문]\n{res['source']}\n\n"
                    f"[상세 - 번역문]\n{res['target']}\n\n"
                    f"[상세 - 대소문자 점검]\n{res['case_section']}\n\n"
                    f"[상세 - 용어집 점검]\n{res['glossary_section']}\n\n"
                    f"[상세 - RAG 일관성 참고]\n{res.get('rag_text', '[별도 지적 사항 없음]')}\n\n"
                    f"[상세 - 역번역]\n{res['back_translation']}\n\n"
                    f"[상세 - AI 검수 결과]\n{res['ai_text']}\n\n"
                    f"[상세 - RAG Payload]\n{res.get('rag_json', '[]')}\n\n"
                    f"[상세 - AI Payload]\n{res['ai_json']}\n"
                )
                
                # Store in the correct slot
                ordered_results[index] = fmt_result
                
                percent = int((completed_count / total_items) * 100)
                yield {
                    "type": "progress",
                    "current": completed_count,
                    "total": total_items,
                    "percent": percent,
                    "log": f"[{res['sheet_name']}] {res['cell_ref']} 검수 완료"
                }
            except Exception as e:
                yield {"type": "log", "message": f"항목 처리 중 오류 발생: {str(e)}"}
                continue

        # Final Header/Footer assembly using ORDERED results
        final_text = []
        header = (
            f"--- 번역 검수 보고서 (Model: {self.model_name}) ---\n"
            f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"총 검수 항목: {total_items}개\n"
            f"용어집 항목: {len(self.glossary)}\n\n"
        )
        final_text.append(header)
        # Filter out None results if any failed
        valid_results = [r for r in ordered_results if r is not None]
        final_text.extend(valid_results)
        
        full_output = "".join(final_text)
        yield {"type": "complete", "total": total_items, "output_data": full_output}

    # ----------------- Integrated Translation & Audit Generator -----------------
    async def run_integrated_pipeline_generator(
        self,
        source_file_path,
        cell_range,
        bx_style_on,
        sheet_lang_map,
        translation_model="gemini-2.5-flash",
        audit_model="gpt-5.2",
        glossary_file_path=None,
        selected_sheets=None,
        source_sheet_name=None,
        skip_audit=False,
        source_lang="English",
        rag_identity_match=True
    ):
        """
        Translates source text, audits it, and saves it to a NEW Excel file.
        Yields SSE events same as inspection.
        """
        yield {"type": "log", "message": "Starting Integrated Translation & Audit Pipeline..."}
        
        # 1. Load Excel
        try:
            wb = openpyxl.load_workbook(source_file_path, rich_text=True)
        except Exception as e:
            yield {"type": "error", "message": f"Excel Load Error: {str(e)}"}
            return

        if not source_sheet_name:
            source_sheet_name = "KR(한국)"
        if source_sheet_name not in wb.sheetnames:
            yield {"type": "error", "message": f"Source sheet '{source_sheet_name}' not found."}
            return

        # Infer source lang
        source_lang = "Korean"
        if source_sheet_name in sheet_lang_map:
            source_lang = sheet_lang_map[source_sheet_name].get("lang", source_lang)

        # 2. Load Glossary
        if glossary_file_path:
            src_lookup = source_lang
            if source_sheet_name in sheet_lang_map:
                s_info = sheet_lang_map[source_sheet_name]
                src_lookup = s_info.get('code', s_info.get('lang', source_lang))

            yield {"type": "log", "message": f"Loading glossary (Match Base: {src_lookup})..."}
            msg = await self.load_glossary_from_file(glossary_file_path, src_lookup)
            yield {"type": "log", "message": msg}

        source_ws = wb[source_sheet_name]
        
        # 3. Extract source data
        source_data = []
        extracted_coords = set()
        # 콤마로 구분된 여러 범위나 단일 셀(c16, c21)을 모두 지원합니다.
        range_parts = [r.strip() for r in cell_range.split(',') if r.strip()]
        
        for current_range in range_parts:
            try:
                rows = source_ws[current_range]
                if not isinstance(rows, tuple) and not isinstance(rows, list):
                    rows = ((rows,),)
                for row in rows:
                    for cell in row:
                        if cell.coordinate in extracted_coords:
                            continue
                        extracted_coords.add(cell.coordinate)
                        
                        val = str(cell.value).strip() if cell.value is not None else ""
                        if val and val.lower() != "x":
                            row_idx = cell.row
                            row_key = str(source_ws[f"B{row_idx}"].value).strip() if source_ws[f"B{row_idx}"].value else ""
                            source_data.append({'text': str(cell.value).strip(), 'coord': cell.coordinate, 'row_key': row_key})
            except Exception as e:
                yield {"type": "error", "message": f"Error accessing range {current_range}: {e}"}
                continue

        if not source_data:
            yield {"type": "error", "message": "No source text found in range."}
            return

        # 4. Identify target sheets
        available_sheets = wb.sheetnames
        if selected_sheets:
            target_sheets = [s for s in selected_sheets if s in available_sheets and s != source_sheet_name]
        else:
            target_sheets = [s for s in available_sheets if s in sheet_lang_map and s != source_sheet_name]

        total_cells = len(source_data) * len(target_sheets)
        yield {"type": "log", "message": f"Plan: {len(target_sheets)} sheets, Total {total_cells} cells."}
        yield {"type": "progress", "current": 0, "total": total_cells, "percent": 0}

        # 5. Process in parallel (cells)
        completed_count = 0
        
        async def cell_worker(ws, index, item, target_lang, target_lang_code):
            nonlocal completed_count
            source_text = item['text']
            coord = item['coord']
            
            # Step 1: Translate
            # Translating: filter out terms explicitly deactivated to save tokens and not force LLM behavior
            glossary_dict = self._get_glossary_context_as_dict(target_lang_code, source_text=source_text, skip_deactivated=True)
            if not glossary_dict:
                glossary_dict = None

            # RAG Injection
            rag_context_str = None
            logs = []
            if getattr(self, "rag_retriever", None) and self.rag_retriever.is_available():
                try:
                    results = self.rag_retriever.retrieve(
                        source_text, 
                        ws.title, 
                        source_lang=source_lang, 
                        n_results=2, 
                        exclude_same_source=False,
                        identity_match_enabled=rag_identity_match
                    )
                    if results:
                        rag_context_str = self.rag_retriever.format_for_prompt(
                            source_text, 
                            ws.title, 
                            source_lang=source_lang, 
                            n_results=2,
                            identity_match_enabled=rag_identity_match
                        )
                        logs.append(f"[RAG] 유사 사례 {len(results)}건 적용됨 (시트: {ws.title}, 셀: {coord})")
                except Exception as e:
                    logs.append(f"[RAG] 검색 오류: {e}")

                translation = await self._run_llm_translation(
                    source_text, 
                    target_lang, 
                    model_name=translation_model,
                    bx_style_on=bx_style_on,
                    glossary_context=glossary_dict,
                    rag_context=rag_context_str,
                    row_key=item.get('row_key', ''),
                    source_lang=source_lang,
                    rag_identity_match=rag_identity_match
                )
            
            # Extract plain glossary targets (removing EXCEPTION strings if any)
            # Use case-insensitive target language code matching
            original_target_terms = []
            
            relevant_terms_for_highlight = self._get_relevant_glossary_terms(source_text)
            if relevant_terms_for_highlight:
                for s_term in relevant_terms_for_highlight:
                    meta = self.glossary.get(s_term)
                    if meta:
                        # Skip deactivated terms for highlighting
                        rule = meta.get("rule", "").lower()
                        clean_rule = rule.replace(" ", "")
                        if "비활성화" in clean_rule or "deactivate" in clean_rule or "disable" in clean_rule:
                            continue

                        t_meta = meta["targets"]
                        # Try exact match, then case-insensitive match
                        target_val = t_meta.get(target_lang_code)
                        if not target_val:
                            # Search by lowercase key
                            for k, v in t_meta.items():
                                if k.lower() == target_lang_code.lower():
                                    target_val = v
                                    break
                        
                        if target_val:
                            # 괄호()가 포함된 경우 제외 조건일 수 있으므로 순수 텍스트만 추출하거나 전체 포함
                            clean_val = re.sub(r'\(.*?\)', '', target_val).strip()
                            
                            # 반점(,)이나 슬래시(/)로 여러 단어가 기재된 경우(예: "단어1, 단어2") 각각 하이라이트 되도록 분리
                            for extract_str in (clean_val, target_val):
                                if not extract_str: continue
                                original_target_terms.append(extract_str.strip())
                                # 정규식으로 분리 (콤마 또는 슬래시 기준, 양옆 공백까지 같이 잡아냄)
                                split_terms = [x.strip() for x in re.split(r'[,/]', extract_str) if x.strip()]
                                if len(split_terms) > 1:
                                    original_target_terms.extend(split_terms)

            target_cell = ws[coord]
            target_cell.value = self._apply_rich_text(translation, original_target_terms, base_font=target_cell.font)

            item_data = {"cell_ref": coord,"sheet_name": ws.title,"source": source_text,"target": translation}
            
            if skip_audit:
                res = {
                    "sheet_name": ws.title,
                    "cell_ref": coord,
                    "source": source_text,
                    "target": translation,
                    "case_section": "[Bypassed]",
                    "glossary_section": "[Bypassed]",
                    "rag_text": "[Bypassed]",
                    "rag_json": "[]",
                    "back_translation": "[Bypassed]",
                    "ai_text": "[Bypassed: Translate Only Mode]",
                    "ai_json": "{}",
                    "logs": logs
                }
            else:
                res = await self.process_item(item_data, source_lang, target_lang, sheet_lang_map, target_lang_code, rag_identity_match=rag_identity_match)
                res["logs"] = logs
            
            return index, res

        # Flatten all cells into a list of tasks
        tasks = []
        idx = 0
        for sheet_name in target_sheets:
            ws = wb[sheet_name]
            sheet_info = sheet_lang_map[sheet_name]
            target_lang = sheet_info['lang']
            target_lang_code = sheet_info['code']
            
            for item in source_data:
                tasks.append(cell_worker(ws, idx, item, target_lang, target_lang_code))
                idx += 1

        ordered_results = [None] * len(tasks)
        
        # Yield per-cell progress
        for future in asyncio.as_completed(tasks):
            try:
                index, res = await future
                completed_count += 1
                
                # yield UI logs from cell_worker
                for msg in res.get("logs", []):
                    yield {"type": "log", "message": msg}

                fmt_result = (
                    f"==========================================================================================\n"
                    f"[시트] {res['sheet_name']} | [셀] {res['cell_ref']}\n"
                    f"------------------------------------------------------------------------------------------\n\n"
                    f"[상세 - 원문]\n{res['source']}\n\n"
                    f"[상세 - 번역문]\n{res['target']}\n\n"
                    f"[상세 - 대소문자 점검]\n{res['case_section']}\n\n"
                    f"[상세 - 용어집 점검]\n{res['glossary_section']}\n\n"
                    f"[상세 - RAG 일관성 참고]\n{res.get('rag_text', '[별도 지적 사항 없음]')}\n\n"
                    f"[상세 - 역번역]\n{res['back_translation']}\n\n"
                    f"[상세 - AI 검수 결과]\n{res['ai_text']}\n\n"
                    f"[상세 - RAG Payload]\n{res.get('rag_json', '[]')}\n\n"
                    f"[상세 - AI Payload]\n{res['ai_json']}\n"
                )
                ordered_results[index] = fmt_result
                
                percent = int((completed_count / total_cells) * 100)
                yield {
                    "type": "progress", 
                    "current": completed_count, 
                    "total": total_cells, 
                    "percent": percent, 
                    "log": f"[{res['sheet_name']}] {res['cell_ref']} 처리 완료"
                }
            except Exception as e:
                yield {"type": "log", "message": f"Cell processing error: {str(e)}"}

        # 6. Save results
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_excel_path = source_file_path.replace(".xlsx", f"_translated_{timestamp}.xlsx")
        wb.save(out_excel_path)
        
        header = (
            f"--- 번역 통합 검수 보고서 (Model: {self.model_name}) ---\n"
            f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"총 항목: {completed_count}개\n\n"
        )
        valid_results = [r for r in ordered_results if r is not None]
        report_text = header + "".join(valid_results)
        
        yield {
            "type": "complete", 
            "output_data": report_text, 
            "excel_path": out_excel_path
        }


    async def run_highlight_only_pipeline_generator(
        self,
        source_file_path,
        cell_range,
        sheet_lang_map,
        glossary_file_path=None,
        selected_sheets=None,
        source_sheet_name=None,
        source_lang="English"
    ):
        """
        Highlights glossary terms in target sheets without translating.
        Relies on the target text already being present in the Excel file.
        """
        yield {"type": "log", "message": "Starting Highlight Only Pipeline..."}
        
        try:
            wb = openpyxl.load_workbook(source_file_path, rich_text=True)
        except Exception as e:
            yield {"type": "error", "message": f"Excel Load Error: {str(e)}"}
            return

        if not source_sheet_name:
            source_sheet_name = "KR(한국)"
        if source_sheet_name not in wb.sheetnames:
            yield {"type": "error", "message": f"Source sheet '{source_sheet_name}' not found."}
            return

        if source_sheet_name in sheet_lang_map:
            source_lang = sheet_lang_map[source_sheet_name].get("lang", source_lang)

        # Load Glossary
        if glossary_file_path:
            src_lookup = source_lang
            if source_sheet_name in sheet_lang_map:
                s_info = sheet_lang_map[source_sheet_name]
                src_lookup = s_info.get('code', s_info.get('lang', source_lang))

            yield {"type": "log", "message": f"Loading glossary (Match Base: {src_lookup})..."}
            msg = await self.load_glossary_from_file(glossary_file_path, src_lookup)
            yield {"type": "log", "message": msg}
            
        if not self.glossary:
            yield {"type": "error", "message": "No glossary data mapped. Skipping highlight."}
            return

        source_ws = wb[source_sheet_name]
        source_data = []
        extracted_coords = set()
        range_parts = [r.strip() for r in cell_range.split(',') if r.strip()]
        
        for current_range in range_parts:
            try:
                rows = source_ws[current_range]
                if not isinstance(rows, tuple) and not isinstance(rows, list):
                    rows = ((rows,),)
                for row in rows:
                    for cell in row:
                        if cell.coordinate in extracted_coords:
                            continue
                        extracted_coords.add(cell.coordinate)
                        val = str(cell.value).strip() if cell.value is not None else ""
                        if val and val.lower() != "x":
                            row_idx = cell.row
                            row_key = str(source_ws[f"B{row_idx}"].value).strip() if source_ws[f"B{row_idx}"].value else ""
                            source_data.append({'text': val, 'coord': cell.coordinate, 'row_key': row_key})
            except Exception as e:
                yield {"type": "error", "message": f"Error accessing range {current_range}: {e}"}
                continue

        if not source_data:
            yield {"type": "error", "message": "No source text found in range."}
            return

        available_sheets = wb.sheetnames
        if selected_sheets:
            target_sheets = [s for s in selected_sheets if s in available_sheets and s != source_sheet_name]
        else:
            target_sheets = [s for s in available_sheets if s in sheet_lang_map and s != source_sheet_name]

        total_cells = len(source_data) * len(target_sheets)
        yield {"type": "log", "message": f"Plan: {len(target_sheets)} sheets, Total {total_cells} cells."}
        yield {"type": "progress", "current": 0, "total": total_cells, "percent": 0}

        completed_count = 0
        highlight_stats = {} # Count of highlights per sheet

        import asyncio

        async def highlight_cell_worker(ws, index, item, target_lang, target_lang_code):
            nonlocal completed_count
            source_text = item['text']
            coord = item['coord']
            
            target_cell = ws[coord]
            target_text = str(target_cell.value).strip() if target_cell.value is not None else ""
            logs = []
            highlight_count = 0
            bracket_issues = []

            # If target has text, calculate highlights based on source
            if target_text and target_text.lower() != "x":
                # 1. Bracket Check
                bracket_issues = self._check_glossary_brackets(source_text, target_text, target_lang_code, target_lang)
                if bracket_issues:
                    logs.extend(bracket_issues)
                
                # 2. Mismatch Check
                mismatch_issues = self._precheck_glossary_mismatch(source_text, target_text, target_lang_code)
                if mismatch_issues:
                    logs.extend(mismatch_issues)
                
                # 3. Casing Check
                casing_issues = self._check_glossary_casing(source_text, target_text, target_lang_code)
                if casing_issues:
                    logs.extend(casing_issues)

                original_target_terms = []
                relevant_terms = self._get_relevant_glossary_terms(source_text)
                if relevant_terms:
                    for s_term in relevant_terms:
                        meta = self.glossary.get(s_term)
                        if meta:
                            # Skip deactivated terms
                            rule = meta.get("rule", "").lower()
                            clean_rule = rule.replace(" ", "")
                            if "비활성화" in clean_rule or "deactivate" in clean_rule or "disable" in clean_rule:
                                continue

                            t_meta = meta["targets"]
                            target_val = t_meta.get(target_lang_code)
                            if not target_val:
                                for k, v in t_meta.items():
                                    if k.lower() == target_lang_code.lower():
                                        target_val = v
                                        break
                            
                            if target_val:
                                clean_val = re.sub(r'\(.*?\)', '', target_val).strip()
                                for extract_str in (clean_val, target_val):
                                    if not extract_str: continue
                                    original_target_terms.append(extract_str.strip())
                                    split_terms = [x.strip() for x in re.split(r'[,/]', extract_str) if x.strip()]
                                    if len(split_terms) > 1:
                                        original_target_terms.extend(split_terms)
                
                # Apply rich text
                if original_target_terms:
                    # We inject the current target cell text instead of a generated translation
                    target_cell.value = self._apply_rich_text(target_text, original_target_terms, base_font=target_cell.font)
                    
                    # Estimate highlights by checking string occurrences
                    sorted_kw = sorted([k.strip() for k in original_target_terms if k.strip()], key=len, reverse=True)
                    if sorted_kw:
                        pattern = '|'.join(re.escape(k) for k in sorted_kw)
                        matches = list(re.finditer(pattern, target_text, flags=re.IGNORECASE))
                        highlight_count = len(matches)
                        if matches:
                            logs.append(f"[{ws.title}] 셀 {coord}: {highlight_count}개 키워드 하이라이트 적용")

            res = {
                "sheet_name": ws.title,
                "cell_ref": coord,
                "highlight_count": highlight_count,
                "logs": logs
            }
            return index, res

        tasks = []
        idx = 0
        for sheet_name in target_sheets:
            ws = wb[sheet_name]
            sheet_info = sheet_lang_map[sheet_name]
            target_lang = sheet_info['lang']
            target_lang_code = sheet_info['code']
            highlight_stats[sheet_name] = {"cells_processed": 0, "total_highlights": 0}
            
            for item in source_data:
                tasks.append(highlight_cell_worker(ws, idx, item, target_lang, target_lang_code))
                idx += 1

        ordered_results = [None] * len(tasks)

        for future in asyncio.as_completed(tasks):
            try:
                index, res = await future
                completed_count += 1
                ordered_results[index] = res # CRITICAL FIX: Populate result list
                
                s_name = res["sheet_name"]
                highlight_stats[s_name]["cells_processed"] += 1
                highlight_stats[s_name]["total_highlights"] += res["highlight_count"]

                for msg in res.get("logs", []):
                    yield {"type": "log", "message": msg}

                percent = int((completed_count / total_cells) * 100)
                yield {
                    "type": "progress", 
                    "current": completed_count, 
                    "total": total_cells, 
                    "percent": percent, 
                    "log": f"[{res['sheet_name']}] {res['cell_ref']} 검토 완료"
                }
            except Exception as e:
                yield {"type": "log", "message": f"Cell processing error: {str(e)}"}

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_excel_path = source_file_path.replace(".xlsx", f"_highlighted_{timestamp}.xlsx")
        wb.save(out_excel_path)
        
        # Build Report Text
        header = (
            f"--- 하이라이트 전용 검수 보고서 ---\n"
            f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"총 대상 셀 수: {total_cells}개\n\n"
        )
        
        # Highlight Only 모드에서도 괄호 오류 등을 기록하기 위해 ordered_results 활용
        report_lines = []
        for s_name, stats in highlight_stats.items():
            report_lines.append(f"시트: {s_name} | 처리 완료 셀: {stats['cells_processed']} | 총 적용된 하이라이트: {stats['total_highlights']} 개")
        
        detail_lines = []
        for res in ordered_results:
            if res and res.get("logs"):
                # logs에 [괄호 오류], [미적용], [대소문자] 등이 있으면 상세 기록에 추가
                issue_logs = [log for log in res["logs"] if any(k in log for k in ["[오류", "[미적용", "[대소문자"])]
                if issue_logs:
                    detail_lines.append(f"\n[{res['sheet_name']} 시트 | {res['cell_ref']} 셀]")
                    for log in issue_logs:
                        detail_lines.append(f"  - {log}")
        
        report_text = header + "\n".join(report_lines)
        if detail_lines:
            report_text += "\n\n[주요 용어집 준수여부 점검 알림]" + "\n".join(detail_lines)
        
        yield {
            "type": "complete", 
            "output_data": report_text, 
            "excel_path": out_excel_path
        }
