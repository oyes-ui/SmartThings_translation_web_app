# -*- coding: utf-8 -*-
"""
backend/checker_service.py
Refactored for Web API usage (FastAPI + SSE)
"""

import openpyxl
import os
import google.generativeai as genai
from datetime import datetime
from dotenv import load_dotenv
import asyncio
import json
import re
import pandas as pd
import io
import urllib.parse

# API Key from Environment
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")

# ----------------- Case Hard-rule Applicable Languages -----------------
CASE_APPLICABLE_LANG_PREFIXES = {
    "English", "German", "French", "Spanish", "Portuguese",
    "Italian", "Dutch", "Swedish", "Polish", "Turkish",
    "Indonesian", "Vietnamese", "Russian"
}

def _is_case_sensitive_language(lang_name: str) -> bool:
    if not lang_name:
        return False
    return any(lang_name.startswith(pref) for pref in CASE_APPLICABLE_LANG_PREFIXES)

class TranslationChecker:
    """
    Excel Translation Checker (Gemini Single Model) - Service Version
    """

    def __init__(
        self,
        model_name: str = "gemini-2.0-flash", # Updated to 2.0-flash as per user mention
        max_concurrency: int = 10,
        short_text_whitelist=None,
        skip_llm_when_glossary_mismatch: bool = False,
        no_backtranslation: bool = False
    ):
        if not GEMINI_API_KEY:
            raise ValueError("API Key is missing.")

        # API Setup
        genai.configure(api_key=GEMINI_API_KEY)
        self.model_name = model_name
        self.qa_model = genai.GenerativeModel(model_name)
        
        self.no_backtranslation = no_backtranslation
        
        # Concurrency
        if max_concurrency < 1:
            max_concurrency = 1
        self.max_concurrency = max_concurrency
        self._sem = asyncio.Semaphore(self.max_concurrency)

        # Whitelist
        default_whitelist = {"ok", "on", "off", "ai", "5g", "go", "up", "usb", "nfc"}
        if short_text_whitelist:
            if isinstance(short_text_whitelist, str):
                extra = {x.strip().lower() for x in short_text_whitelist.split(",") if x.strip()}
            else:
                extra = {str(x).strip().lower() for x in short_text_whitelist}
            default_whitelist |= extra
        self.short_text_whitelist = default_whitelist

        self.skip_llm_when_glossary_mismatch = skip_llm_when_glossary_mismatch

        # Glossary Structure: { source_term: { 'targets': {code: term, ...}, 'rule': ... } }
        self.glossary = {}
        self.glossary_headers = []
        self.source_lang_code = None
    
    # ----------------- Glossary Loader (Google Sheets via Pandas) -----------------
    async def load_glossary_from_url(self, url: str, source_lang_code: str):
        """
        Loads glossary from Google Sheets export URL using pandas.
        Auto-converts 'edit' URLs to 'export?format=csv'.
        """
        try:
            # Convert Google Sheets URL if needed
            if "docs.google.com/spreadsheets" in url and "/edit" in url:
                # Extract Doc ID
                match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
                if match:
                    doc_id = match.group(1)
                    # Extract GID if present
                    gid_match = re.search(r"[#&]gid=([0-9]+)", url)
                    gid_param = f"&gid={gid_match.group(1)}" if gid_match else ""
                    
                    url = f"https://docs.google.com/spreadsheets/d/{doc_id}/export?format=csv{gid_param}"
                    print(f"Converted Google Sheets URL to: {url}")

            # Safe encode URL (handles Korean/Unicode in query params)
            # safe characters are those allowed in URL structure (: / ? = &)
            encoded_url = urllib.parse.quote(url, safe=':/?=&')
            
            # Using pandas to read CSV directly from URL
            df = pd.read_csv(encoded_url)
            
            # Basic cleanup: header guard (if first row is duplicate, though pandas usually handles headers well)
            # If the first row values equal columns, drop it.
            if df.shape[0] > 0:
                first_row_vals = [str(x) for x in df.iloc[0].values]
                col_names = [str(x) for x in df.columns]
                # specific check might be needed if pure duplicate
                pass 

            self.glossary_headers = list(df.columns)
            self.source_lang_code = source_lang_code
            
            # Identify 'explanation/rule' column
            rule_header = None
            if "설명/규칙" in df.columns:
                rule_header = "설명/규칙"
            elif "규칙" in df.columns:
                rule_header = "규칙"

            if source_lang_code not in df.columns:
                return f"Warning: Source column '{source_lang_code}' not found in glossary. Available: {list(df.columns)}"

            count = 0
            for _, row in df.iterrows():
                # Source Term
                source_term = str(row[source_lang_code]).strip() if pd.notna(row[source_lang_code]) else ""
                if not source_term:
                    continue

                # Rule
                rule = str(row[rule_header]).strip() if rule_header and pd.notna(row[rule_header]) else ""

                # Targets
                targets = {}
                for col in df.columns:
                    if col == source_lang_code:
                        continue
                    if rule_header and col == rule_header:
                        continue
                    
                    val = str(row[col]).strip() if pd.notna(row[col]) else ""
                    if val:
                        targets[col] = val
                
                if targets:
                    self.glossary[source_term] = {"targets": targets, "rule": rule}
                    count += 1
            
            return f"Glossary loaded: {count} entries."

        except Exception as e:
            return f"Glossary load failed: {str(e)}"

    async def load_glossary_from_file(self, file_path: str, source_lang_code: str):
        """
        사용자 JSON의 'code' 필드(한글 명칭)를 기준으로 용어집 1행(Index 0)을 매칭합니다.
        """
        try:
            if not os.path.exists(file_path):
                return f"Glossary file not found: {file_path}"

            # [핵심] 1행(Index 0)을 헤더로 지정 (한국어, 영어_미국 등)
            import pandas as pd
            df = pd.read_csv(file_path, header=0, encoding='utf-8-sig')
            df = df.fillna("")
            
            # 컬럼명 정리
            df.columns = [str(c).strip() for c in df.columns]
            
            # JSON에서 넘어온 'code' 값이 용어집 헤더에 있는지 확인
            # 예: source_lang_code가 "영어_미국"인 경우
            if source_lang_code not in df.columns:
                return f"⚠ Warning: 용어집 1행에서 '{source_lang_code}' 컬럼을 찾을 수 없습니다. (제공된 JSON 'code' 값 확인 필요)"

            self.glossary_headers = list(df.columns)
            self.source_lang_code = source_lang_code
            
            # 설명/규칙 컬럼 찾기
            rule_header = None
            for col in df.columns:
                if "설명" in col or "규칙" in col:
                    rule_header = col
                    break

            count = 0
            # 실제 데이터는 Lng 행(Index 2)을 지나 4행(Index 3)부터 시작하므로
            # 데이터프레임 기준으로는 Index 2부터 읽습니다.
            for i in range(2, len(df)):
                row = df.iloc[i]
                source_term = str(row[source_lang_code]).strip()
                if not source_term or source_term.lower() == 'lng': # 'Lng' 행 방어 로직
                    continue

                rule = str(row[rule_header]).strip() if rule_header and pd.notna(row[rule_header]) else ""

                targets = {}
                for col in df.columns:
                    # 원문 컬럼, 규칙 컬럼, key 컬럼 제외하고 모두 타겟으로 저장
                    if col in [source_lang_code, rule_header, 'key'] or not col.strip():
                        continue
                    
                    val = str(row[col]).strip() if pd.notna(row[col]) else ""
                    if val:
                        targets[col] = val
                
                if targets:
                    self.glossary[source_term] = {"targets": targets, "rule": rule}
                    count += 1
            
            return f"✓ 용어집 로드 성공: {count}개 항목 (매칭 기준: {source_lang_code})"

        except Exception as e:
            return f"Glossary load failed: {str(e)}"

    # ----------------- Excel Loader -----------------
    def load_excel_files(self, source_path, target_path, selected_sheets=None):
        pass

    def load_excel_data(self, source_path, target_path, cell_range="A:Z", selected_sheets=None):
        all_data = []
        try:
            wb_source = openpyxl.load_workbook(source_path, data_only=True)
            wb_target = openpyxl.load_workbook(target_path, data_only=True)
        except Exception as e:
            raise Exception(f"Excel load error: {e}")

        source_order = wb_source.sheetnames
        target_set = set(wb_target.sheetnames)
        common_sheets = [s for s in source_order if s in target_set]

        if not common_sheets:
            raise Exception("No matching sheets found between Source and Target files.")

        if selected_sheets:
            target_sheets = [s for s in selected_sheets if s in common_sheets]
            if not target_sheets:
                raise Exception(f"Selected sheets {selected_sheets} not found in common sheets.")
        else:
            target_sheets = common_sheets

        processed_sheets = []

        for sheet_name in target_sheets:
            ws_source = wb_source[sheet_name]
            ws_target = wb_target[sheet_name]
            
            extracted_count = 0
            try:
                # Handle full range like "A:Z" or specific "A1:C10"
                src_rows = ws_source[cell_range]
                tgt_rows = ws_target[cell_range]
                
                # Normalize to list of rows if single cell
                if not isinstance(src_rows, tuple) and not isinstance(src_rows, list): 
                    # it might be a single cell if range is "A1"
                    src_rows = ((src_rows,),)
                    tgt_rows = ((tgt_rows,),)
                elif isinstance(src_rows, tuple) and not isinstance(src_rows[0], tuple): 
                     # Single column/row tuple structure variation check
                     pass

                for source_row, target_row in zip(src_rows, tgt_rows):
                    for s_cell, t_cell in zip(source_row, target_row):
                        s_val = str(s_cell.value).strip() if s_cell.value is not None else ""
                        t_val = str(t_cell.value).strip() if t_cell.value is not None else ""
                        
                        if s_val and t_val:
                            all_data.append({
                                "cell_ref": s_cell.coordinate,
                                "sheet_name": sheet_name,
                                "source": s_val,
                                "target": t_val,
                            })
                            extracted_count += 1
            except Exception as e:
                print(f"Error accessing range {cell_range} in {sheet_name}: {e}")
                continue
            
            if extracted_count > 0:
                processed_sheets.append(sheet_name)
        
        return all_data, processed_sheets

    # ----------------- Helpers -----------------
    async def _with_semaphore(self, coro):
        async with self._sem:
            return await coro

    def _build_glossary_lines_for_code(self, target_lang_code: str):
        if not self.glossary:
            return "용어집 없음"
        out = []
        for source_term, meta in self.glossary.items():
            tgt = meta["targets"].get(target_lang_code)
            if not tgt:
                continue
            rule = meta.get("rule")
            rule_info = f" (규칙: {rule})" if rule else ""
            out.append(f"- 원어: {source_term} → 대상어({target_lang_code}): {tgt}{rule_info}")
        return "\n".join(out) if out else f"용어집에 '{target_lang_code}' 타겟 항목 없음"

    def _precheck_glossary_mismatch(self, source_text: str, target_text: str, target_lang_code: str):
        if not self.glossary or not target_lang_code:
            return []
        mismatches = []
        src_lower = source_text.lower()
        tgt_lower = target_text.lower()
        for s_term, meta in self.glossary.items():
            t_term = meta["targets"].get(target_lang_code)
            if not t_term: continue
            if s_term and s_term.lower() in src_lower:
                if t_term.lower() not in tgt_lower:
                    mismatches.append(f"'{s_term}' → '{t_term}' 미적용({target_lang_code})")
        return mismatches

    def _check_glossary_casing(self, source_text: str, target_text: str, target_lang_code: str):
        if not self.glossary or not target_lang_code or not source_text or not target_text:
            return []
        issues = []
        src_lower = source_text.lower()
        tgt_lower = target_text.lower()
        for s_term, meta in self.glossary.items():
            t_term = meta["targets"].get(target_lang_code)
            if not t_term: continue
            if s_term.lower() in src_lower and t_term.lower() in tgt_lower:
                idx = tgt_lower.find(t_term.lower())
                if idx == -1: continue
                actual = target_text[idx:idx + len(t_term)]
                if actual.lower() == t_term.lower() and actual != t_term:
                    issues.append(f"용어집 '{t_term}'의 대소문자 표기가 '{actual}'로 사용됨")
        return issues

    def _analyze_sentence_case(self, target_text: str, target_lang: str):
        if not target_text or not _is_case_sensitive_language(target_lang):
            return None, None
        
        text = target_text.strip()
        sentences = re.split(r'(?<=[\.!?])\s+', text)
        sentences = [s for s in sentences if s.strip()]
        if not sentences:
            return None, None
            
        report_lines = []
        fixed_sentences = []
        
        for idx, sent in enumerate(sentences, start=1):
            s = sent
            first_alpha_index = None
            for i, ch in enumerate(s):
                if ch.isalpha():
                    first_alpha_index = i
                    break
            
            if first_alpha_index is None:
                fixed_sentences.append(s)
                continue
                
            first_char = s[first_alpha_index]
            rest = s[first_alpha_index+1:]
            
            is_sentence_case = first_char.isupper() and rest == rest.lower()
            extra_caps = sum(1 for ch in rest if ch.isalpha() and ch.isupper())
            
            status = "문장형(첫 글자만 대문자)" if is_sentence_case else "문장형 아님"
            report_lines.append(f"- 문장 {idx}: {status}, 추가 대문자 수: {extra_caps}개")
            
            fixed_rest = rest.lower()
            fixed_sent = s[:first_alpha_index] + first_char.upper() + fixed_rest
            fixed_sentences.append(fixed_sent)
            
        simple_fixed_text = " ".join(fixed_sentences)
        report = "\n".join(report_lines)
        if simple_fixed_text == target_text:
            simple_fixed_text = None
        return report, simple_fixed_text

    # ----------------- LLM Calls -----------------
    async def check_with_gemini_qa(self, source_text, target_text, source_lang, target_lang, target_lang_code):
        glossary_text = self._build_glossary_lines_for_code(target_lang_code)
        
        prompt = f"""당신은 전문 번역 검수 전문가입니다.

[Context]
- 원문({source_lang}): {source_text}
- 번역문({target_lang}/{target_lang_code}): {target_text}

[용어집({target_lang_code})]
{glossary_text}

다음 항목에 대해 상세하고 객관적인 검수 결과를 **한국어 불렛 포인트**로 정리해 주세요:
1. **문법/유창성**: 번역문의 어색한 표현이나 문법 오류.
2. **문화/문맥 적절성**: 뉘앙스 손실, 문화적으로 부적절한 요소. **(참고: 'German for Casual Audience (Du-form mandatory)' 등 언어 지정이 있는 경우, 호칭(Du/Sie) 일관성 확인)**
3. **대소문자(Casing) 및 문장형**:
   - 각 문장에서 첫 단어만 대문자이고 나머지는 소문자인지(문장형) 평가하세요.
   - 문장 중간에 등장하는 대문자·ALL CAPS 단어가 고유명/브랜드명/기능명/약어인지 여부를 설명하세요.
4. **용어집 및 규칙 준수**:
   - 용어집에 있는 용어가 올바르게 번역되었는지뿐만 아니라,
   - 용어집에 명시된 대소문자 표기(SmartThings, Galaxy Watch 등)가 그대로 지켜졌는지도 평가하세요.
5. **수정 제안**:
   - 문제가 있다고 판단되는 부분이 있을 경우,
   - 문장형(첫 글자만 대문자) 기준을 해치지 않으면서
   - 고유명/기능명/약어/용어집 표기를 그대로 유지하는 **안전한 수정안**을 제시하세요.

문제가 없다면 마지막 줄에:
"최종 평가: 우수, 주요 문제 없음."
"""
        try:
            response = await self.qa_model.generate_content_async(
                prompt,
                generation_config={"temperature": 0.2},
                request_options={"timeout": 90}
            )
            return response.text.strip() if response.text else "[응답 비어있음]"
        except Exception as e:
            return f"[Gemini QA 오류]: {e}"

    async def get_back_translation(self, target_text, target_lang, source_lang):
        prompt = (
            f"다음 {target_lang} 텍스트를 {source_lang}으로 다시 번역해주세요. "
            f"오직 번역된 텍스트만 제공해야 합니다. 다른 설명이나 텍스트는 포함하지 마세요.\n\n{target_text}"
        )
        try:
            response = await self.qa_model.generate_content_async(
                prompt,
                generation_config={"temperature": 0.1},
                request_options={"timeout": 90}
            )
            return response.text.strip() if response.text else "[응답 비어있음]"
        except Exception as e:
            return f"[Gemini 역번역 오류]: {e}"

    # ----------------- Process Single Item -----------------
    async def process_item(self, item, source_lang, default_target_lang, sheet_lang_map, default_target_lang_code):
        cell_ref = item["cell_ref"]
        sheet_name = item["sheet_name"]
        source = item["source"]
        target = item["target"]
        
        tgt_lang = sheet_lang_map.get(sheet_name, {}).get("lang", default_target_lang)
        tgt_code = sheet_lang_map.get(sheet_name, {}).get("code", default_target_lang_code)
        
        # Skip detection
        is_placeholder = (
            (len(source) <= 2 and len(target) <= 2 and source.lower() == target.lower()) or
            (not source.strip() and not target.strip()) or
            (source.strip().lower() == target.strip().lower() and len(source.strip()) < 10)
        )
        if source.strip().lower() in self.short_text_whitelist or target.strip().lower() in self.short_text_whitelist:
            is_placeholder = False
        
        pre_mismatch = self._precheck_glossary_mismatch(source, target, tgt_code)
        skip_llm = self.skip_llm_when_glossary_mismatch and bool(pre_mismatch)
        
        case_report, simple_case_fix = self._analyze_sentence_case(target, tgt_lang)
        glossary_case_issues = self._check_glossary_casing(source, target, tgt_code)
        
        # Construct Partial Report Sections
        case_section = "대소문자 하드룰(문장형) 점검:\n" + case_report if case_report else "별도 지적 사항 없음."
        if simple_case_fix:
            case_section += f"\n\n[단순 규칙 기반 문장형 변환안]:\n{simple_case_fix}"

        glossary_parts = []
        if pre_mismatch: glossary_parts.append("용어집 사전 감지:\n- " + "\n- ".join(pre_mismatch))
        if glossary_case_issues: glossary_parts.append("용어집 대소문자 표기 점검:\n" + "\n".join(f"- {msg}" for msg in glossary_case_issues))
        glossary_section = "\n\n".join(glossary_parts) if glossary_parts else "별도 지적 사항 없음."

        # Result container
        res = {
            "sheet_name": sheet_name,
            "cell_ref": cell_ref,
            "source": source,
            "target": target,
            "case_section": case_section,
            "glossary_section": glossary_section,
            "back_translation": "",
            "gemini_review": ""
        }

        # Skip Logic
        if is_placeholder and not pre_mismatch:
            res["back_translation"] = "[건너뜀: 짧은/무의미]"
            res["gemini_review"] = "[건너뜀: 짧은/무의미]"
            return res

        # LLM Logic
        if skip_llm:
            res["back_translation"] = "[사전 감지로 LLM 호출 생략]"
            res["gemini_review"] = "※ 용어집 사전 감지 결과를 우선 검토하세요."
        else:
             # LLM Calls
            qa_task = self._with_semaphore(
                self.check_with_gemini_qa(source, target, source_lang, tgt_lang, tgt_code)
            )
            
            if self.no_backtranslation:
                res["gemini_review"] = await qa_task
                res["back_translation"] = "[역번역 비활성화됨]"
            else:
                bt_task = self._with_semaphore(
                    self.get_back_translation(target, tgt_lang, source_lang)
                )
                review, bt = await asyncio.gather(qa_task, bt_task)
                res["gemini_review"] = review
                res["back_translation"] = bt
        
        return res

    # ----------------- Generator Method for SSE -----------------
    async def run_inspection_async_generator(
        self,
        source_file_path,
        target_file_path,
        cell_range,
        source_lang,
        target_lang,
        target_lang_code,
        sheet_lang_map,
        glossary_url=None,
        selected_sheets=None,
        glossary_file_path=None
    ):
        """
        Yields events:
        {"type": "log", "message": "..."}
        {"type": "progress", "current": n, "total": m, "percent": ...}
        {"type": "result_chunk", "data": formatted_string}
        {"type": "complete", "total": m, "output_data": all_text}
        """
        yield {"type": "log", "message": "용어집 로드 시작..."}
        
        # Load Glossary
        if glossary_file_path:
             yield {"type": "log", "message": f"용어집 파일 로드 중: {os.path.basename(glossary_file_path)}"}
             msg = await self.load_glossary_from_file(glossary_file_path, source_lang)
             yield {"type": "log", "message": msg}
        elif glossary_url:
            yield {"type": "log", "message": "Google Sheets 용어집 URL 로드 중..."}
            msg = await self.load_glossary_from_url(glossary_url, source_lang)
            yield {"type": "log", "message": msg}
        
        yield {"type": "log", "message": "엑셀 파일 및 시트 분석 중..."}
        try:
            # Use explicit selected_sheets if provided, otherwise fallback to map keys
            sel_sheets = selected_sheets if selected_sheets else (list(sheet_lang_map.keys()) if sheet_lang_map else None)
            
            all_data, processed_sheets = self.load_excel_data(
                source_file_path, target_file_path, cell_range=cell_range, selected_sheets=sel_sheets
            )
        except Exception as e:
            yield {"type": "error", "message": str(e)}
            return

        total_items = len(all_data)
        yield {"type": "log", "message": f"검수 대상: {len(processed_sheets)}개 시트, 총 {total_items}개 항목"}
        yield {"type": "progress", "current": 0, "total": total_items, "percent": 0}

        if total_items == 0:
            yield {"type": "complete", "total": 0, "output_data": "검수할 데이터가 없습니다."}
            return

        # Wrapper to track index
        async def process_with_index(index, item):
            res = await self.process_item(item, source_lang, target_lang, sheet_lang_map, target_lang_code)
            return index, res

        # Create tasks with index
        tasks = [
            process_with_index(i, item)
            for i, item in enumerate(all_data)
        ]

        completed_count = 0
        # Initialize list to store results in order
        ordered_results = [None] * total_items
        
        # Use asyncio.as_completed to yield progress, but store by index
        for future in asyncio.as_completed(tasks):
            try:
                index, res = await future
                completed_count += 1
                
                # Format result
                fmt_result = (
                    f"\n\n{'='*90}\n"
                    f"[시트] {res['sheet_name']} | [셀] {res['cell_ref']}\n"
                    f"{'-'*90}\n\n"
                    f"[상세 - 원문]\n{res['source']}\n\n"
                    f"[상세 - 번역문]\n{res['target']}\n\n"
                    f"[상세 - 대소문자 점검]\n{res['case_section']}\n\n"
                    f"[상세 - 용어집 점검]\n{res['glossary_section']}\n\n"
                    f"[상세 - 역번역]\n{res['back_translation']}\n\n"
                    f"[상세 - Gemini 검수 결과]\n{res['gemini_review']}\n"
                    f"{'='*90}\n"
                )
                
                # Store in the correct slot
                ordered_results[index] = fmt_result
                
                percent = int((completed_count / total_items) * 100)
                yield {
                    "type": "progress",
                    "current": completed_count,
                    "total": total_items,
                    "percent": percent,
                    "log": f"[{res['sheet_name']}] {res['cell_ref']} 검수 완료"
                }
            except Exception as e:
                yield {"type": "log", "message": f"항목 처리 중 오류 발생: {str(e)}"}
                continue

        # Final Header/Footer assembly using ORDERED results
        final_text = []
        header = (
            f"--- 번역 검수 보고서 (Gemini 2.0-flash / Web App) ---\n"
            f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"총 검수 항목: {total_items}개\n"
            f"용어집 항목: {len(self.glossary)}\n\n"
        )
        final_text.append(header)
        # Filter out None results if any failed
        valid_results = [r for r in ordered_results if r is not None]
        final_text.extend(valid_results)
        
        full_output = "".join(final_text)
        yield {"type": "complete", "total": total_items, "output_data": full_output}

