import openpyxl
import csv
import re
import sys
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont


# 1️⃣ CSV에서 단어 리스트 불러오기
def load_keywords_from_csv(filename, key_column):
    keywords = []
    with open(filename, newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            keyword = row.get(key_column)
            if keyword:
                keywords.append(keyword.strip())
    return sorted(keywords, key=len, reverse=True)


# 2️⃣ 워크북 내 단어 하이라이트 처리
def highlight_keywords_in_workbook(input_file, output_file, keywords, cell_range=None, selected_sheets=None):
    from openpyxl import load_workbook
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont

    wb = load_workbook(input_file, rich_text=True)
    
    # 정규표현식 패턴 미리 생성
    if not keywords:
        print("⚠ 키워드가 없습니다.")
        return
    pattern = '|'.join(re.escape(k) for k in keywords)

    for ws in wb.worksheets:
        if selected_sheets and ws.title not in selected_sheets:
            continue
            
        print(f"🔍 워크시트 처리 중: {ws.title} (범위: {cell_range or '전체'})")
        
        # 특정 범위가 있으면 해당 범위만, 없으면 전체 셀
        if cell_range:
            rows = []
            range_parts = [r.strip() for r in cell_range.split(',')]
            for current_range in range_parts:
                part_rows = ws[current_range]
                # 단일 셀인 경우 처리
                if not isinstance(part_rows, (tuple, list)):
                    part_rows = ((part_rows,),)
                rows.extend(part_rows)
        else:
            rows = ws.iter_rows()

        processed_coords = set()

        for row in rows:
            for cell in row:
                if cell.coordinate in processed_coords:
                    continue
                processed_coords.add(cell.coordinate)

                text = str(cell.value) if cell.value is not None else ""
                if text:
                    matches = list(re.finditer(pattern, text, flags=re.IGNORECASE))
                    if not matches:
                        continue

                    # 기존 스타일 유지 로직 (InlineFont는 name→rFont 사용)
                    base_font = cell.font
                    font_params = {
                        "rFont": base_font.name,
                        "sz": base_font.sz,
                        "b": base_font.b,
                        "i": base_font.i,
                        "u": base_font.u,
                        "strike": base_font.strike,
                        "family": base_font.family,
                        "charset": base_font.charset,
                        "outline": base_font.outline,
                        "shadow": base_font.shadow,
                        "condense": base_font.condense,
                        "extend": base_font.extend,
                        "vertAlign": base_font.vertAlign,
                        "scheme": base_font.scheme,
                    }

                    def get_font(is_keyword=False):
                        params = font_params.copy()
                        if is_keyword:
                            params["color"] = "0000FF"
                        else:
                            # 기존 셀 색상 유지
                            if base_font and base_font.color and hasattr(base_font.color, 'rgb'):
                                params["color"] = base_font.color
                        return InlineFont(**params)

                    parts = []
                    last_end = 0
                    for m in matches:
                        start, end = m.span()
                        if start > last_end:
                            parts.append((text[last_end:start], get_font(False)))
                        parts.append((text[start:end], get_font(True)))
                        last_end = end
                    if last_end < len(text):
                        parts.append((text[last_end:], get_font(False)))

                    rt = CellRichText()
                    for segment_text, segment_font in parts:
                        rt.append(TextBlock(text=segment_text, font=segment_font))
                    cell.value = rt

    wb.save(output_file)
    print(f"✅ 완료! 결과 파일: {output_file}")


# 3️⃣ 실행부
if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Excel Glossary Keyword Highlighter")
    parser.add_argument("csv_file", help="용어집 CSV 파일")
    parser.add_argument("column_key", help="매칭 기준 컬럼명")
    parser.add_argument("input_xlsx", help="입력 엑셀 파일")
    parser.add_argument("output_xlsx", help="출력 엑셀 파일")
    parser.add_argument("--range", help="범위 지정 (예: C7:C28)", default=None)
    parser.add_argument("--sheets", help="시트 지정 (콤마 구분)", default=None)

    args = parser.parse_args()

    print(f"📘 CSV 파일: {args.csv_file}")
    print(f"🔑 번역 키: {args.column_key}")
    print(f"📂 입력 파일: {args.input_xlsx}")
    
    selected_sheets = None
    if args.sheets:
        selected_sheets = [s.strip() for s in args.sheets.split(",")]

    keywords = load_keywords_from_csv(args.csv_file, args.column_key)
    highlight_keywords_in_workbook(args.input_xlsx, args.output_xlsx, keywords, cell_range=args.range, selected_sheets=selected_sheets)