from dotenv import load_dotenv
# Load environment variables FIRST before other imports
load_dotenv()

from fastapi import FastAPI, UploadFile, File, BackgroundTasks, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
import shutil
import os
import uuid
import asyncio
import json
import openpyxl
from datetime import datetime
from contextlib import asynccontextmanager
from checker_service import TranslationChecker
import zipfile

# RAG 모듈 임포트 (DB 없으면 graceful fallback)
try:
    import rag_db_builder as _rag_builder
    _rag_builder_available = True
except ImportError:
    _rag_builder_available = False

# Directory for temp files
UPLOAD_DIR = "uploads"
if not os.path.exists(UPLOAD_DIR):
    os.makedirs(UPLOAD_DIR)

app = FastAPI()

# 1. 현재 main.py 파일이 있는 실제 경로를 계산합니다.
current_dir = os.path.dirname(os.path.abspath(__file__))
static_dir = os.path.join(current_dir, "static")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)



# Store for active tasks: { task_id: { "queue": asyncio.Queue, "result_path": str, "status": ... } }
TASK_STORE = {}

class StartRequest(BaseModel):
    source_file_id: str
    source_file_name: str = None
    target_file_id: str = None
    glossary_file_id: str = None
    source_sheet: str = None # Added for explicit source sheet selection
    sheets: list[str] = None # Used as target sheets in integrated mode
    sheet_langs: dict = {} # {"Sheet1": {"lang": "Korean", "code": "ko_KR"}}
    source_lang: str = "English"
    target_lang: str = "Korean"
    target_code: str = "ko_KR"
    max_concurrency: int = 5
    cell_range: str = "C7:C28" # Default range
    model_name: str
    translation_model: str = "gemini-2.5-flash"
    audit_model: str = "gpt-5-mini"
    bx_style_enabled: bool = False
    task_mode: str = "integrated"
    rag_identity_match: bool = True

async def background_inspection_task(task_id, params):
    queue = TASK_STORE[task_id]["queue"]
    try:
        checker = TranslationChecker(
            model_name=params.model_name,
            max_concurrency=params.max_concurrency,
            skip_llm_when_glossary_mismatch=False,
            no_backtranslation=True # Disabled as per user request
        )
        
        source_path = os.path.join(UPLOAD_DIR, params.source_file_id)
        target_path = os.path.join(UPLOAD_DIR, params.target_file_id)
        
        glossary_path = None
        if params.glossary_file_id:
            glossary_path = os.path.join(UPLOAD_DIR, params.glossary_file_id)

        # Run generator
        gen = checker.run_inspection_async_generator(
            source_file_path=source_path,
            target_file_path=target_path,
            cell_range=params.cell_range,
            source_lang=params.source_lang,
            target_lang=params.target_lang,
            target_lang_code=params.target_code,
            sheet_lang_map=params.sheet_langs,
            glossary_file_path=glossary_path,
            selected_sheets=params.sheets, # These are the sheets to inspect
            source_sheet_name=params.source_sheet,
            rag_identity_match=params.rag_identity_match
        )
        
        async for event in gen:
            # If complete, save output
            if event["type"] == "complete":
                base_name = os.path.splitext(params.source_file_name)[0] if params.source_file_name else f"review_{task_id}"
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_filename = f"{base_name}_review_report_{timestamp}.txt"
                output_path = os.path.join(UPLOAD_DIR, output_filename)
                with open(output_path, "w", encoding="utf-8") as f:
                    f.write(event["output_data"])
                
                TASK_STORE[task_id]["result_path"] = output_path
                TASK_STORE[task_id]["txt_path"] = output_path
                # Notify completion without large data payload
                await queue.put({
                    "type": "complete", 
                    "download_url": f"/api/download/{task_id}",
                    "download_file_name": output_filename,
                    "is_zip": False
                })
            else:
                await queue.put(event)
                
    except Exception as e:
        print(f"Background Task Error: {e}")
        await queue.put({"type": "error", "message": f"검수 시작 중 오류 발생: {str(e)}"})
    finally:
        # cleanup files maybe? or keep them for now.
        pass

async def integrated_translation_task(task_id, params):
    queue = TASK_STORE[task_id]["queue"]
    try:
        checker = TranslationChecker(
            model_name=params.audit_model,
            max_concurrency=params.max_concurrency,
            no_backtranslation=True # Disabled as per user request
        )
        
        source_path = os.path.join(UPLOAD_DIR, params.source_file_id)
        
        glossary_path = None
        if params.glossary_file_id:
            glossary_path = os.path.join(UPLOAD_DIR, params.glossary_file_id)

        if params.task_mode == "highlight_only":
            await queue.put({"type": "log", "message": "Starting highlight only pipeline..."})
            gen = checker.run_highlight_only_pipeline_generator(
                source_file_path=source_path,
                cell_range=params.cell_range,
                sheet_lang_map=params.sheet_langs,
                glossary_file_path=glossary_path,
                selected_sheets=params.sheets,
                source_sheet_name=params.source_sheet,
                source_lang=params.source_lang
            )
        else:
            await queue.put({"type": "log", "message": "Starting integrated translation and audit pipeline..."})
            gen = checker.run_integrated_pipeline_generator(
                source_file_path=source_path,
                cell_range=params.cell_range,
                bx_style_on=params.bx_style_enabled,
                sheet_lang_map=params.sheet_langs,
                translation_model=params.translation_model,
                audit_model=params.audit_model,
                glossary_file_path=glossary_path,
                selected_sheets=params.sheets,
                source_sheet_name=params.source_sheet,
                skip_audit=(params.task_mode == "translate_only"),
                source_lang=params.source_lang,
                rag_identity_match=params.rag_identity_match
            )
        
        async for event in gen:
            if event["type"] == "complete":
                base_name = os.path.splitext(params.source_file_name)[0] if params.source_file_name else f"result_{task_id}"
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                
                output_filename = f"{base_name}_review_{timestamp}.txt"
                output_path = os.path.join(UPLOAD_DIR, output_filename)
                with open(output_path, "w", encoding="utf-8") as f:
                    f.write(event["output_data"])
                
                excel_out = event.get("excel_path")
                
                # Create ZIP
                zip_name = f"{base_name}_result_{timestamp}.zip"
                zip_path = os.path.join(UPLOAD_DIR, zip_name)
                
                with zipfile.ZipFile(zip_path, 'w') as zipf:
                    if output_path and os.path.exists(output_path): 
                        zipf.write(output_path, os.path.basename(output_path))
                    if excel_out and os.path.exists(excel_out): 
                        zipf.write(excel_out, os.path.basename(excel_out))
                
                TASK_STORE[task_id]["result_path"] = zip_path
                TASK_STORE[task_id]["txt_path"] = output_path
                await queue.put({"type": "log", "message": "Processing complete. Generating ZIP..."})
                await queue.put({
                    "type": "complete", 
                    "download_url": f"/api/download/{task_id}",
                    "result_file_id": os.path.basename(excel_out) if excel_out else None,
                    "download_file_name": zip_name,
                    "is_zip": True
                })
            else:
                await queue.put(event)
                
    except Exception as e:
        print(f"Integrated Task Error: {e}")
        await queue.put({"type": "error", "message": f"작업 중 오류 발생: {str(e)}"})

@app.post("/api/upload")
async def upload_files(
    source: UploadFile = File(None),
    target: UploadFile = File(None),
    glossary: UploadFile = File(None)
):
    result = {}
    
    try:
        # Handle Source Upload
        if source:
            s_id = f"src_{uuid.uuid4()}.xlsx"
            s_path = os.path.join(UPLOAD_DIR, s_id)
            content = await source.read()
            if not content:
                raise Exception("Uploaded source file is empty.")
            
            with open(s_path, "wb") as f:
                f.write(content)
                
            # Extract sheets from Source
            try:
                print(f"Opening workbook at {s_path}...")
                # data_only=True is usually what we want for checking values
                wb = openpyxl.load_workbook(s_path, read_only=True, data_only=True)
                sheets = wb.sheetnames
                print(f"Sheets found: {sheets}")
                wb.close()
            except Exception as e:
                print(f"Error loading workbook: {e}")
                # Re-try without read_only as some files might need it
                try:
                    wb = openpyxl.load_workbook(s_path, data_only=True)
                    sheets = wb.sheetnames
                    wb.close()
                except Exception as e2:
                    raise Exception(f"Failed to read Excel sheets: {str(e2)}")
                
            result["source_file_id"] = s_id
            result["source_file_name"] = source.filename
            result["sheets"] = sheets

        # Handle Target Upload
        if target:
            t_id = f"tgt_{uuid.uuid4()}.xlsx"
            t_path = os.path.join(UPLOAD_DIR, t_id)
            content = await target.read()
            if not content:
                 raise Exception("Uploaded target file is empty.")
            with open(t_path, "wb") as f:
                f.write(content)
            result["target_file_id"] = t_id

        # Handle Glossary Upload
        if glossary:
            g_id = f"glossary_{uuid.uuid4()}.csv"
            g_path = os.path.join(UPLOAD_DIR, g_id)
            content = await glossary.read()
            if not content:
                raise Exception("Uploaded glossary file is empty.")
            with open(g_path, "wb") as f:
                f.write(content)
            result["glossary_file_id"] = g_id

        if not result:
             raise HTTPException(status_code=400, detail="No files provided in the request.")
             
        return result

    except Exception as e:
        print(f"Upload error: {e}")
        raise HTTPException(status_code=400, detail=str(e))


class GlossaryCheckRequest(BaseModel):
    url: str
    source_lang: str = "English"

@app.post("/api/check_glossary")
async def check_glossary(req: GlossaryCheckRequest):
    # This endpoint is mostly for testing Google Sheets support which we removed
    # We'll just return an error or disable it.
    raise HTTPException(status_code=400, detail="Google Sheets glossary support has been removed. Please upload a CSV file.")

@app.post("/api/start")
async def start_inspection(req: StartRequest, background_tasks: BackgroundTasks):
    task_id = str(uuid.uuid4())
    TASK_STORE[task_id] = {
        "queue": asyncio.Queue(),
        "result_path": None
    }
    
    if req.task_mode == "inspect_only":
        background_tasks.add_task(background_inspection_task, task_id, req)
    else:
        # For 'integrated', 'translate_only', and 'highlight_only', we use integrated_translation_task
        background_tasks.add_task(integrated_translation_task, task_id, req)
    
    return {"task_id": task_id}

@app.get("/api/stream/{task_id}")
async def stream_progress(task_id: str):
    if task_id not in TASK_STORE:
        raise HTTPException(status_code=404, detail="Task not found")
        
    async def event_generator():
        queue = TASK_STORE[task_id]["queue"]
        while True:
            data = await queue.get()
            yield f"data: {json.dumps(data)}\n\n"
            if data["type"] in ["complete", "error"]:
                break
    
    return StreamingResponse(event_generator(), media_type="text/event-stream")

@app.get("/api/download/{task_id}/{filename:path}")
@app.get("/api/download/{task_id}")
async def download_result(task_id: str, filename: str = None):
    if task_id not in TASK_STORE or not TASK_STORE[task_id]["result_path"]:
        raise HTTPException(status_code=404, detail="Result not ready or task not found")
    
    path = TASK_STORE[task_id]["result_path"]
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Result file missing from server")
    
    filename = os.path.basename(path)
    return FileResponse(path, filename=filename)

@app.get("/api/report/{task_id}")
async def get_report_txt(task_id: str):
    if task_id not in TASK_STORE or not TASK_STORE[task_id].get("txt_path"):
        raise HTTPException(status_code=404, detail="Report not ready or task not found")
    
    path = TASK_STORE[task_id]["txt_path"]
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Report file missing from server")
    
    return FileResponse(path, media_type="text/plain")

# ─── RAG 엔드포인트 ──────────────────────────────────────────────────────────

@app.get("/api/rag_status")
async def rag_status():
    """RAG DB 현황 조회"""
    if not _rag_builder_available:
        return {"available": False, "message": "rag_db_builder 모듈 미설정"}
    try:
        conn = _rag_builder.get_sqlite_conn()
        _, col_kr, col_us = _rag_builder.get_chroma_collections()
        total = conn.execute("SELECT COUNT(*) FROM rag_pairs").fetchone()[0]
        files = conn.execute("SELECT COUNT(*) FROM processed_files").fetchone()[0]
        return {
            "available": True,
            "total_pairs": total,
            "processed_files": files,
            "kr_vectors": col_kr.count(),
            "us_vectors": col_us.count(),
        }
    except Exception as e:
        return {"available": False, "message": str(e)}


@app.get("/api/rag_browse")
async def rag_browse(
    story_id: str = None,
    target_lang: str = None,
    source_group: str = None,
    search: str = None,
    page: int = 1,
    page_size: int = 30
):
    """RAG DB 레코드 조회 (필터 + 페이지네이션)"""
    if not _rag_builder_available:
        raise HTTPException(status_code=500, detail="rag_db_builder 모듈 미설정")
    try:
        conn = _rag_builder.get_sqlite_conn()

        # 스토리/언어 목록
        stories = [r[0] for r in conn.execute(
            "SELECT DISTINCT story_id FROM rag_pairs ORDER BY story_id"
        ).fetchall()]
        langs = [r[0] for r in conn.execute(
            "SELECT DISTINCT target_lang FROM rag_pairs ORDER BY target_lang"
        ).fetchall()]

        # 필터 쿼리
        where_clauses, params = [], []
        if story_id:
            where_clauses.append("story_id = ?"); params.append(story_id)
        if target_lang:
            where_clauses.append("target_lang = ?"); params.append(target_lang)
        if source_group:
            where_clauses.append("source_group = ?"); params.append(source_group)
        if search:
            where_clauses.append("(source_text LIKE ? OR target_text LIKE ?)")
            params += [f"%{search}%", f"%{search}%"]

        where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""
        total_count = conn.execute(
            f"SELECT COUNT(*) FROM rag_pairs {where_sql}", params
        ).fetchone()[0]

        offset = (page - 1) * page_size
        rows = conn.execute(
            f"""SELECT story_id, section_code, source_group, source_text,
                       target_lang, target_text, original_file
                FROM rag_pairs {where_sql}
                ORDER BY story_id, target_lang, section_code
                LIMIT ? OFFSET ?""",
            params + [page_size, offset]
        ).fetchall()

        return {
            "stories": stories,
            "langs": langs,
            "total": total_count,
            "page": page,
            "page_size": page_size,
            "records": [dict(r) for r in rows]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/rag_similar")
async def rag_similar(query: str, target_lang: str = None, n: int = 5):
    """유사도 검색 테스트 (뷰어용)"""
    if not _rag_builder_available:
        raise HTTPException(status_code=500, detail="rag_db_builder 모듈 미설정")
    try:
        from rag_retriever import get_retriever
        retriever = get_retriever()
        if not retriever.is_available():
            return {"results": [], "message": "RAG DB가 비어있습니다"}
        
        # 쿼리 언어 자동 감지 (한글 포함 시 Korean 소스 DB 검색)
        import re
        source_lang = "English"
        if re.search(r'[가-힣]', query):
            source_lang = "Korean"
            
        results = retriever.retrieve(query, target_lang, source_lang=source_lang, n_results=n, exclude_same_source=False)
        return {"results": results}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))



from fastapi import Query

@app.post("/api/build_rag")
async def build_rag(background_tasks: BackgroundTasks, force: bool = Query(False)):
    """RAG DB 빌드 (SSE 스트리밍 진행률 제공)"""
    if not _rag_builder_available:
        raise HTTPException(status_code=500, detail="rag_db_builder 모듈 미설정")

    task_id = str(uuid.uuid4())
    TASK_STORE[task_id] = {"queue": asyncio.Queue(), "result_path": None}

    async def rag_build_task(task_id, force_val):
        queue = TASK_STORE[task_id]["queue"]
        loop = asyncio.get_event_loop()

        def log_fn(msg):
            loop.call_soon_threadsafe(
                queue.put_nowait, {"type": "log", "message": msg}
            )

        try:
            total = await _rag_builder.build_all_async(log_fn, force=force_val)
            await queue.put({"type": "complete", "total": total})
        except Exception as e:
            await queue.put({"type": "error", "message": str(e)})

    background_tasks.add_task(rag_build_task, task_id, force)
    return {"task_id": task_id}


@app.post("/api/update_rag_story")
async def update_rag_story(story_id: str, background_tasks: BackgroundTasks):
    """RAG DB 특정 스토리 증분 업데이트"""
    if not _rag_builder_available:
        raise HTTPException(status_code=500, detail="rag_db_builder 모듈 미설정")

    task_id = str(uuid.uuid4())
    TASK_STORE[task_id] = {"queue": asyncio.Queue(), "result_path": None}

    async def rag_update_task(task_id, story_id):
        queue = TASK_STORE[task_id]["queue"]
        loop = asyncio.get_event_loop()  # async 컨텍스트에서 미리 캡처

        def log_fn(msg):
            loop.call_soon_threadsafe(
                queue.put_nowait, {"type": "log", "message": msg}
            )

        try:
            total = await _rag_builder.update_story_async(story_id, log_fn)
            await queue.put({"type": "complete", "total": total})
        except Exception as e:
            await queue.put({"type": "error", "message": str(e)})

    background_tasks.add_task(rag_update_task, task_id, story_id)
    return {"task_id": task_id}


# 1. 'static' 폴더를 웹에 연결
app.mount("/", StaticFiles(directory=static_dir, html=True), name="static")

# 2. 접속 시 첫 화면(index.html) 보내주기
@app.get("/")
async def read_index():
    return FileResponse('static/index.html')

# 3. 브라우저 favicon.ico 404 에러 방지
@app.get("/favicon.ico", include_in_schema=False)
async def favicon():
    from fastapi.responses import Response
    return Response(status_code=204)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
