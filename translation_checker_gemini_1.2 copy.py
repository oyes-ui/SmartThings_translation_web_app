# -*- coding: utf-8 -*-
"""
translation_checker_gemini_1.0.py

# 업데이트 내역
# 0.2 용어집 인식 수정, GPT 결제문제로 gpt 제거
# 0.3 비동기 처리 최적화
# 0.4 비동기 처리시 순서가 뒤죽박죽으로 기입되던 것 수정
# 0.6 용어집 내 규칙 내용도 프롬포트에 반영하도록 수정
# 0.7 검수 프롬포트 대문자 검수 강화, API Timeout 90초 설정
# 0.8 시트 이름(sheet_names) 인수를 통한 선택적 검수 기능 추가 및 버그 수정 (FINAL)
# 0.8A 세마포어(동시성 제한), 짧은 텍스트 화이트리스트, 용어집 사전 불일치 감지 추가
# 0.9 시트별 언어/코드 매핑(--sheet_langs / --sheet_langs_file) + 용어집 다언어 컬럼 지원 + 디버그
# 1.0 대소문자(문장형) 하드룰 강화 + 용어집 케이스 검수 + LLM 프롬프트에 케이스/고유명/기능명 평가·수정안 명시

필수:
- pip install google-generativeai openpyxl python-dotenv

환경:
- .env 파일에 GEMINI_API_KEY=... 설정
"""

import openpyxl
import csv
import os
import argparse
import google.generativeai as genai
from datetime import datetime
from dotenv import load_dotenv
import asyncio
import json
import re

# 환경 변수 로드 (.env 파일에서 API 키를 읽어옵니다)
load_dotenv()

# ----------------- 대소문자 하드룰 적용 대상 언어 -----------------
# sheet_langs.json 에서 오는 lang 값 기준 (예: "English", "German", "French" ...)
CASE_APPLICABLE_LANG_PREFIXES = {
    "English", "German", "French", "Spanish", "Portuguese",
    "Italian", "Dutch", "Swedish", "Polish", "Turkish",
    "Indonesian", "Vietnamese", "Russian"
}


def _is_case_sensitive_language(lang_name: str) -> bool:
    if not lang_name:
        return False
    return any(lang_name.startswith(pref) for pref in CASE_APPLICABLE_LANG_PREFIXES)


# ----------------- 경로 해석 유틸 (상대/절대/홈/스크립트 폴더 탐색) -----------------
def _resolve_path(p):
    """
    주어진 경로 p를 다음 순서로 해석하여 존재하는 경로를 반환:
    1) 절대경로면 그대로
    2) ~ 홈 확장
    3) 현재 작업 폴더(CWD) 기준
    4) 이 스크립트 파일이 있는 폴더 기준
    존재하지 않으면 마지막 후보(CWD 기준) 경로를 반환
    """
    if not p:
        return None
    p = os.path.expanduser(p)
    if os.path.isabs(p) and os.path.exists(p):
        return p
    # 후보군
    candidates = [
        os.path.join(os.getcwd(), p),
        os.path.join(os.path.dirname(__file__), p),
    ]
    for c in candidates:
        if os.path.exists(c):
            return c
    # 마지막으로 CWD 기준 경로를 반환(파일이 없어도 경로 문자열은 리턴)
    return os.path.join(os.getcwd(), p)


class TranslationChecker:
    """
    엑셀 번역 품질 검수기 (Gemini 단일 모델 기반)
    - 비동기 처리 + 세마포어로 동시성 제어
    - 무의미/짧은 텍스트 스킵 + 화이트리스트 예외
    - 용어집 사전 불일치 감지(시트별 타겟 언어 코드 사용)
    - 시트별 언어/언어코드 매핑(--sheet_langs / --sheet_langs_file) 지원
    - 용어집 CSV의 다언어 컬럼을 한 번에 로드하여 시트별 코드에 맞게 사용
    - (1.0) 타겟 언어에 대한 대소문자(문장형) 하드룰 & 용어집 케이스 체크
    """

    def __init__(
        self,
        gemini_api_key: str,
        model_name: str = "gemini-2.5-flash",
        max_concurrency: int = 10,
        short_text_whitelist=None,
        skip_llm_when_glossary_mismatch: bool = False,
        no_backtranslation = False
    ):
        if not gemini_api_key:
            raise ValueError("Gemini API 키가 설정되지 않았습니다. '.env' 파일을 확인해주세요.")

        # API 설정
        genai.configure(api_key=gemini_api_key)
        self.model_name = model_name
        self.qa_model = genai.GenerativeModel(model_name)

        # ★ 추가된 플래그 저장
        self.no_backtranslation = no_backtranslation

        # 동시성 제한(세마포어)
        if max_concurrency < 1:
            max_concurrency = 1
        self.max_concurrency = max_concurrency
        self._sem = asyncio.Semaphore(self.max_concurrency)

        # 화이트리스트(짧아도 중요한 라벨/약어 예외)
        default_whitelist = {"ok", "on", "off", "ai", "5g", "go", "up", "usb", "nfc"}
        if short_text_whitelist:
            if isinstance(short_text_whitelist, str):
                extra = {x.strip().lower() for x in short_text_whitelist.split(",") if x.strip()}
            else:
                extra = {str(x).strip().lower() for x in short_text_whitelist}
            default_whitelist |= extra
        self.short_text_whitelist = default_whitelist

        # 용어집 사전 불일치 감지 시 LLM 호출 생략 옵션
        self.skip_llm_when_glossary_mismatch = skip_llm_when_glossary_mismatch

        # 용어집 구조:
        # {
        #   source_term: {
        #       'targets': { '한국어': '스마트싱스', '영어_미국': 'SmartThings', ... },
        #       'rule': '...'
        #   },
        #   ...
        # }
        self.glossary = {}
        self.glossary_headers = []  # 전체 헤더 유지(디버그용)
        self.source_lang_code = None  # 로딩 시 지정
        self.rule_header = None       # '설명/규칙' 또는 '규칙' 또는 없음

    # ----------------- CSV Header Guard -----------------
    def _read_csv_with_header_guard(self, csv_path):
        """
        헤더 이중행 방지:
        - 1행과 2행이 동일 헤더면 둘째 행부터 데이터로 간주
        """
        with open(csv_path, "r", encoding="utf-8-sig") as f:
            lines = f.readlines()
        if not lines:
            return csv.DictReader([])

        first = lines[0].strip()
        second = lines[1].strip() if len(lines) > 1 else None
        if second and first == second:
            return csv.DictReader(lines[1:])
        return csv.DictReader(lines)

    def load_glossary_multi(self, csv_path, source_lang_code: str):
        """
        다언어 컬럼을 모두 로드.
        - source_lang_code: 원문이 들어있는 컬럼명(예: '영어_미국' 혹은 'en_US'가 아닌 CSV 실제 헤더명)
        - 나머지 모든 언어 컬럼을 targets에 수집
        """
        try:
            reader = self._read_csv_with_header_guard(csv_path)
            headers = reader.fieldnames or []
            self.glossary_headers = headers[:]
            self.source_lang_code = source_lang_code
            self.rule_header = (
                "설명/규칙" if "설명/규칙" in headers else ("규칙" if "규칙" in headers else None)
            )

            if source_lang_code not in headers:
                print(f"⚠ 경고: 용어집에 원문 컬럼 '{source_lang_code}'이 없습니다. 사용 가능한 컬럼: {', '.join(headers)}")
                return

            count = 0
            for row in reader:
                source_term = (row.get(source_lang_code) or "").strip()
                if not source_term:
                    continue

                # rule
                rule = (row.get(self.rule_header) or "").strip() if self.rule_header else ""

                # targets: 모든 컬럼(원문/규칙 제외)을 타겟 후보로 수집
                targets = {}
                for col in headers:
                    if col == source_lang_code:
                        continue
                    if self.rule_header and col == self.rule_header:
                        continue
                    val = (row.get(col) or "").strip()
                    if val:
                        targets[col] = val

                if targets:
                    self.glossary[source_term] = {"targets": targets, "rule": rule}
                    count += 1

            print(f"✓ 용어집 로드 완료: {count}개 항목 (다언어 타겟 포함)")

        except FileNotFoundError:
            print(f"⚠ 용어집 파일 '{csv_path}'을 찾을 수 없습니다. 경로를 확인해주세요.")
        except Exception as e:
            print(f"⚠ 용어집 로드 실패: {e}")

    # ----------------- 엑셀 로딩 -----------------
    def load_excel_files(self, source_file, target_file, cell_range, selected_sheets=None):
        """엑셀 파일에서 지정된 셀 범위 데이터를 '시트 순서 → 셀 순서'로 평탄화해서 반환"""
        all_data = []

        try:
            wb_source = openpyxl.load_workbook(source_file)
            wb_target = openpyxl.load_workbook(target_file)
        except FileNotFoundError as e:
            print(f"⚠ 엑셀 파일 로드 실패: {e}")
            return []

        # 엑셀 원본 시트 순서 유지
        source_order = wb_source.sheetnames
        target_set = set(wb_target.sheetnames)

        # 공통 시트를 '원본 순서'로만 선별
        common_sheets_in_order = [s for s in source_order if s in target_set]
        if not common_sheets_in_order:
            print("⚠ 경고: 원문 파일과 번역본 파일 사이에 이름이 일치하는 시트가 없습니다!")
            return []

        # 최종 대상 시트: 사용자가 지정했으면 그 '지정 순서' 유지, 아니면 공통 전체
        if selected_sheets:
            # 입력된 순서 그대로, 실제 공통 시트에 존재하는 것만
            target_sheets_to_process = [s for s in selected_sheets if s in common_sheets_in_order]
            if not target_sheets_to_process:
                print(f"⚠ 경고: 지정된 시트({', '.join(selected_sheets)}) 중 공통 파일에 존재하는 시트가 없습니다.")
                return []
        else:
            target_sheets_to_process = common_sheets_in_order

        print(f"\n✓ 검수 대상 시트: {', '.join(target_sheets_to_process)}")

        # 시트 순서 유지 + 셀 순서(행→열)대로 평탄화
        for sheet_name in target_sheets_to_process:
            ws_source = wb_source[sheet_name]
            ws_target = wb_target[sheet_name]
            extracted = 0
            try:
                for source_row, target_row in zip(ws_source[cell_range], ws_target[cell_range]):
                    for s_cell, t_cell in zip(source_row, target_row):
                        s_val = str(s_cell.value).strip() if s_cell.value is not None else ""
                        t_val = str(t_cell.value).strip() if t_cell.value is not None else ""
                        if s_val and t_val:
                            all_data.append({
                                "cell_ref": s_cell.coordinate,
                                "sheet_name": sheet_name,
                                "source": s_val,
                                "target": t_val,
                            })
                            extracted += 1
            except Exception as e:
                print(f"  ⚠ '{sheet_name}' 시트 '{cell_range}' 범위 처리 중 오류: {e}")
                continue

            if extracted:
                print(f"  ✓ '{sheet_name}': {extracted}개 항목 추출 완료")

        return all_data  # 이미 '시트 순서 → 셀 순서'로 정렬된 단일 리스트

    # ----------------- 유틸 -----------------
    async def _with_semaphore(self, coro):
        async with self._sem:
            return await coro

    def _get_target_term_for_code(self, source_term: str, target_lang_code: str):
        """해당 source_term에 대해 타겟 언어 코드의 용어를 돌려줍니다. 없으면 None."""
        entry = self.glossary.get(source_term)
        if not entry:
            return None
        return entry["targets"].get(target_lang_code)

    def _build_glossary_lines_for_code(self, target_lang_code: str):
        """프롬프트에 넣을 용어집 라인 생성(현재 타겟 언어 코드 전용)."""
        if not self.glossary:
            return "용어집 없음"

        out = []
        for source_term, meta in self.glossary.items():
            tgt = meta["targets"].get(target_lang_code)
            if not tgt:
                continue
            rule = meta.get("rule")
            rule_info = f" (규칙: {rule})" if rule else ""
            out.append(f"- 원어: {source_term} → 대상어({target_lang_code}): {tgt}{rule_info}")
        return "\n".join(out) if out else f"용어집에 '{target_lang_code}' 타겟 항목 없음"

    # ----------------- 사전 불일치 감지 -----------------
    def _precheck_glossary_mismatch(self, source_text: str, target_text: str, target_lang_code: str):
        """
        매우 간단한 사전 감지:
        - 원문에 source_term이 포함되면, 번역문에 해당 타겟코드의 target_term이 포함되는지만 확인(대소문자 무시)
        """
        if not self.glossary or not target_lang_code:
            return []

        mismatches = []
        src_lower = source_text.lower()
        tgt_lower = target_text.lower()

        for s_term, meta in self.glossary.items():
            t_term = meta["targets"].get(target_lang_code)
            if not t_term:
                continue  # 해당 코드에 대한 타겟 미정의
            if s_term and s_term.lower() in src_lower:
                if t_term.lower() not in tgt_lower:
                    mismatches.append(f"'{s_term}' → '{t_term}' 미적용({target_lang_code})")
        return mismatches

    # ----------------- 용어집 "대소문자" 체크 -----------------
    def _check_glossary_casing(self, source_text: str, target_text: str, target_lang_code: str):
        """
        - 원문에 source_term이 나오고
        - 번역문에 해당 target_lang_code의 용어가 포함되어 있을 때
        - 실제 번역문 안의 표기가 용어집 표기와 '철자는 같지만 케이스만 다른지' 확인
        """
        if not self.glossary or not target_lang_code:
            return []

        if not source_text or not target_text:
            return []

        issues = []
        src_lower = source_text.lower()
        tgt_lower = target_text.lower()

        for s_term, meta in self.glossary.items():
            t_term = meta["targets"].get(target_lang_code)
            if not t_term:
                continue

            if s_term.lower() in src_lower and t_term.lower() in tgt_lower:
                idx = tgt_lower.find(t_term.lower())
                if idx == -1:
                    continue
                actual = target_text[idx:idx + len(t_term)]
                # 철자는 같은데 케이스만 다르면 이슈로 보고
                if actual.lower() == t_term.lower() and actual != t_term:
                    issues.append(f"용어집 '{t_term}'의 대소문자 표기가 '{actual}'로 사용됨")

        return issues

    # ----------------- 대소문자 하드룰(문장형) 분석 -----------------
    def _analyze_sentence_case(self, target_text: str, target_lang: str):
        """
        - 타겟 언어가 CASE_APPLICABLE_LANG_PREFIXES 에 해당하면
        - 문장을 .!? 기준으로 나누고
        - 각 문장에서 '첫 알파벳만 대문자 + 나머지 소문자' 여부를 단순 판정
        - 전체 텍스트에 대한 '단순 룰 기반 문장형 변환안'도 함께 반환 (참고용)
        - ⚠ 알파벳 판정은 유니코드 기반(str.isalpha/isupper/islower)으로,
          라틴·키릴·베트남어 등 케이스 있는 문자 전부 지원
        """
        if not target_text:
            return None, None
        if not _is_case_sensitive_language(target_lang):
            # 한국어/중국어/일본어 등 케이스 개념 없는 언어는 스킵
            return None, None

        text = target_text.strip()
        # 문장 분리: ., !, ? 뒤의 공백 기준 (아주 단순한 기준)
        sentences = re.split(r'(?<=[\.!?])\s+', text)
        sentences = [s for s in sentences if s.strip()]
        if not sentences:
            return None, None

        report_lines = []
        fixed_sentences = []

        for idx, sent in enumerate(sentences, start=1):
            s = sent

            # ✅ 유니코드 기반: 처음 나오는 "글자"(isalpha=True)를 찾음
            first_alpha_index = None
            for i, ch in enumerate(s):
                if ch.isalpha():  # 라틴, 키릴, 베트남어 등 모두 포함
                    first_alpha_index = i
                    break

            if first_alpha_index is None:
                # 알파벳/글자 자체가 없으면 케이스 판정 불필요
                report_lines.append(f"- 문장 {idx}: 알파벳/문자 없음 → 대소문자 판정 생략")
                fixed_sentences.append(s)
                continue

            first_char = s[first_alpha_index]
            rest = s[first_alpha_index + 1 :]

            # 유니코드 기반 소문자/대문자 판정
            is_sentence_case = first_char.isupper() and rest == rest.lower()

            # 첫 글자 이후의 '추가 대문자' 갯수
            extra_caps = sum(
                1 for ch in rest
                if ch.isalpha() and ch.isupper()
            )

            status = "문장형(첫 글자만 대문자)" if is_sentence_case else "문장형 아님"
            report_lines.append(
                f"- 문장 {idx}: {status}, 추가 대문자 수: {extra_caps}개"
            )

            # 🔧 단순 룰 기반 변환안:
            #   - 첫 글자는 대문자 유지
            #   - 나머지 알파벳은 전부 소문자로 변환
            fixed_rest = rest.lower()
            fixed_sent = s[:first_alpha_index] + first_char.upper() + fixed_rest
            fixed_sentences.append(fixed_sent)

        simple_fixed_text = " ".join(fixed_sentences)
        report = "\n".join(report_lines)

        # 원문과 완전히 같으면 굳이 제안 안 함
        if simple_fixed_text == target_text:
            simple_fixed_text = None

        return report, simple_fixed_text

    # ----------------- LLM 호출 -----------------
    async def check_with_gemini_qa(
        self,
        source_text,
        target_text,
        source_lang,
        target_lang,
        target_lang_code: str,
        max_retries=3,
    ):
        """Gemini 상세 검수(한국어 결과). 타겟 언어 코드 기반 용어집 라인만 프롬프트에 포함."""
        glossary_text = self._build_glossary_lines_for_code(target_lang_code)

        prompt = f"""당신은 전문 번역 검수 전문가입니다.

[Context]
- 원문({source_lang}): {source_text}
- 번역문({target_lang}/{target_lang_code}): {target_text}

[용어집({target_lang_code})]
{glossary_text}

다음 항목에 대해 상세하고 객관적인 검수 결과를 **한국어 불렛 포인트**로 정리해 주세요:
1. **문법/유창성**: 번역문의 어색한 표현이나 문법 오류.
2. **문화/문맥 적절성**: 뉘앙스 손실, 문화적으로 부적절한 요소. **(참고: 'German for Casual Audience (Du-form mandatory)' 등 언어 지정이 있는 경우, 호칭(Du/Sie) 일관성 확인)**
3. **대소문자(Casing) 및 문장형**:
   - 각 문장에서 첫 단어만 대문자이고 나머지는 소문자인지(문장형) 평가하세요.
   - 문장 중간에 등장하는 대문자·ALL CAPS 단어가 고유명/브랜드명/기능명/약어인지 여부를 설명하세요.
4. **용어집 및 규칙 준수**:
   - 용어집에 있는 용어가 올바르게 번역되었는지뿐만 아니라,
   - 용어집에 명시된 대소문자 표기(SmartThings, Galaxy Watch 등)가 그대로 지켜졌는지도 평가하세요.
5. **수정 제안**:
   - 문제가 있다고 판단되는 부분이 있을 경우,
   - 문장형(첫 글자만 대문자) 기준을 해치지 않으면서
   - 고유명/기능명/약어/용어집 표기를 그대로 유지하는 **안전한 수정안**을 제시하세요.

문제가 없다면 마지막 줄에:
"최종 평가: 우수, 주요 문제 없음."
"""

        for attempt in range(max_retries):
            try:
                response = await self.qa_model.generate_content_async(
                    prompt,
                    generation_config={"temperature": 0.2},
                    request_options={"timeout": 90},
                )
                text = getattr(response, "text", None)
                return text.strip() if text else "[응답 비어있음]"
            except Exception as e:
                if attempt < max_retries - 1:
                    await asyncio.sleep(2 ** attempt)  # 1,2,4초 백오프
                else:
                    return f"[Gemini QA 오류] 최대 재시도 횟수 초과: {e}"
        return "[Gemini QA 오류] 알 수 없는 오류"

    async def get_back_translation(self, target_text, target_lang, source_lang, max_retries=3):
        """Gemini 역번역(설명 없이 번역문만)"""
        prompt = (
            f"다음 {target_lang} 텍스트를 {source_lang}으로 다시 번역해주세요. "
            f"오직 번역된 텍스트만 제공해야 합니다. 다른 설명이나 텍스트는 포함하지 마세요.\n\n{target_text}"
        )
        for attempt in range(max_retries):
            try:
                response = await self.qa_model.generate_content_async(
                    prompt,
                    generation_config={"temperature": 0.1},
                    request_options={"timeout": 90},
                )
                text = getattr(response, "text", None)
                return text.strip() if text else "[응답 비어있음]"
            except Exception as e:
                if attempt < max_retries - 1:
                    await asyncio.sleep(2 ** attempt)
                else:
                    return f"[Gemini 역번역 오류] 최대 재시도 횟수 초과: {e}"
        return "[Gemini 역번역 오류] 알 수 없는 오류"

    # ----------------- 메인 처리 -----------------
    async def process_translation_pair_async(
        self,
        data_pair,
        source_lang,
        default_target_lang,
        sheet_lang_map,  # dict: {sheet: {'lang': 'Korean', 'code': '한국어'}}
        default_target_lang_code,
    ):
        """단일 번역 쌍 처리 및 결과 문자열 반환"""
        cell_ref = data_pair["cell_ref"]
        sheet_name = data_pair["sheet_name"]
        source = data_pair["source"]
        target = data_pair["target"]

        # 시트별 언어/코드 결정
        tgt_lang = sheet_lang_map.get(sheet_name, {}).get("lang", default_target_lang)
        tgt_code = sheet_lang_map.get(sheet_name, {}).get("code", default_target_lang_code)

        # DEBUG: 실제 적용된 타겟 언어/코드 출력
        print(f"[{sheet_name}] 적용 타겟 = {tgt_lang} / {tgt_code}")

        # --- 스킵 로직 + 화이트리스트 예외 ---
        is_placeholder = (
            (len(source) <= 2 and len(target) <= 2 and source.lower() == target.lower())
            or (not source.strip() and not target.strip())
            or (source.strip().lower() == target.strip().lower() and len(source.strip()) < 10)
        )
        if source.strip().lower() in self.short_text_whitelist or target.strip().lower() in self.short_text_whitelist:
            is_placeholder = False

        # --- 용어집 사전 불일치 감지(시트 타겟 코드 기준) ---
        pre_mismatch = self._precheck_glossary_mismatch(source, target, tgt_code)
        skip_llm = self.skip_llm_when_glossary_mismatch and bool(pre_mismatch)

        # --- 대소문자 하드룰(문장형) 분석 & 용어집 케이스 체크 (Python 레벨) ---
        case_report, simple_case_fix = self._analyze_sentence_case(target, tgt_lang)
        glossary_case_issues = self._check_glossary_casing(source, target, tgt_code)

        # 대소문자 점검 섹션 텍스트 구성
        if case_report:
            case_section = "대소문자 하드룰(문장형) 점검:\n" + case_report
            if simple_case_fix:
                case_section += "\n\n[단순 규칙 기반 문장형 변환안(참고용)]:\n" + simple_case_fix
        else:
            case_section = "별도 지적 사항 없음."

        # 용어집 점검 섹션 텍스트 구성 (사전 감지 + 케이스 이슈 묶어서)
        glossary_parts = []
        if pre_mismatch:
            glossary_parts.append("용어집 사전 감지:\n- " + "\n- ".join(pre_mismatch))
        if glossary_case_issues:
            glossary_parts.append(
                "용어집 대소문자 표기 점검:\n" +
                "\n".join(f"- {msg}" for msg in glossary_case_issues)
            )
        if glossary_parts:
            glossary_section = "\n\n".join(glossary_parts)
        else:
            glossary_section = "별도 지적 사항 없음."

        # --- placeholder (짧은/무의미 텍스트) 처리 ---
        if is_placeholder and not pre_mismatch:
            back_translation = "[건너뜀: 텍스트가 짧거나 무의미하여 AI 호출을 생략했습니다.]"
            gemini_qa_review = "[건너뜀: 텍스트가 짧거나 무의미하여 AI 호출을 생략했습니다.]"

            result_content = (
                f"\n\n{'='*90}\n"
                f"[시트] {sheet_name} | [셀] {cell_ref}\n"
                f"{'-'*90}\n\n"
                f"[상세 - 원문]\n"
                f"{source}\n\n"
                f"[상세 - 번역문]\n"
                f"{target}\n\n"
                f"[상세 - 대소문자 점검]\n"
                f"{case_section}\n\n"
                f"[상세 - 용어집 점검]\n"
                f"{glossary_section}\n\n"
                f"[상세 - 역번역]\n"
                f"{back_translation}\n\n"
                f"[상세 - Gemini 검수 결과]\n"
                f"{gemini_qa_review}\n"
                f"{'='*90}\n"
            )
            return result_content

        # --- LLM 호출 ---
        print(f"[{sheet_name}] {cell_ref} 비동기 작업 시작...")

        if skip_llm:
            back_translation = "[사전 감지로 LLM 호출 생략]"
            gemini_qa_review = "※ 용어집 사전 감지 결과를 우선 검토하세요. (옵션에 의해 LLM 호출 생략됨)"
        else:
            qa_task = self._with_semaphore(
                self.check_with_gemini_qa(source, target, source_lang, tgt_lang, tgt_code)
            )

            # no_backtranslation 옵션이 켜져 있으면 역번역 스킵
            if getattr(self, "no_backtranslation", False):
                gemini_qa_review = await qa_task
                back_translation = "[역번역 비활성화됨 (--no_backtranslation)]"
            else:
                bt_task = self._with_semaphore(
                    self.get_back_translation(target, tgt_lang, source_lang)
                )
                gemini_qa_review, back_translation = await asyncio.gather(qa_task, bt_task)

        print(f"[{sheet_name}] {cell_ref} 처리 완료.")

        # 결과 블록 최종 조립
        result_content = (
            f"\n\n{'='*90}\n"
            f"[시트] {sheet_name} | [셀] {cell_ref}\n"
            f"{'-'*90}\n\n"
            f"[상세 - 원문]\n"
            f"{source}\n\n"
            f"[상세 - 번역문]\n"
            f"{target}\n\n"
            f"[상세 - 대소문자 점검]\n"
            f"{case_section}\n\n"
            f"[상세 - 용어집 점검]\n"
            f"{glossary_section}\n\n"
            f"[상세 - 역번역]\n"
            f"{back_translation}\n\n"
            f"[상세 - Gemini 검수 결과]\n"
            f"{gemini_qa_review}\n"
            f"{'='*90}\n"
        )
        return result_content

    async def main_async(
        self,
        source_file,
        target_file,
        cell_range,
        source_lang,
        target_lang,
        source_lang_code=None,      # 원문 컬럼
        target_lang_code=None,      # (기본) 타겟 컬럼
        glossary_file=None,
        sheet_names=None,
        sheet_langs=None,           # "Sheet:LangName:LangHeader"
        sheet_langs_file=None,      # JSON 인자
    ):
        """메인 비동기 실행 함수"""
        print("=" * 90)
        print("번역 검수 스크립트 ver1.0 (Casing 강화 / Semaphore + Whitelist + Precheck + SheetLangs)")
        print(f"사용 모델: {self.model_name}")
        print("=" * 90)

        # 1) 시트별 언어 매핑
        sheet_lang_map = {}

        # ★ DEBUG: 넘겨받은 JSON 경로/존재 여부 + CWD/해석경로 출력
        print(f"[DEBUG] sheet_langs_file 인자값 = {sheet_langs_file}")
        print(f"[DEBUG] CWD = {os.getcwd()}")
        resolved = _resolve_path(sheet_langs_file) if sheet_langs_file else None
        print(f"[DEBUG] resolved path = {resolved}")
        if resolved:
            print(f"[DEBUG] os.path.exists(resolved) = {os.path.exists(resolved)}")

        # 1-1) JSON 파일 우선
        if resolved and os.path.exists(resolved):
            try:
                with open(resolved, 'r', encoding='utf-8') as f:
                    loaded = json.load(f)
                # 기대 스키마: { "KR(한국)": {"lang":"Korean","code":"한국어"}, ... }
                # 키/값 정규화(양끝 공백 제거)
                for k, v in loaded.items():
                    if not isinstance(v, dict):
                        print(f"⚠ 무시됨: '{k}' 값이 객체가 아닙니다.")
                        continue
                    name = str(k).strip()
                    lang = str(v.get("lang", "")).strip()
                    code = str(v.get("code", "")).strip()
                    if not name or not lang or not code:
                        print(f"⚠ 무시됨: '{k}' 매핑에 lang/code가 비었습니다.")
                        continue
                    sheet_lang_map[name] = {"lang": lang, "code": code}
                print(f"✓ 시트 언어 매핑(JSON) 로드: {len(sheet_lang_map)}개")
            except Exception as e:
                print(f"⚠ sheet_langs.json 로드 실패: {e}")
        else:
            if sheet_langs_file:
                print(f"⚠ 경고: 지정된 JSON 경로를 찾을 수 없습니다 → {sheet_langs_file} (resolved: {resolved})")

        # 1-2) 문자열 인자 파싱(폴백)
        if not sheet_lang_map and sheet_langs:
            pairs = [s.strip() for s in sheet_langs.split(",") if ":" in s]
            for p in pairs:
                try:
                    name, lang_name, lang_code = [x.strip() for x in p.split(":")]
                    if name and lang_name and lang_code:
                        sheet_lang_map[name] = {"lang": lang_name, "code": lang_code}
                except ValueError:
                    print(f"⚠ 시트 언어 매핑 구문 오류: {p}")

        # 2) 용어집 로드(다언어)
        if glossary_file and os.path.exists(glossary_file):
            if source_lang_code:
                self.load_glossary_multi(glossary_file, source_lang_code)
            else:
                print("⚠ 경고: 용어집 사용 시 원문 언어 코드(--src_code)가 필요합니다.")
        else:
            if glossary_file:
                print(f"⚠ 용어집 파일을 찾을 수 없습니다: {glossary_file}")

        # 3) 엑셀 로드
        print(f"\n▶ 엑셀 파일 로드 중...")
        selected_sheets_list = [s.strip() for s in sheet_names.split(",")] if sheet_names else None
        all_data = self.load_excel_files(
            source_file, target_file, cell_range, selected_sheets=selected_sheets_list
        )
        if not all_data:
            print("\n⚠ 처리할 데이터가 없거나 파일 매칭에 실패했습니다. 스크립트를 종료합니다.")
            return

        total_items = len(all_data)
        print(f"\n✓ 로드 완료: 총 {total_items}개 항목")

        # (선택) 매핑 진단: 실제 시트와 매핑 대조
        actual_sheets = {item["sheet_name"] for item in all_data}
        if sheet_lang_map:
            missing = [name for name in sheet_lang_map if name not in actual_sheets]
            if missing:
                print(f"⚠ 경고: JSON/문자열에만 있고 엑셀에 없는 시트 → {', '.join(missing)}")
            unmapped = [s for s in sorted(actual_sheets) if s not in sheet_lang_map]
            if unmapped:
                print(f"⚠ 경고: 매핑 없는 시트는 기본 타겟({target_lang}/{target_lang_code})으로 처리 → {', '.join(unmapped)}")

        # 4) 보고서 파일 헤더
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"translation_review_{timestamp}.txt"
        with open(output_file, "w", encoding="utf-8") as f:
            f.write(f"--- 번역 검수 보고서 (Gemini 단일 모델 기반 / 규칙 + Casing 반영) ---\n")
            f.write(f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"총 검수 항목: {total_items}개\n")
            f.write(f"사용 모델: {self.model_name}\n")
            if self.glossary:
                f.write(f"용어집 사용: {len(self.glossary)}개 원어 항목 (다언어 타겟 포함)\n")
            sheets_display = ", ".join(selected_sheets_list) if selected_sheets_list else "전체"
            f.write(f"검수 대상 시트: {sheets_display}\n")
            f.write(f"동시성 제한: {self.max_concurrency}\n")
            f.write(f"짧은 텍스트 화이트리스트: {', '.join(sorted(self.short_text_whitelist))}\n")
            f.write(f"용어집 사전 불일치 시 LLM 스킵: {self.skip_llm_when_glossary_mismatch}\n")
            if sheet_lang_map:
                f.write("시트별 언어 설정:\n")
                for name, info in sheet_lang_map.items():
                    f.write(f" - {name}: {info['lang']} ({info['code']})\n")
            else:
                f.write(f"(기본 타겟) {target_lang} ({target_lang_code})\n")

        # 5) 비동기 작업 실행(순서 보장: gather)
        tasks = [
            self.process_translation_pair_async(
                data_pair=dp,
                source_lang=source_lang,
                default_target_lang=target_lang,
                sheet_lang_map=sheet_lang_map,
                default_target_lang_code=target_lang_code,
            )
            for dp in all_data
        ]
        print(f"\nAPI 호출 병렬 처리 시작 (총 {total_items}개 항목, 동시성 {self.max_concurrency})")
        all_results = await asyncio.gather(*tasks)

        # 6) 결과 파일 기록
        print("✓ 비동기 작업 완료. 결과를 순차적으로 파일에 작성 중...")
        with open(output_file, "a", encoding="utf-8") as f:
            for result_content in all_results:
                f.write(result_content)

        print(f"\n{'='*90}")
        print(f"★★★ 최종 검수 완료! ★★★")
        print(f"결과 파일: {output_file}")
        print(f"결과 파일은 이 스크립트가 실행된 폴더에 저장되었습니다.")
        print(f"{'='*90}")

    def run(self, *args, **kwargs):
        return asyncio.run(self.main_async(*args, **kwargs))


# ----------------- CLI -----------------
if __name__ == "__main__":
    GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")

    parser = argparse.ArgumentParser(description="엑셀 기반 번역 품질 검수 스크립트 (Gemini 단일 모델 기반)")
    parser.add_argument("--source_file", required=True, help="원문 엑셀 파일 경로 (예: original.xlsx)")
    parser.add_argument("--target_file", required=True, help="번역본 엑셀 파일 경로 (예: translation.xlsx)")
    parser.add_argument("--range", required=True, help="검수할 셀 범위 (예: A2:A100)")
    parser.add_argument("--src_lang", default="English", help="원문 언어 (예: English)")
    parser.add_argument("--tgt_lang", default="Korean", help="기본 번역 언어 (예: Korean)")
    parser.add_argument("--src_code", default="en_US", help="용어집 CSV의 원문 언어 컬럼명 (예: en_US 또는 CSV 실제 헤더명)")
    parser.add_argument("--tgt_code", default="ko_KR", help="(기본) 타겟 언어 컬럼명 (예: ko_KR 또는 CSV 실제 헤더명)")
    parser.add_argument("--glossary", default="glossary.csv", help="용어집 CSV 파일 경로 (선택 사항)")
    parser.add_argument(
        "--sheet_names",
        help="검수할 시트 이름을 쉼표(,)로 지정 (예: KR(한국),US(미국))",
    )
    parser.add_argument(
        "--sheet_langs",
        help="시트별 언어/코드 (예: AE(아랍에메리트):Arabic:아랍에미리트)",
    )
    parser.add_argument(
        "--sheet_langs_file",
        help="시트별 언어/코드 매핑 JSON 파일 경로 (예: sheet_langs.json)"
    )
    parser.add_argument(
        "--max_concurrency",
        type=int,
        default=10,
        help="동시 호출 제한 (세마포어). 기본 10, 환경에 맞게 5~15 권장.",
    )
    parser.add_argument(
        "--whitelist",
        default="",
        help="짧은 텍스트 스킵 예외 화이트리스트(콤마 구분). 예: OK,ON,OFF,NFC",
    )
    parser.add_argument(
        "--skip_llm_when_glossary_mismatch",
        action="store_true",
        help="용어집 사전 불일치가 감지되면 LLM 호출을 생략하고 사전 감지만 보고합니다.",
    )
    parser.add_argument(
        "--model",
        default="gemini-2.5-flash",
        help="사용할 Gemini 모델명(기본: gemini-2.5-flash)",
    )
    parser.add_argument(
    "--no_backtranslation",
    action="store_true",
    help="역번역 단계를 비활성화합니다.",
    )

    args = parser.parse_args()

    try:
        checker = TranslationChecker(
            gemini_api_key=GEMINI_API_KEY,
            model_name=args.model,
            max_concurrency=args.max_concurrency,
            short_text_whitelist=args.whitelist,
            skip_llm_when_glossary_mismatch=args.skip_llm_when_glossary_mismatch,
            no_backtranslation = args.no_backtranslation
        )

        checker.run(
            source_file=args.source_file,
            target_file=args.target_file,
            cell_range=args.range,
            source_lang=args.src_lang,
            target_lang=args.tgt_lang,
            source_lang_code=args.src_code,
            target_lang_code=args.tgt_code,
            glossary_file=args.glossary,
            sheet_names=args.sheet_names,
            sheet_langs=args.sheet_langs,
            sheet_langs_file=args.sheet_langs_file,  # JSON 경로 전달
        )

    except ValueError as e:
        print(f"\n[FATAL ERROR] 실행 오류: {e}")
        print("API 키를 환경 변수(.env 파일)에 정확히 설정했는지 확인해주세요.")
    except Exception as e:
        print(f"\n[FATAL ERROR] 예기치 않은 오류 발생: {e}")