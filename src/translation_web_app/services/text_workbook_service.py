"""Create source workbooks from structured text input."""

from __future__ import annotations

import re
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from uuid import uuid4

import openpyxl

from translation_web_app.paths import SOURCE_WORKBOOK_TEMPLATE, UPLOAD_DIR


TEXT_INPUT_RANGE = "C7:C28"
SECTION_COUNT_MAX = 4

STORY_CELLS = {
    "title": "C7",
    "description": "C8",
}

SECTION_CELLS = [
    {"title": "C10", "description": "C11", "disclaimer": "C12", "button": "C13"},
    {"title": "C15", "description": "C16", "disclaimer": "C17", "button": "C18"},
    {"title": "C20", "description": "C21", "disclaimer": "C22", "button": "C23"},
    {"title": "C25", "description": "C26", "disclaimer": "C27", "button": "C28"},
]

WRITABLE_TEXT_CELLS = [
    "C7", "C8",
    "C10", "C11", "C12", "C13",
    "C15", "C16", "C17", "C18",
    "C20", "C21", "C22", "C23",
    "C25", "C26", "C27", "C28",
]


@dataclass(frozen=True)
class GeneratedWorkbook:
    path: Path
    file_id: str
    file_name: str
    story_id: str
    cell_range: str = TEXT_INPUT_RANGE


def normalize_story_id(story_number: str | int | None) -> str:
    if story_number is None or str(story_number).strip() == "":
        return f"story_{datetime.now().strftime('%H%M%S')[-3:]}"

    raw = str(story_number).strip()
    match = re.search(r"(\d+)$", raw)
    if not match:
        raise ValueError("story_number must end with digits, for example 52 or story_052.")

    return f"story_{int(match.group(1)):03d}"


def normalize_update_date(update_date: str | None) -> str:
    if update_date and update_date.strip():
        return update_date.strip()
    return datetime.now().strftime("%Y.%m.%d")


def create_text_source_workbook(
    *,
    source_sheet: str,
    story_number: str | int | None,
    update_date: str | None,
    story: dict,
    sections: list[dict],
    output_dir: Path = UPLOAD_DIR,
    template_path: Path = SOURCE_WORKBOOK_TEMPLATE,
) -> GeneratedWorkbook:
    if not template_path.exists():
        raise FileNotFoundError(f"Source workbook template not found: {template_path}")

    if not source_sheet:
        raise ValueError("source_sheet is required.")

    if len(sections or []) > SECTION_COUNT_MAX:
        raise ValueError(f"sections supports at most {SECTION_COUNT_MAX} items.")

    output_dir.mkdir(parents=True, exist_ok=True)
    story_id = normalize_story_id(story_number)
    file_id = f"text_src_{uuid4().hex}.xlsx"
    out_path = output_dir / file_id
    shutil.copy2(template_path, out_path)

    wb = openpyxl.load_workbook(out_path, rich_text=True)
    if source_sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"source_sheet '{source_sheet}' not found in template.")

    date_value = normalize_update_date(update_date)
    for ws in wb.worksheets:
        ws["C2"] = date_value
        ws["C5"] = story_id
        for coord in WRITABLE_TEXT_CELLS:
            ws[coord] = None

    source_ws = wb[source_sheet]
    story = story or {}
    source_ws[STORY_CELLS["title"]] = _clean_text(story.get("title"))
    source_ws[STORY_CELLS["description"]] = _clean_text(story.get("description"))

    for idx, section in enumerate(sections or []):
        cells = SECTION_CELLS[idx]
        section = section or {}
        for key, coord in cells.items():
            source_ws[coord] = _clean_text(section.get(key))

    wb.save(out_path)
    wb.close()

    return GeneratedWorkbook(
        path=out_path,
        file_id=file_id,
        file_name=f"{story_id}_source.xlsx",
        story_id=story_id,
    )


def _clean_text(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None

