from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from workbook_story_apply import (  # noqa: E402
    _changed_cells,
    _split_sheets,
    _verify_highlight_scope,
    _verify_values,
)


class WorkbookStoryApplyTests(unittest.TestCase):
    def test_split_sheets_rejects_empty_and_duplicates(self):
        self.assertEqual(_split_sheets("VN(베트남), TH(태국)"), ["VN(베트남)", "TH(태국)"])
        with self.assertRaises(ValueError):
            _split_sheets("")
        with self.assertRaises(ValueError):
            _split_sheets("VN(베트남),VN(베트남)")

    def test_verify_values_accepts_only_approved_changes(self):
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "source.xlsx"
            final = Path(tmp) / "final.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "VN(베트남)"
            ws["C8"] = "before"
            ws["C10"] = "unchanged"
            wb.save(source)

            wb["VN(베트남)"]["C8"] = "after"
            wb.save(final)

            result = _verify_values(source, final, {("VN(베트남)", "C8")})
            self.assertEqual(result, {"expected_value_changes": 1, "actual_value_changes": 1})

    def test_verify_values_rejects_unapproved_changes(self):
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "source.xlsx"
            final = Path(tmp) / "final.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "VN(베트남)"
            ws["C8"] = "before"
            ws["C10"] = "also before"
            wb.save(source)

            ws["C8"] = "after"
            ws["C10"] = "unexpected"
            wb.save(final)

            with self.assertRaises(RuntimeError):
                _verify_values(source, final, {("VN(베트남)", "C8")})

    def test_changed_cells_ignores_noop(self):
        changes = [
            {"sheet": "VN(베트남)", "cell": "C8", "old_value": "same", "new_value": "same"},
            {"sheet": "TH(태국)", "cell": "C10", "old_value": "old", "new_value": "new"},
        ]
        self.assertEqual(_changed_cells(changes), {("TH(태국)", "C10")})

    def test_verify_highlight_scope_requires_every_delivery_sheet(self):
        output = (
            "[그룹 2: US(미국)] VN(베트남) | 처리 완료 셀: 6 | 총 하이라이트: 18개\n"
            "[그룹 2: US(미국)] TH(태국) | 처리 완료 셀: 6 | 총 하이라이트: 17개\n"
        )
        result = _verify_highlight_scope(output, ["VN(베트남)", "TH(태국)"])
        self.assertEqual(result["completed_delivery_sheets"]["VN(베트남)"], 6)
        with self.assertRaises(RuntimeError):
            _verify_highlight_scope(output, ["VN(베트남)", "ID(인도네시아)"])


if __name__ == "__main__":
    unittest.main()
