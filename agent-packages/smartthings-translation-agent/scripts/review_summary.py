#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
review_summary.py — 원어민/번역사 감수본과 AI 검수 txt 요약 (읽기 전용)

Excel 원본을 수정하지 않는다. 감수본의 F/H열, AI 검수 txt의 등급/수정안,
선택적으로 md 리포트의 반영 판단 표를 집계해 최종 summary용 숫자를 출력한다.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - 환경 안내용
    raise SystemExit("openpyxl이 필요합니다. app venv 또는 requirements-excel.txt 환경에서 실행하세요.") from exc


DEFAULT_SOURCE_SHEETS = {"KR(한국)", "US(미국)"}
DEFAULT_ROWS = range(7, 29)


def _cell_text(value: Any) -> str:
    return "" if value is None else str(value)


def _load_workbook(path: Path):
    if not path.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {path}")
    return openpyxl.load_workbook(path, data_only=True, read_only=True)


def _target_sheets(wb, source_sheets: set[str]) -> list[str]:
    return [name for name in wb.sheetnames if name not in source_sheets]


def summarize_review_workbook(
    review_path: Path,
    count_path: Path | None,
    source_sheets: set[str],
) -> tuple[dict[str, Any], set[tuple[str, int]]]:
    review_wb = _load_workbook(review_path)
    count_wb = _load_workbook(count_path) if count_path else review_wb
    try:
        sheets = _target_sheets(count_wb, source_sheets)
        total_target_cells = 0
        for sheet in sheets:
            ws = count_wb[sheet]
            total_target_cells += sum(
                1 for row in DEFAULT_ROWS
                if _cell_text(ws[f"C{row}"].value).strip()
            )

        touched: set[tuple[str, int]] = set()
        changed: set[tuple[str, int]] = set()
        commented: set[tuple[str, int]] = set()
        per_sheet = []
        for sheet in _target_sheets(review_wb, source_sheets):
            ws = review_wb[sheet]
            sheet_touched = []
            sheet_changed = []
            sheet_comments = []
            for row in DEFAULT_ROWS:
                current = ws[f"C{row}"].value
                revised = ws[f"F{row}"].value
                comment = ws[f"H{row}"].value
                key = (sheet, row)
                if revised is not None or comment is not None:
                    touched.add(key)
                    sheet_touched.append(row)
                if comment:
                    commented.add(key)
                    sheet_comments.append(row)
                if revised is not None and _cell_text(revised) != _cell_text(current):
                    changed.add(key)
                    sheet_changed.append(row)
            if sheet_touched or sheet_changed or sheet_comments:
                per_sheet.append({
                    "sheet": sheet,
                    "touched": len(sheet_touched),
                    "changed": len(sheet_changed),
                    "comments": len(sheet_comments),
                    "changed_rows": sheet_changed,
                })

        return {
            "review_workbook": str(review_path),
            "count_workbook": str(count_path or review_path),
            "source_sheets_excluded": sorted(source_sheets),
            "sheets_counted": len(sheets),
            "total_target_cells": total_target_cells,
            "reviewer_touched_cells": len(touched),
            "reviewer_changed_cells": len(changed),
            "comment_cells": len(commented),
            "per_sheet": per_sheet,
        }, changed
    finally:
        review_wb.close()
        if count_wb is not review_wb:
            count_wb.close()


def parse_ai_review(path: Path | None) -> tuple[dict[str, Any] | None, set[tuple[str, int]]]:
    if not path:
        return None, set()
    if not path.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {path}")
    text = path.read_text(encoding="utf-8")
    marker = re.compile(r"\[시트\] ([^|]+) \| \[셀\] C(\d+)")
    matches = list(marker.finditer(text))
    grade_counts: Counter[str] = Counter()
    suggested_cells: set[tuple[str, int]] = set()
    blocks = 0

    for idx, match in enumerate(matches):
        start = match.start()
        end = matches[idx + 1].start() if idx + 1 < len(matches) else len(text)
        chunk = text[start:end]
        blocks += 1
        if "최종 평가: 수정 필요" in chunk:
            grade_counts["Needs Revision"] += 1
        elif "최종 평가: 양호" in chunk:
            grade_counts["Good"] += 1
        elif "최종 평가: 우수" in chunk:
            grade_counts["Excellent"] += 1

        suggested = re.search(r"\[수정안 제안\]:[^\n]*\n(.*?)(?:\n\n|\n\[상세|\n=)", chunk, re.S)
        if suggested and suggested.group(1).strip():
            suggested_cells.add((match.group(1).strip(), int(match.group(2))))

    return {
        "ai_review_txt": str(path),
        "ai_blocks": blocks,
        "ai_grade_counts": dict(grade_counts),
        "ai_suggested_fix_count": len(suggested_cells),
    }, suggested_cells


def _classify_judgment(judgment: str) -> str:
    if any(token in judgment for token in ("수용 비추천", "미반영", "감수안 수용 비추천")):
        return "reject"
    if "부분 수용" in judgment or "부분 반영" in judgment:
        return "partial"
    if "추가 적용 없음" in judgment or judgment == "유지" or "적용 대상 없음" in judgment:
        return "no_apply"
    if any(token in judgment for token in (
        "전체 수용",
        "수용 가능",
        "수용 권장",
        "번역사 의견 전체 반영",
        "반영 가능",
        "수용.",
    )):
        return "accept"
    return "conditional_or_other"


def parse_report_md(path: Path | None) -> dict[str, Any] | None:
    if not path:
        return None
    if not path.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {path}")
    text = path.read_text(encoding="utf-8")
    if "## 10. 재감수 대조 기록" in text:
        text = text.split("## 10. 재감수 대조 기록", 1)[1]
    if "## 11. 최종 Summary" in text:
        text = text.split("## 11. 최종 Summary", 1)[0]

    rows = []
    counts: Counter[str] = Counter()
    for line in text.splitlines():
        if not line.startswith("| `"):
            continue
        parts = [part.strip() for part in line.strip("|").split("|")]
        if len(parts) < 4 or not re.match(r"`[A-Z]{2} C\d+", parts[0]):
            continue
        category = _classify_judgment(parts[3])
        counts[category] += 1
        rows.append({"cell": parts[0].strip("`"), "classification": category})

    return {
        "report_md": str(path),
        "report_rows": len(rows),
        "report_classification_counts": dict(counts),
    }


def build_summary(args: argparse.Namespace) -> dict[str, Any]:
    source_sheets = {item.strip() for item in args.source_sheets.split(",") if item.strip()}
    review, reviewer_changed = summarize_review_workbook(
        Path(args.review_workbook),
        Path(args.fix_workbook) if args.fix_workbook else None,
        source_sheets,
    )
    ai, ai_suggested = parse_ai_review(Path(args.ai_review_txt) if args.ai_review_txt else None)
    report = parse_report_md(Path(args.report_md) if args.report_md else None)

    result: dict[str, Any] = {"review": review}
    if ai:
        result["ai_review"] = ai
        result["ai_overlap"] = {
            "reviewer_changed_with_ai_suggestion": len(reviewer_changed & ai_suggested),
            "reviewer_changed_without_ai_suggestion": len(reviewer_changed - ai_suggested),
            "ai_suggested_without_reviewer_change": len(ai_suggested - reviewer_changed),
        }
    if report:
        result["report"] = report
    return result


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="감수본/AI 검수 결과 요약 수치 산출 (읽기 전용)")
    parser.add_argument("--review-workbook", required=True, help="F/H열이 포함된 원어민/번역사 감수본 xlsx")
    parser.add_argument("--fix-workbook", help="총 번역 대상 셀 카운트 기준이 될 fix/source xlsx")
    parser.add_argument("--ai-review-txt", help="AI/RAG 검수 txt")
    parser.add_argument("--report-md", help="재감수 대조 판단이 정리된 md 리포트")
    parser.add_argument("--source-sheets", default="KR(한국),US(미국)", help="집계에서 제외할 source sheet 목록")
    parser.add_argument("--json", action="store_true", help="JSON으로 출력")
    args = parser.parse_args(argv)

    try:
        result = build_summary(args)
    except Exception as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 1

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        review = result["review"]
        print(f"총 번역 대상: {review['total_target_cells']}셀")
        print(f"번역사 수정 제안: {review['reviewer_changed_cells']}셀")
        print(f"코멘트/검토 흔적: {review['reviewer_touched_cells']}셀")
        if "ai_overlap" in result:
            overlap = result["ai_overlap"]
            print(f"AI와 겹친 항목: {overlap['reviewer_changed_with_ai_suggestion']}건")
            print(f"AI 수정안 중 최종 미변경: {overlap['ai_suggested_without_reviewer_change']}건")
        if "report" in result:
            print(f"리포트 판단: {result['report']['report_classification_counts']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
