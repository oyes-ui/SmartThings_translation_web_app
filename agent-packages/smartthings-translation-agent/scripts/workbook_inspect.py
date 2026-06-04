#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
workbook_inspect.py — SmartThings 번역 워크북 분석 (읽기 전용)

원본 파일을 절대 수정하지 않는다. 시트 목록 / story_id / 섹션·내용 행 요약 /
언어 매핑 힌트를 출력해, 에이전트가 어떤 셀이 편집 후보인지 파악하도록 돕는다.

파싱 위치/범위 상수는 rag_db_builder.py 에서 import 하여 재사용한다(중복 하드코딩 금지).

사용 예:
  python workbook_inspect.py path/to/story.xlsx
  python workbook_inspect.py path/to/story.xlsx --sheet "JA(일본)" --json
  python workbook_inspect.py path/to/story.xlsx --cell-range C7:C28
"""

import re
import sys
import json
import argparse
from pathlib import Path


def _bootstrap_project():
    """app_root 의 src/translation_web_app 를 sys.path 에 추가 (없어도 비치명적).

    bootstrap.py(sibling)의 resolve_app_root 를 우선 사용(--app-root/env/config/탐색).
    이 스크립트는 Excel 분석만 하므로 RAG 스택 없는 경량 환경에서도 동작해야 한다.
    따라서 src 를 못 찾아도 예외를 던지지 않는다(상수는 fallback 으로 대체).
    """
    try:
        sys.path.insert(0, str(Path(__file__).resolve().parent))
        import bootstrap as _bs
        app_root, _src = _bs.resolve_app_root(_bs.cli_app_root_from_argv())
        if app_root:
            src_dir = Path(app_root) / "src"
            if src_dir.is_dir() and str(src_dir) not in sys.path:
                sys.path.insert(0, str(src_dir))
            return Path(app_root)
    except Exception:
        pass

    here = Path(__file__).resolve()
    for parent in [here.parent, *here.parents]:
        if (parent / "src" / "translation_web_app").is_dir():
            src_dir = parent / "src"
            if str(src_dir) not in sys.path:
                sys.path.insert(0, str(src_dir))
            return parent
    return None


_bootstrap_project()

import openpyxl  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

# 파싱 위치/범위 상수: 정상 환경에서는 rag_db_builder 에서 재사용(동기화 유지),
# chromadb/google-genai 가 없는 경량 환경에서는 로컬 fallback 으로 openpyxl 만으로 동작.
# ⚠ fallback 값은 rag_db_builder.py 의 상수와 반드시 동일하게 유지할 것.
try:
    from translation_web_app.rag_db_builder import (  # noqa: E402
        STORY_ID_CELL,
        CONTENT_ROW_START,
        CONTENT_ROW_END,
        SECTION_COL,
        CONTENT_COL,
        GROUP_A_KEY,
        GROUP_B_KEY,
    )
except Exception:  # ImportError 또는 무거운 의존성 부재
    STORY_ID_CELL = "C5"
    CONTENT_ROW_START = 7
    CONTENT_ROW_END = 28
    SECTION_COL = 2   # B열
    CONTENT_COL = 3   # C열
    GROUP_A_KEY = "KR(한국)"
    GROUP_B_KEY = "US(미국)"


def _lang_code(sheet_name: str) -> str:
    """시트명에서 언어 코드 추출: 'JA(일본)' → 'JA'."""
    return re.split(r"[(（]", sheet_name.strip(), maxsplit=1)[0].strip().upper()


def inspect_sheet(ws) -> dict:
    """단일 시트의 story_id 와 섹션/내용 행 요약."""
    story_id = ws[STORY_ID_CELL].value
    rows = []
    populated = 0
    for r in range(CONTENT_ROW_START, CONTENT_ROW_END + 1):
        section = ws.cell(r, SECTION_COL).value
        content = ws.cell(r, CONTENT_COL).value
        if section is None and content is None:
            continue
        populated += 1
        text = "" if content is None else str(content)
        rows.append({
            "row": r,
            "section_code": section,
            "content_preview": text[:80] + ("…" if len(text) > 80 else ""),
            "content_len": len(text),
        })
    return {
        "story_id": story_id,
        "populated_rows": populated,
        "rows": rows,
    }


# ─── section-level grouping (--sections) ──────────────────────────────────────
# SmartThings 워크북은 story(C7/C8)와 4개 section(각 4행)으로 구성된다.
# 각 그룹의 행을 위치 순서대로 title/description/disclaimer/button 에 매핑한다.
SECTION_GROUPS = [
    ("story", "story", (7, 8)),
    ("section", "section_1", (10, 11, 12, 13)),
    ("section", "section_2", (15, 16, 17, 18)),
    ("section", "section_3", (20, 21, 22, 23)),
    ("section", "section_4", (25, 26, 27, 28)),
]
FIELD_NAMES = ("title", "description", "disclaimer", "button")


def _is_empty_or_placeholder(raw) -> bool:
    """빈 값 또는 'x' placeholder 인지."""
    if raw is None:
        return True
    s = str(raw).strip()
    return s == "" or s.lower() == "x"


def _make_field(ws, row: int) -> dict:
    section_code = ws.cell(row, SECTION_COL).value
    raw = ws.cell(row, CONTENT_COL).value
    return {
        "row": row,
        "cell": f"{get_column_letter(CONTENT_COL)}{row}",
        "section_code": section_code,
        "text": "" if raw is None else str(raw),
        "is_empty_or_placeholder": _is_empty_or_placeholder(raw),
    }


def _has_content(field: dict | None) -> bool:
    """섹션코드가 있거나 텍스트가 비어있지 않으면 실제 콘텐츠로 간주."""
    if not field:
        return False
    return field["section_code"] is not None or not field["is_empty_or_placeholder"]


def build_groups(ws, story_id, sheet_name: str) -> list[dict]:
    """story / section 단위로 title·description·disclaimer·button 을 묶는다.

    - title/description: 항상 포함 (비어 있어도 is_empty_or_placeholder 로 표시).
    - disclaimer/button: optional — 해당 행에 section_code 도 없고 비어 있으면 생략.
    - title·description 이 모두 비어 있는 그룹(예: 미사용 section)은 생성하지 않는다.
    """
    groups = []
    for gtype, gid, rows in SECTION_GROUPS:
        fields: dict[str, dict] = {}
        for idx, row in enumerate(rows):
            fname = FIELD_NAMES[idx]
            fld = _make_field(ws, row)
            if fname in ("disclaimer", "button"):
                # optional: 내용도 없고 섹션코드도 없으면 생략
                if fld["section_code"] is None and fld["is_empty_or_placeholder"]:
                    continue
            fields[fname] = fld

        if not (_has_content(fields.get("title")) or _has_content(fields.get("description"))):
            continue  # 미사용 섹션 skip

        groups.append({
            "group_type": gtype,
            "group_id": gid,
            "story_id": story_id,
            "sheet": sheet_name,
            "fields": fields,
        })
    return groups


def inspect_workbook(
    path: Path,
    only_sheet: str | None,
    cell_range: str | None,
    with_sections: bool = False,
) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)  # 읽기 전용 로드
    result = {
        "workbook": str(path),
        "format": {
            "story_id_cell": STORY_ID_CELL,
            "content_rows": f"{CONTENT_ROW_START}-{CONTENT_ROW_END}",
            "section_col": SECTION_COL,
            "content_col": CONTENT_COL,
            "source_sheets": {"group_a": GROUP_A_KEY, "group_b": GROUP_B_KEY},
        },
        "sheet_names": wb.sheetnames,
        "language_codes": {s: _lang_code(s) for s in wb.sheetnames},
        "sheets": {},
    }

    targets = [only_sheet] if only_sheet else wb.sheetnames
    for name in targets:
        if name not in wb.sheetnames:
            result["sheets"][name] = {"error": "해당 시트가 없습니다."}
            continue
        ws = wb[name]
        if cell_range:
            cells = []
            for row in ws[cell_range]:
                for cell in row:
                    if cell.value is not None:
                        cells.append({"coord": cell.coordinate, "value": str(cell.value)})
            result["sheets"][name] = {"cell_range": cell_range, "cells": cells}
        else:
            sheet_data = inspect_sheet(ws)
            if with_sections:
                # 기존 rows 는 유지하고 groups 만 추가 (backward compatibility)
                sheet_data["groups"] = build_groups(ws, sheet_data["story_id"], name)
            result["sheets"][name] = sheet_data
    return result


def _print_human(res: dict) -> None:
    print(f"📄 workbook: {res['workbook']}")
    fmt = res["format"]
    print(f"   format  : story_id={fmt['story_id_cell']}, "
          f"rows {fmt['content_rows']}, section=col{fmt['section_col']}, "
          f"content=col{fmt['content_col']}")
    print(f"   sheets  : {len(res['sheet_names'])} → {', '.join(res['sheet_names'])}")
    print("-" * 60)
    for name, data in res["sheets"].items():
        if "error" in data:
            print(f"[{name}] {data['error']}")
            continue
        if "cell_range" in data:
            print(f"[{name}] range {data['cell_range']}: {len(data['cells'])} non-empty")
            for c in data["cells"]:
                print(f"    {c['coord']}: {c['value'][:70]}")
            continue
        print(f"[{name}] story_id={data['story_id']} "
              f"populated={data['populated_rows']}")
        for row in data["rows"]:
            print(f"    row{row['row']:>2} [{row['section_code']}] "
                  f"({row['content_len']}자) {row['content_preview']}")
        if data.get("groups"):
            print(f"    ── groups ({len(data['groups'])}) ──")
            for g in data["groups"]:
                print(f"    ▸ {g['group_id']} ({g['group_type']})")
                for fname, f in g["fields"].items():
                    flag = " ⚠empty/x" if f["is_empty_or_placeholder"] else ""
                    preview = f["text"][:48] + ("…" if len(f["text"]) > 48 else "")
                    print(f"        {fname:11} {f['cell']} [{f['section_code']}]{flag}: {preview}")


def main():
    parser = argparse.ArgumentParser(description="SmartThings 워크북 분석 (읽기 전용)")
    parser.add_argument("workbook", help="분석할 .xlsx 경로")
    parser.add_argument("--sheet", default=None, help="특정 시트만 분석 (예: 'JA(일본)')")
    parser.add_argument("--cell-range", default=None, help="셀 범위 덤프 (예: C7:C28)")
    parser.add_argument("--sections", action="store_true",
                        help="story/section 단위로 title·description·disclaimer·button 그룹핑")
    parser.add_argument("--app-root", help="app repo 경로 명시 (Excel-only 에선 불필요)")
    parser.add_argument("--json", action="store_true", help="JSON 출력")
    args = parser.parse_args()

    path = Path(args.workbook).expanduser()
    if not path.is_file():
        msg = {"error": f"파일을 찾을 수 없습니다: {path}"}
        print(json.dumps(msg, ensure_ascii=False) if args.json else f"❌ {msg['error']}")
        sys.exit(2)

    res = inspect_workbook(path, args.sheet, args.cell_range, with_sections=args.sections)
    if args.json:
        print(json.dumps(res, ensure_ascii=False, indent=2))
    else:
        _print_human(res)


if __name__ == "__main__":
    main()
