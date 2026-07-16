#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""승인된 story 편집을 납품용 복사본에 반영하고 glossary 하이라이트를 복구한다.

``workbook_apply_edits.py``는 셀 값만 바꾸는 저수준 도구다. openpyxl 저장 뒤에는
편집하지 않은 시트의 rich-text 하이라이트도 풀릴 수 있으므로, 납품본은 반드시 이
스크립트처럼 명시된 delivery scope 전체를 재하이라이트한 산출물만 사용한다.
"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import re
from pathlib import Path
from types import SimpleNamespace

import openpyxl

import _app_pipeline as ap
from workbook_apply_edits import _load_edits, apply_edits
from workbook_highlight_glossary import _report_path_for, run_highlight


def _split_sheets(value: str) -> list[str]:
    sheets = [item.strip() for item in value.split(",") if item.strip()]
    if not sheets:
        raise ValueError("--delivery-sheets에 하나 이상의 시트를 지정해야 합니다.")
    if len(sheets) != len(set(sheets)):
        raise ValueError("--delivery-sheets에 중복된 시트가 있습니다.")
    return sheets


def _cell_text(value) -> str:
    return "" if value is None else str(value)


def _validate_delivery_scope(workbook: Path, edits: list[dict], delivery_sheets: list[str]) -> None:
    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=False)
    unknown = [sheet for sheet in delivery_sheets if sheet not in wb.sheetnames]
    if unknown:
        raise ValueError(f"워크북에 없는 delivery 시트: {', '.join(unknown)}")

    unsupported = [sheet for sheet in delivery_sheets if sheet not in ap.DEFAULT_SHEET_LANGS]
    if unsupported:
        raise ValueError(
            "표준 source-group 매핑이 없는 delivery 시트: "
            + ", ".join(unsupported)
            + ". 시트 매핑을 먼저 추가하거나 이번 delivery scope에서 제외하세요."
        )

    outside = [edit.get("sheet") for edit in edits if edit.get("sheet") not in delivery_sheets]
    if outside:
        raise ValueError(
            "편집 대상은 --delivery-sheets 안에 있어야 합니다: " + ", ".join(sorted(set(outside)))
        )


def _changed_cells(change_log: list[dict]) -> set[tuple[str, str]]:
    return {
        (change["sheet"], change["cell"])
        for change in change_log
        if change["old_value"] != change["new_value"]
    }


def _verify_values(source: Path, final: Path, expected: set[tuple[str, str]]) -> dict:
    before = openpyxl.load_workbook(source, data_only=False)
    after = openpyxl.load_workbook(final, data_only=False, rich_text=True)
    actual: set[tuple[str, str]] = set()

    for name in before.sheetnames:
        ws_before = before[name]
        ws_after = after[name]
        max_row = max(ws_before.max_row, ws_after.max_row)
        max_col = max(ws_before.max_column, ws_after.max_column)
        for row in ws_before.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                if _cell_text(cell.value) != _cell_text(ws_after[cell.coordinate].value):
                    actual.add((name, cell.coordinate))

    if actual != expected:
        unexpected = sorted(actual - expected)
        missing = sorted(expected - actual)
        raise RuntimeError(
            f"값 검증 실패: 예기치 않은 변경={unexpected or '-'}, 누락 변경={missing or '-'}"
        )

    return {"expected_value_changes": len(expected), "actual_value_changes": len(actual)}


def _verify_highlight_scope(output_data: str, delivery_sheets: list[str]) -> dict:
    """앱 로그에서 이번 납품 시트가 실제로 처리됐는지 확인한다."""
    completed = {
        sheet: int(count)
        for sheet, count in re.findall(r"\] (.+?) \| 처리 완료 셀: (\d+)", output_data or "")
    }
    missing = [sheet for sheet in delivery_sheets if sheet not in completed]
    if missing:
        raise RuntimeError("하이라이트 scope 검증 실패: 처리되지 않은 delivery 시트=" + ", ".join(missing))
    return {"completed_delivery_sheets": completed}


async def run_apply(args) -> dict:
    app_root = ap.bootstrap_project(args.app_root)
    ap.maybe_reexec_with_app_venv(app_root)

    source = Path(args.workbook).expanduser()
    glossary = Path(args.glossary).expanduser()
    if not source.is_file():
        raise FileNotFoundError(f"워크북을 찾을 수 없습니다: {source}")
    if not glossary.is_file():
        raise FileNotFoundError(f"용어집 CSV를 찾을 수 없습니다: {glossary}")

    delivery_sheets = _split_sheets(args.delivery_sheets)
    edits = _load_edits(args.edits)
    _validate_delivery_scope(source, edits, delivery_sheets)

    applied = apply_edits(source, edits)
    if applied["status"] != "ok":
        return applied

    highlight_args = SimpleNamespace(
        app_root=str(app_root),
        workbook=applied["revised"],
        glossary=str(glossary),
        sheets=",".join(delivery_sheets),
        cell_range=args.cell_range,
        sheet_langs=None,
        single_source=False,
        source_sheet="US(미국)",
        include_source_sheets=True,
        max_concurrency=args.max_concurrency,
        json=True,
        verbose=False,
    )
    highlighted = await run_highlight(highlight_args)
    if highlighted["status"] != "ok" or not highlighted.get("excel_path"):
        raise RuntimeError(f"하이라이트 재생성 실패: {highlighted.get('errors', [])}")

    report_path = _report_path_for(highlighted)
    if report_path:
        highlighted["report_path"] = report_path

    validation = _verify_values(source, Path(highlighted["excel_path"]), _changed_cells(applied["applied"]))
    highlight_validation = _verify_highlight_scope(highlighted.get("output_data", ""), delivery_sheets)
    summary = {
        "status": "ok",
        "source": str(source),
        "revised": applied["revised"],
        "final": highlighted["excel_path"],
        "change_log": applied["change_log"],
        "highlight_report": report_path,
        "glossary": str(glossary),
        "cell_range": args.cell_range,
        "delivery_sheets": delivery_sheets,
        "source_groups": highlighted.get("source_groups"),
        "value_validation": validation,
        "highlight_validation": highlight_validation,
    }
    manifest = Path(highlighted["excel_path"]).with_suffix(".delivery.json")
    temp = manifest.with_suffix(manifest.suffix + ".tmp")
    temp.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(temp, manifest)
    summary["delivery_manifest"] = str(manifest)
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(description="승인된 story 편집을 납품용 하이라이트 복사본으로 생성")
    parser.add_argument("workbook", help="원본 .xlsx 경로 (수정되지 않음)")
    parser.add_argument("edits", help="승인된 edits JSON 파일 또는 inline JSON")
    parser.add_argument("--delivery-sheets", required=True, help="이번 납품 대상 시트 CSV")
    parser.add_argument("--glossary", required=True, help="이번 납품에 적용할 glossary CSV")
    parser.add_argument("--cell-range", default="C7:C28", help="하이라이트 대상 셀 범위")
    parser.add_argument("--max-concurrency", type=int, default=10)
    parser.add_argument("--app-root", help="SmartThings app repo 경로")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()

    try:
        result = asyncio.run(run_apply(args))
    except Exception as exc:
        result = {"status": "error", "error": str(exc)}

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    elif result["status"] == "ok":
        print("✅ 납품용 복사본 생성 완료")
        print(f"   최종본: {result['final']}")
        print(f"   변경 검증: {result['value_validation']['actual_value_changes']}개 셀")
        print(f"   하이라이트 리포트: {result['highlight_report']}")
    else:
        print(f"❌ {result.get('error') or result.get('errors')}")

    if result["status"] != "ok":
        raise SystemExit(1)


if __name__ == "__main__":
    main()
