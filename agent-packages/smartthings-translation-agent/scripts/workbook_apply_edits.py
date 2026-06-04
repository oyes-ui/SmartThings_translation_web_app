#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
workbook_apply_edits.py — 승인된 편집을 워크북에 적용 (원본 불변)

⚠ 이 스크립트는 사용자가 명시적으로 승인한 뒤에만 실행해야 한다.

안전 정책:
  - 원본 파일은 절대 수정하지 않는다. 항상 타임스탬프가 붙은 '복사본'에 기록한다.
  - 출력은 .tmp 로 먼저 쓴 뒤 os.replace() 로 교체한다 (프로젝트 atomic write 규칙).
  - cell.value 만 갱신하여 서식(폰트/색/병합)을 보존한다 (openpyxl 기본 동작).

edits JSON 형식 (파일 경로 또는 inline 문자열):
  [
    {"sheet": "JA(일본)", "row": 10, "col": "C", "new_value": "新しいテキスト"},
    {"sheet": "DE(독일)", "cell": "C11", "new_value": "Neuer Text"}
  ]
  - col 은 문자("C") 또는 숫자(3) 모두 허용. row + col 또는 cell 둘 중 하나로 지정.

사용 예:
  python workbook_apply_edits.py story.xlsx edits.json
  python workbook_apply_edits.py story.xlsx '[{"sheet":"JA(일본)","cell":"C10","new_value":"..."}]'
  python workbook_apply_edits.py story.xlsx edits.json --json
"""

import os
import sys
import json
import argparse
from pathlib import Path
from datetime import datetime


def _bootstrap_project():
    """app_root 의 src/ 를 sys.path 에 추가 (없어도 비치명적).

    이 스크립트는 openpyxl 만으로 동작하므로(Level 1, Excel-only) app repo 가 없어도
    예외를 던지지 않는다. bootstrap.py(sibling)가 있으면 --app-root/env/config 를 반영한다.
    """
    try:
        sys.path.insert(0, str(Path(__file__).resolve().parent))
        import bootstrap as _bs
        app_root, _src = _bs.resolve_app_root(_bs.cli_app_root_from_argv())
        if app_root:
            src_dir = Path(app_root) / "src"
            if src_dir.is_dir() and str(src_dir) not in sys.path:
                sys.path.insert(0, str(src_dir))
            return Path(app_root)
    except Exception:
        pass

    here = Path(__file__).resolve()
    for parent in [here.parent, *here.parents]:
        if (parent / "src" / "translation_web_app").is_dir():
            src_dir = parent / "src"
            if str(src_dir) not in sys.path:
                sys.path.insert(0, str(src_dir))
            return parent
    return None


_bootstrap_project()

import openpyxl  # noqa: E402
from openpyxl.utils import get_column_letter, column_index_from_string  # noqa: E402
from openpyxl.utils.cell import coordinate_from_string  # noqa: E402


def _load_edits(edits_arg: str) -> list[dict]:
    """edits 인자를 파일 경로 또는 inline JSON 문자열로 해석."""
    p = Path(edits_arg).expanduser()
    if p.is_file():
        raw = p.read_text(encoding="utf-8")
    else:
        raw = edits_arg
    data = json.loads(raw)
    if not isinstance(data, list):
        raise ValueError("edits 는 객체들의 리스트여야 합니다.")
    return data


def _resolve_coord(edit: dict) -> str:
    """edit 항목을 셀 좌표 문자열('C10')로 정규화."""
    if "cell" in edit and edit["cell"]:
        return str(edit["cell"]).strip().upper()
    if "row" in edit and "col" in edit:
        col = edit["col"]
        col_letter = col if isinstance(col, str) and col.isalpha() \
            else get_column_letter(int(col))
        return f"{col_letter.upper()}{int(edit['row'])}"
    raise ValueError(f"편집 항목에 'cell' 또는 'row'+'col' 이 필요합니다: {edit}")


def apply_edits(src_path: Path, edits: list[dict]) -> dict:
    wb = openpyxl.load_workbook(src_path)  # 서식 유지 위해 data_only 미사용
    change_log = []
    errors = []

    for i, edit in enumerate(edits):
        try:
            sheet = edit["sheet"]
            new_value = edit["new_value"]
            coord = _resolve_coord(edit)
            if sheet not in wb.sheetnames:
                raise ValueError(f"시트 '{sheet}' 가 워크북에 없습니다.")
            ws = wb[sheet]
            old_value = ws[coord].value
            ws[coord].value = new_value  # cell.value 만 갱신 → 서식 보존
            change_log.append({
                "index": i,
                "sheet": sheet,
                "cell": coord,
                "old_value": None if old_value is None else str(old_value),
                "new_value": str(new_value),
            })
        except Exception as e:
            errors.append({"index": i, "edit": edit, "error": str(e)})

    if errors:
        # 하나라도 실패하면 파일을 쓰지 않고 중단 (부분 적용 방지)
        return {"status": "aborted", "errors": errors, "applied": []}

    # 타임스탬프 복사본 경로 (원본은 그대로 둔다)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = src_path.with_name(f"{src_path.stem}_revised_{ts}{src_path.suffix}")

    # atomic write: .tmp 로 먼저 저장 후 os.replace()
    tmp_path = out_path.with_suffix(out_path.suffix + ".tmp")
    wb.save(tmp_path)
    os.replace(tmp_path, out_path)

    # 변경 로그도 atomic write 로 함께 기록
    log_path = out_path.with_suffix(".changes.json")
    log_payload = {
        "source": str(src_path),
        "revised": str(out_path),
        "timestamp": ts,
        "changes": change_log,
    }
    log_tmp = log_path.with_suffix(log_path.suffix + ".tmp")
    log_tmp.write_text(
        json.dumps(log_payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    os.replace(log_tmp, log_path)

    return {
        "status": "ok",
        "source": str(src_path),
        "revised": str(out_path),
        "change_log": str(log_path),
        "applied": change_log,
    }


def main():
    parser = argparse.ArgumentParser(
        description="승인된 편집을 워크북 복사본에 적용 (원본 불변, atomic write)"
    )
    parser.add_argument("workbook", help="원본 .xlsx 경로 (수정되지 않음)")
    parser.add_argument("edits", help="edits JSON 파일 경로 또는 inline JSON 문자열")
    parser.add_argument("--app-root", help="app repo 경로 명시 (Excel-only 에선 불필요)")
    parser.add_argument("--json", action="store_true", help="JSON 출력")
    args = parser.parse_args()

    src_path = Path(args.workbook).expanduser()
    if not src_path.is_file():
        msg = {"error": f"파일을 찾을 수 없습니다: {src_path}"}
        print(json.dumps(msg, ensure_ascii=False) if args.json else f"❌ {msg['error']}")
        sys.exit(2)

    try:
        edits = _load_edits(args.edits)
    except (json.JSONDecodeError, ValueError) as e:
        msg = {"error": f"edits 파싱 실패: {e}"}
        print(json.dumps(msg, ensure_ascii=False) if args.json else f"❌ {msg['error']}")
        sys.exit(2)

    res = apply_edits(src_path, edits)

    if args.json:
        print(json.dumps(res, ensure_ascii=False, indent=2))
    else:
        if res["status"] == "ok":
            print(f"✅ 수정본 생성: {res['revised']}")
            print(f"   변경 로그  : {res['change_log']}")
            print(f"   원본 (불변): {res['source']}")
            for c in res["applied"]:
                print(f"   [{c['sheet']}!{c['cell']}] "
                      f"{c['old_value']!r} → {c['new_value']!r}")
        else:
            print(f"❌ 적용 중단 ({len(res['errors'])}건 오류, 파일 미생성):")
            for e in res["errors"]:
                print(f"   #{e['index']}: {e['error']}")
            sys.exit(1)


if __name__ == "__main__":
    main()
