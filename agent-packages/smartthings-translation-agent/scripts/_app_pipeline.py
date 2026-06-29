#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""_app_pipeline.py — 앱 파이프라인 래퍼 스크립트 공용 헬퍼.

workbook_highlight_glossary / workbook_translate / workbook_audit 가 공유하는
부트스트랩·venv 재실행·시트 매핑·source group·이벤트 요약 로직을 한 곳에 둔다.
하이라이트/번역/검수 로직은 여기서 재구현하지 않고 app repo의
translation_web_app.checker_service 를 호출한다.
"""

from __future__ import annotations

import json
import os
import sys
from pathlib import Path


# app 표준 시트 → 언어 매핑 (rag_db_builder / checker_service 와 동일 키 체계)
DEFAULT_SHEET_LANGS = {
    "KR(한국)": {"lang": "Korean", "code": "한국어"},
    "US(미국)": {"lang": "English", "code": "영어_미국"},
    "UK(영국)": {"lang": "English_UK", "code": "영어_영국"},
    "AU(호주)": {"lang": "English_AU", "code": "영어_호주"},
    "SG(싱가포르)": {"lang": "English_SG", "code": "영어_싱가폴"},
    "FR(프랑스)": {"lang": "French", "code": "프랑스어_프랑스"},
    "BE(벨기에)": {"lang": "French_BE", "code": "프랑스어_벨기에"},
    "CA(캐나다)": {"lang": "French_CA", "code": "프랑스어_캐나다"},
    "DE(독일)": {"lang": "German", "code": "독어_독일"},
    "IT(이탈리아)": {"lang": "Italian", "code": "이탈리아"},
    "ES(스페인)": {"lang": "Spanish", "code": "스페인어_스페인"},
    "NL(네덜란드)": {"lang": "Dutch", "code": "네덜란드어_네덜란드"},
    "SE(스웨덴)": {"lang": "Swedish", "code": "스웨덴"},
    "AE(아랍에메리트)": {"lang": "Arabic", "code": "아랍에미리트"},
    "PT(포르투갈)": {"lang": "European Portuguese", "code": "포르투갈_포르투갈"},
    "BR(브라질)": {"lang": "Brazilian Portuguese", "code": "포르투갈_브라질"},
    "RU(러시아)": {"lang": "Russian", "code": "러시아"},
    "TR(터키)": {"lang": "Turkish", "code": "터키"},
    "CN(중국)": {"lang": "Simplified Chinese", "code": "간체_중국어"},
    "TW(대만)": {"lang": "Traditional Chinese", "code": "번체_대만"},
    "JA(일본)": {"lang": "Japanese", "code": "일본"},
    "PL(폴란드)": {"lang": "Polish", "code": "폴란드"},
    "VN(베트남)": {"lang": "Vietnamese", "code": "베트남"},
    "TH(태국)": {"lang": "Thai", "code": "태국"},
    "ID(인도네시아)": {"lang": "Indonesian", "code": "인도네시아"},
}

# 기본 source grouping: KR 그룹(A) vs US 그룹(B)
GROUP_A_SOURCE = "KR(한국)"
GROUP_B_SOURCE = "US(미국)"
GROUP_A_TARGETS = {"US(미국)", "JA(일본)", "CN(중국)", "TW(대만)"}

# venv 재실행 1회 가드 (모든 파이프라인 래퍼 공용)
REEXEC_GUARD_ENV = "STTA_PIPELINE_REEXEC"


def bootstrap_project(cli_app_root: str | None) -> Path:
    """app_root 를 해석하고 src 를 sys.path 에 추가한다."""
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    import bootstrap as _bs

    app_root, _source = _bs.resolve_app_root(cli_app_root)
    if not app_root:
        raise RuntimeError(
            "app repo를 찾지 못했습니다. --app-root <경로>를 지정하거나 bootstrap.py --save를 먼저 실행하세요."
        )
    src_dir = app_root / "src"
    if not (src_dir / "translation_web_app").is_dir():
        raise RuntimeError(f"translation_web_app src를 찾지 못했습니다: {src_dir}")
    if str(src_dir) not in sys.path:
        sys.path.insert(0, str(src_dir))
    os.environ.setdefault("SMARTTHINGS_APP_ROOT", str(app_root))
    return app_root


def maybe_reexec_with_app_venv(app_root: Path) -> None:
    """앱 의존성이 현재 Python에 없을 때 app venv로 1회 재실행한다."""
    if os.environ.get(REEXEC_GUARD_ENV) == "1":
        return
    candidates = [
        app_root / "venv" / "bin" / "python",
        app_root / ".venv" / "bin" / "python",
        app_root / "venv" / "Scripts" / "python.exe",
        app_root / ".venv" / "Scripts" / "python.exe",
    ]
    for py in candidates:
        # venv/bin/python is often a symlink to the same system interpreter.
        # Compare the executable path itself, not the resolved target, so we still
        # enter the venv and get its site-packages.
        if py.exists() and Path(sys.executable) != py:
            env = os.environ.copy()
            env[REEXEC_GUARD_ENV] = "1"
            os.execve(str(py), [str(py), *sys.argv], env)


def split_sheets(raw: str | None) -> list[str] | None:
    if not raw:
        return None
    sheets = [x.strip() for x in raw.split(",") if x.strip()]
    return sheets or None


def load_sheet_langs(path_arg: str | None) -> dict:
    if not path_arg:
        return DEFAULT_SHEET_LANGS
    path = Path(path_arg).expanduser()
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError("sheet language mapping JSON은 객체여야 합니다.")
    return data


def default_source_groups(
    selected_sheets: list[str] | None,
    workbook_sheets: list[str],
    include_source_sheets: bool = False,
) -> list[dict]:
    """선택 시트를 KR/US 두 source group으로 자동 분배한다."""
    available = [s for s in workbook_sheets if s in DEFAULT_SHEET_LANGS]
    selected = set(selected_sheets) if selected_sheets else set(available)

    group_a = [s for s in available if s in GROUP_A_TARGETS and s in selected]
    group_b = [
        s for s in available
        if s in selected and s not in {GROUP_A_SOURCE, GROUP_B_SOURCE} and s not in GROUP_A_TARGETS
    ]
    if include_source_sheets:
        if GROUP_A_SOURCE in workbook_sheets and (group_a or not selected_sheets or GROUP_A_SOURCE in selected):
            group_a = [GROUP_A_SOURCE, *[s for s in group_a if s != GROUP_A_SOURCE]]
        if GROUP_B_SOURCE in workbook_sheets and (group_b or not selected_sheets or GROUP_B_SOURCE in selected):
            group_b = [GROUP_B_SOURCE, *[s for s in group_b if s != GROUP_B_SOURCE]]
    groups = []
    if GROUP_A_SOURCE in workbook_sheets and group_a:
        groups.append({"source_sheet": GROUP_A_SOURCE, "target_sheets": group_a})
    if GROUP_B_SOURCE in workbook_sheets and group_b:
        groups.append({"source_sheet": GROUP_B_SOURCE, "target_sheets": group_b})
    return groups


def event_summary(events: list[dict]) -> dict:
    """파이프라인 이벤트 스트림을 요약한다(앱 공통 type:complete 기준).

    status 판정:
      - error 이벤트가 있으면 'error'
      - error 는 없지만 complete 이벤트가 없으면 'incomplete' (조기 종료/누락 방어)
      - 정상 complete 이면 'ok'
    """
    logs = [
        e.get("message") or e.get("log")
        for e in events
        if e.get("type") == "log" and (e.get("message") or e.get("log"))
    ]
    complete = next((e for e in reversed(events) if e.get("type") == "complete"), None)
    errors = [e for e in events if e.get("type") == "error"]
    if errors:
        status = "error"
    elif complete is None:
        status = "incomplete"
    else:
        status = "ok"
    return {
        "status": status,
        "excel_path": complete.get("excel_path") if complete else None,
        "output_data": complete.get("output_data") if complete else None,
        "errors": errors,
        "log_count": len(logs),
        "logs": logs,
    }


def write_text_atomic(path, text: str) -> Path:
    """`.tmp` 에 먼저 쓰고 os.replace() 로 교체한다(이 repo 공통 atomic write 규칙)."""
    path = Path(path)
    tmp = path.with_name(path.name + ".tmp")
    tmp.write_text(text, encoding="utf-8")
    os.replace(tmp, path)
    return path


def workbook_sheetnames(workbook_path: Path) -> list[str]:
    import openpyxl

    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()
