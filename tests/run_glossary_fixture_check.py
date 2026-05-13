# -*- coding: utf-8 -*-
"""
Manual fixture check for disclaimer glossary prompts and Excel highlighting.

This script reuses tests/045_TEST.xlsx and tests/glossary_TEST_ver2.csv, copies
the workbook to /private/tmp, injects disclaimer rows, and runs the highlight-only
pipeline without calling any LLM APIs.
"""

from __future__ import annotations

import asyncio
import shutil
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from checker_service import TranslationChecker
from prompt_builder import PromptBuilder


SOURCE_XLSX = ROOT / "tests" / "045_TEST.xlsx"
GLOSSARY_CSV = ROOT / "tests" / "glossary_TEST_ver2.csv"
TMP_DIR = Path("/private/tmp")

TEST_ROWS = (30, 31)

SHEET_LANG_MAP = {
    "KR(한국)": {"lang": "Korean", "code": "ko_KR"},
    "US(미국)": {"lang": "English", "code": "en_US"},
    "CN(중국)": {"lang": "Simplified Chinese", "code": "zh_CN"},
    "JA(일본)": {"lang": "Japanese", "code": "ja_JP"},
    "BR(브라질)": {"lang": "Brazilian Portuguese", "code": "pt_BR"},
    "RU(러시아)": {"lang": "Russian", "code": "ru_RU"},
}

KOREAN_SOURCES = {
    30: (
        "//test_disclaimer_01",
        "[웰컴 에어 케어]는 [기기 > 에어컨 > 기기 제어 > 웰컴 에어 케어]에서 설정할 수 있습니다.",
    ),
    31: (
        "//test_disclaimer_02",
        "[부재 절전]은 [(스탠드형 모델) 기기 > 에어컨 > 기기 제어 > 동작 감지 > 부재 절전], "
        "[(벽걸이형 모델) 기기 > 에어컨 > 기기 제어 > 동작 감지 > 사람이 없을 때]에서 설정할 수 있습니다.",
    ),
}

ENGLISH_SOURCES = {
    30: (
        "//test_disclaimer_01",
        '[Welcome air care] can be set in "Device > Air Conditioner > Device control > Welcome air care."',
    ),
    31: (
        "//test_disclaimer_02",
        '[Power saving while away] can be set in "(Floor-standing model) Device > Air Conditioner > Device control > '
        'Motion detection > Power saving while away, (Wall-mount model) Device > Air Conditioner > Device control > '
        'Motion detection > While away."',
    ),
}

TARGET_TEXT = {
    "CN(중국)": {
        30: "[智能净化]可在\"Settings > 设备 > 空调 > 设备控制 > 智能净化\"中设置。",
        31: "[外出时省电]可在\"Settings > (落地式型号)设备 > 空调 > 设备控制 > 运动检测 > 外出时省电、"
        "(壁挂式型号)设备 > 空调 > 设备控制 > 运动检测 > 不在家时\"中设置。",
    },
    "JA(일본)": {
        30: "「Welcome air care」は\"Settings > デバイス > エアコン > デバイス操作 > Welcome air care\"で設定できます。",
        31: "「外出中に省電力」は\"Settings > (床置きモデル)デバイス > エアコン > デバイス操作 > Motion detection > 外出中に省電力、"
        "(壁掛けモデル)デバイス > エアコン > デバイス操作 > Motion detection > While away\"で設定できます。",
    },
    "BR(브라질)": {
        30: "[Cuidado do ar de boas-vindas] pode ser definido em \"Device > Air Conditioner > Contr. do aparelho > "
        "Cuidado do ar de boas-vindas.\"",
        31: "[Economia de energia] pode ser definida em \"(Floor-standing model) Device > Air Conditioner > Contr. do aparelho > "
        "Detecção de movimentos > Economia de energia, (Wall-mount model) Device > Air Conditioner > Contr. do aparelho > "
        "Detecção de movimentos > Estiver ausente.\"",
    },
    "RU(러시아)": {
        30: "[Предварительная обработка воздуха] можно настроить в \"Device > Air Conditioner > Управление > "
        "Предварительная обработка воздуха.\"",
        31: "[Энергосбер. в отсутствие] можно настроить в \"(Floor-standing model) Device > Air Conditioner > Управление > "
        "Обнаружение движения > Энергосбер. в отсутствие, (Wall-mount model) Device > Air Conditioner > Управление > "
        "Обнаружение движения > Когда в доме нет людей.\"",
    },
}


def assert_prompt_builder() -> list[str]:
    builder = PromptBuilder()
    results = []
    for sheet_name in ("CN(중국)", "JA(일본)", "BR(브라질)", "RU(러시아)"):
        info = SHEET_LANG_MAP[sheet_name]
        source_lang = "Korean" if sheet_name in {"CN(중국)", "JA(일본)"} else "English"
        prompt = builder.build_translation_prompt(
            target_lang=info["lang"],
            source_lang=source_lang,
            row_key="//test_disclaimer_01",
            glossary_context={"fixture": "fixture"},
            target_lang_code=info["code"],
        )
        expected_wrap = "Wrap glossary terms in '「' and '」'" if sheet_name == "JA(일본)" else "Wrap glossary terms in '[' and ']'"
        assert builder.get_glossary_context_mode("//test_disclaimer_01") == "disclaimer"
        assert expected_wrap in prompt, f"{sheet_name}: missing {expected_wrap}"
        assert "navigation paths" in prompt, f"{sheet_name}: missing navigation path exception"
        assert "double quotation marks" in prompt, f"{sheet_name}: missing path quote rule"
        assert "inside the closing quote" in prompt, f"{sheet_name}: missing intl period placement rule"
        assert "US English" not in prompt, f"{sheet_name}: should use concrete intl period placement rule"
        results.append(f"PASS prompt: {sheet_name} uses disclaimer wrap, path quote, and period placement rules")
    return results


def make_workbook_copy() -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = TMP_DIR / f"glossary_fixture_{timestamp}.xlsx"
    shutil.copy2(SOURCE_XLSX, dst)

    wb = openpyxl.load_workbook(dst)
    for row_idx, (row_key, text) in KOREAN_SOURCES.items():
        ws = wb["KR(한국)"]
        ws[f"B{row_idx}"] = row_key
        ws[f"C{row_idx}"] = text

    for row_idx, (row_key, text) in ENGLISH_SOURCES.items():
        ws = wb["US(미국)"]
        ws[f"B{row_idx}"] = row_key
        ws[f"C{row_idx}"] = text

    for sheet_name, rows in TARGET_TEXT.items():
        ws = wb[sheet_name]
        for row_idx, text in rows.items():
            source_row = KOREAN_SOURCES[row_idx] if sheet_name in {"CN(중국)", "JA(일본)"} else ENGLISH_SOURCES[row_idx]
            ws[f"B{row_idx}"] = source_row[0]
            ws[f"C{row_idx}"] = text

    wb.save(dst)
    return dst


async def run_highlight(source_path: Path, source_sheet: str, targets: list[str]) -> tuple[Path, list[str]]:
    checker = TranslationChecker()
    logs = []
    output_path = None
    async for event in checker.run_highlight_only_pipeline_generator(
        source_file_path=str(source_path),
        cell_range="C30:C31",
        sheet_lang_map=SHEET_LANG_MAP,
        glossary_file_path=str(GLOSSARY_CSV),
        selected_sheets=targets,
        source_sheet_name=source_sheet,
    ):
        if event.get("type") == "log":
            logs.append(event.get("message", ""))
        if event.get("type") == "complete":
            output_path = Path(event["excel_path"])
            logs.append(event.get("output_data", ""))

    if output_path is None:
        raise RuntimeError(f"highlight pipeline did not complete for {source_sheet}")
    return output_path, logs


def count_blue_rich_text(path: Path) -> dict[str, int]:
    wb = openpyxl.load_workbook(path, rich_text=True)
    counts = {}
    for sheet_name in ("CN(중국)", "JA(일본)", "BR(브라질)", "RU(러시아)"):
        total = 0
        ws = wb[sheet_name]
        for row_idx in TEST_ROWS:
            value = ws[f"C{row_idx}"].value
            if isinstance(value, CellRichText):
                for part in value:
                    if isinstance(part, TextBlock) and getattr(part.font, "color", None):
                        rgb = getattr(part.font.color, "rgb", None)
                        if rgb and str(rgb).upper().endswith("0000FF"):
                            total += 1
        counts[sheet_name] = total
    return counts


def plain_cell_text(value) -> str:
    if isinstance(value, CellRichText):
        return "".join(value.as_list())
    return str(value or "")


async def collect_direct_issues(path: Path) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(path, rich_text=True)
    issues = []

    runs = [
        ("KR(한국)", ["CN(중국)", "JA(일본)"]),
        ("US(미국)", ["BR(브라질)", "RU(러시아)"]),
    ]

    for source_sheet, target_sheets in runs:
        checker = TranslationChecker()
        await checker.load_glossary_from_file(GLOSSARY_CSV, SHEET_LANG_MAP[source_sheet]["code"])
        source_ws = wb[source_sheet]

        for target_sheet in target_sheets:
            target_ws = wb[target_sheet]
            target_info = SHEET_LANG_MAP[target_sheet]
            for row_idx in TEST_ROWS:
                source_text = plain_cell_text(source_ws[f"C{row_idx}"].value)
                target_text = plain_cell_text(target_ws[f"C{row_idx}"].value)
                row_key = plain_cell_text(source_ws[f"B{row_idx}"].value)
                checks = [
                    ("mismatch", checker._precheck_glossary_mismatch(source_text, target_text, target_info["code"])),
                    ("casing", checker._check_glossary_casing(source_text, target_text, target_info["code"])),
                    (
                        "bracket",
                        checker._check_glossary_brackets(
                            source_text,
                            target_text,
                            target_info["code"],
                            target_info["lang"],
                            row_key=row_key,
                        ),
                    ),
                ]
                for issue_type, messages in checks:
                    for message in messages:
                        issues.append(
                            {
                                "type": issue_type,
                                "sheet": target_sheet,
                                "cell": f"C{row_idx}",
                                "message": message,
                            }
                        )
    return issues


async def main() -> None:
    prompt_results = assert_prompt_builder()
    input_path = make_workbook_copy()

    first_output, korean_logs = await run_highlight(input_path, "KR(한국)", ["CN(중국)", "JA(일본)"])
    final_output, english_logs = await run_highlight(first_output, "US(미국)", ["BR(브라질)", "RU(러시아)"])

    _ = korean_logs + english_logs
    direct_issues = await collect_direct_issues(final_output)
    bracket_logs = [issue for issue in direct_issues if issue["type"] == "bracket"]
    mismatch_logs = [issue for issue in direct_issues if issue["type"] == "mismatch"]
    casing_logs = [issue for issue in direct_issues if issue["type"] == "casing"]
    sheet_highlights = count_blue_rich_text(final_output)

    missing_highlights = [sheet for sheet, count in sheet_highlights.items() if count == 0]
    if missing_highlights:
        raise AssertionError(f"Expected non-zero rich-text highlights for: {', '.join(missing_highlights)}")
    if mismatch_logs:
        details = "\n".join(f"{i['sheet']} {i['cell']}: {i['message']}" for i in mismatch_logs)
        raise AssertionError("[미적용] glossary mismatch logs found:\n" + details)
    if casing_logs:
        details = "\n".join(f"{i['sheet']} {i['cell']}: {i['message']}" for i in casing_logs)
        raise AssertionError("[대소문자] glossary casing logs found:\n" + details)

    grouped_brackets = defaultdict(list)
    for issue in bracket_logs:
        grouped_brackets[(issue["sheet"], issue["cell"])].append(issue["message"])

    print("=== Prompt checks ===")
    for result in prompt_results:
        print(result)
    print()
    print("=== Generated files ===")
    print(f"Input copy: {input_path}")
    print(f"Korean highlight output: {first_output}")
    print(f"Final highlight output: {final_output}")
    print()
    print("=== Rich text highlight counts ===")
    for sheet_name, count in sheet_highlights.items():
        print(f"{sheet_name}: {count}")
    print()
    print("=== Issue summary ===")
    print(f"[미적용] count: {len(mismatch_logs)}")
    print(f"[대소문자] count: {len(casing_logs)}")
    print(f"[괄호 오류] false-positive candidates: {len(bracket_logs)}")
    for (sheet_name, cell), lines in grouped_brackets.items():
        print(f"- {sheet_name} {cell}: {len(lines)}")
        for line in lines:
            print(f"  {line}")
    if not direct_issues:
        print("No glossary issue logs found.")


if __name__ == "__main__":
    asyncio.run(main())
