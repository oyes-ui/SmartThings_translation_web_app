from pathlib import Path

import openpyxl
from fastapi.testclient import TestClient

from translation_web_app.main import app
from translation_web_app.paths import SOURCE_WORKBOOK_TEMPLATE
from translation_web_app.services.text_workbook_service import create_text_source_workbook


def test_source_workbook_template_is_canonical():
    assert SOURCE_WORKBOOK_TEMPLATE.exists()
    assert SOURCE_WORKBOOK_TEMPLATE.name == "source_workbook_blank.xlsx"
    assert not (SOURCE_WORKBOOK_TEMPLATE.parent / "source_workbook_blank.xlsx.xlsx").exists()
    assert not (SOURCE_WORKBOOK_TEMPLATE.parent / "(CX Center) SmartThings_2.0_Story_Contents_.xlsx").exists()


def test_create_text_source_workbook_fills_structured_cells(tmp_path: Path):
    generated = create_text_source_workbook(
        source_sheet="KR(한국)",
        story_number=52,
        update_date="2026.06.02",
        story={"title": "Story title", "description": "Story description"},
        sections=[
            {"title": "S1 title", "description": "S1 desc", "disclaimer": "S1 disclaimer", "button": "S1 button"},
            {"title": "S2 title"},
            {},
            {"button": "S4 button"},
        ],
        output_dir=tmp_path,
    )

    wb = openpyxl.load_workbook(generated.path, data_only=False)
    try:
        for ws in wb.worksheets:
            assert ws["C2"].value == "2026.06.02"
            assert ws["C5"].value == "story_052"

        source_ws = wb["KR(한국)"]
        assert source_ws["B7"].value == '="//story_"&RIGHT($C$5, 3)&"_title"'
        assert source_ws["C7"].value == "Story title"
        assert source_ws["C8"].value == "Story description"
        assert source_ws["C10"].value == "S1 title"
        assert source_ws["C11"].value == "S1 desc"
        assert source_ws["C12"].value == "S1 disclaimer"
        assert source_ws["C13"].value == "S1 button"
        assert source_ws["C15"].value == "S2 title"
        assert source_ws["C25"].value is None
        assert source_ws["C28"].value == "S4 button"

        target_ws = wb["US(미국)"]
        assert target_ws["C7"].value is None
        assert target_ws["C28"].value is None
    finally:
        wb.close()


def test_create_text_source_workbook_rejects_more_than_four_sections(tmp_path: Path):
    try:
        create_text_source_workbook(
            source_sheet="KR(한국)",
            story_number="story_001",
            update_date=None,
            story={},
            sections=[{} for _ in range(5)],
            output_dir=tmp_path,
        )
    except ValueError as exc:
        assert "at most 4" in str(exc)
    else:
        raise AssertionError("Expected ValueError")


def test_text_workbook_router_validates_target_sheets():
    client = TestClient(app)
    response = client.post(
        "/api/text-workbooks/start",
        json={
            "source_sheet": "KR(한국)",
            "sheet_langs": {"KR(한국)": {"lang": "Korean", "code": "ko_KR"}},
            "story": {"title": "Hello"},
            "sections": [],
        },
    )
    assert response.status_code == 400
    assert "target sheet" in response.json()["detail"]

