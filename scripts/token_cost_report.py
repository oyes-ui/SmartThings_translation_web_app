#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Token & Cost Report — Story 015 (min) vs Story 006 (max)

Usage:
  # Phase 1: translation tokens only (no audit TXT needed)
  python scripts/token_cost_report.py

  # Phase 2: include audit output tokens (4 TXT files, one per story×workflow)
  python scripts/token_cost_report.py \
    --audit-txt-015-a outputs/audit_015_wf_a_ja.txt \
    --audit-txt-015-b outputs/audit_015_wf_b_fr.txt \
    --audit-txt-006-a outputs/audit_006_wf_a_ja.txt \
    --audit-txt-006-b outputs/audit_006_wf_b_fr.txt

  각 TXT는 해당 workflow의 대표 타겟 검수 output (JA=Workflow A 대표, FR=Workflow B 대표).
  같은 workflow 내 다른 타겟에는 동일 출력 토큰 수가 적용됩니다.
"""

import argparse
import asyncio
import json
import os
import re
import sys

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "src")
if SRC not in sys.path:
    sys.path.insert(0, SRC)

from translation_web_app.model_handler import ModelHandler
from translation_web_app.prompt_builder import PromptBuilder
from translation_web_app.checker_service import TranslationChecker
from translation_web_app.paths import EXCEL_DATA_DIR, UPLOAD_DIR

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

EXCEL_DIR = str(EXCEL_DATA_DIR)

STORY_015 = os.path.join(EXCEL_DIR, "(CX Center) SmartThings_2.0_Story_Contents_015_0227.xlsx")
STORY_006 = os.path.join(EXCEL_DIR, "(CX Center) SmartThings_2.0_Story_Contents_006_0319.xlsx")

# Most recent glossary (uploads/, May 19) — used in actual workflow
GLOSSARY_CSV = str(UPLOAD_DIR / "glossary_e2bc0367-f624-4d28-ab50-ce02e330b1e3.csv")

TRANSLATION_MODEL = "gemini-3-flash-preview"
AUDIT_MODEL       = "gpt-5.4-mini"

# USD per 1M tokens (공식 가격 기준 2026-05-19)
# Gemini 3 Flash Preview: ai.google.dev/pricing
# gpt-5.4-mini: platform.openai.com/pricing (Short context 기준)
PRICING = {
    TRANSLATION_MODEL: {"input": 0.50 / 1_000_000, "output": 3.00 / 1_000_000},
    AUDIT_MODEL:       {"input": 0.75 / 1_000_000, "output": 4.50 / 1_000_000},
}

# ---------------------------------------------------------------------------
# Language target tables
# target_lang strings must match PromptBuilder._LANGUAGE_RULE_LABELS keys
# ---------------------------------------------------------------------------

# Workflow A: Korean source → 4 CJK + US targets
KR_TO_4 = [
    ("US(미국)",    "en_US", "English_US",           "Korean"),
    ("CN(중국)",    "zh_CN", "Simplified Chinese",   "Korean"),
    ("TW(대만)",    "zh_TW", "Traditional Chinese",  "Korean"),
    ("JA(일본)",    "ja_JP", "Japanese",             "Korean"),
]

# Workflow B: English source → 20 other countries
EN_TO_20 = [
    ("UK(영국)",       "en_GB", "English_UK",             "English"),
    ("AU(호주)",       "en_AU", "English_AU",             "English"),
    ("SG(싱가포르)",   "en_SG", "English_SG",             "English"),
    ("FR(프랑스)",     "fr_FR", "French",                 "English"),
    ("BE(벨기에)",     "fr_BE", "French_BE",              "English"),
    ("CA(캐나다)",     "fr_CA", "French_CA",              "English"),
    ("DE(독일)",       "de_DE", "German",                 "English"),
    ("IT(이탈리아)",   "it_IT", "Italian",                "English"),
    ("ES(스페인)",     "es_ES", "Spanish",                "English"),
    ("NL(네덜란드)",   "nl_NL", "Dutch",                  "English"),
    ("SE(스웨덴)",     "sv_SE", "Swedish",                "English"),
    ("AE(아랍에메리트)","ar_AE","Arabic",                 "English"),
    ("PT(포르투갈)",   "pt_PT", "European Portuguese",    "English"),
    ("BR(브라질)",     "pt_BR", "Brazilian Portuguese",   "English"),
    ("RU(러시아)",     "ru_RU", "Russian",                "English"),
    ("TR(터키)",       "tr_TR", "Turkish",                "English"),
    ("PL(폴란드)",     "pl_PL", "Polish",                 "English"),
    ("VN(베트남)",     "vi_VN", "Vietnamese",             "English"),
    ("TH(태국)",       "th_TH", "Thai",                   "English"),
    ("ID(인도네시아)", "id_ID", "Indonesian",             "English"),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def extract_source(xlsx_path: str, sheet_name: str) -> dict:
    """Return {key: text} from C7:C28 (col B=key, col C=text)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    result = {}
    for r in range(7, 29):
        key  = ws.cell(r, 2).value
        text = ws.cell(r, 3).value
        if key and text:
            result[str(key)] = str(text)
    wb.close()
    return result


def build_translation_sys_text(
    pb: PromptBuilder, target_lang: str, source_lang: str, lang_code: str,
    glossary_context: dict | None = None,
) -> str:
    sections = pb.build_translation_prompt_sections(
        target_lang=target_lang,
        source_lang=source_lang,
        row_key="description",
        target_lang_code=lang_code,
        glossary_context=glossary_context or None,
    )
    return "\n\n".join(s["content"] for s in sections if s.get("active") and s.get("content"))


def build_translation_user_text(source_dict: dict, target_lang: str, glossary_dict: dict) -> str:
    pb = PromptBuilder()
    formatting = pb.build_input_formatting(target_lang, row_key="description")
    msg = {
        "context_key": "description",
        "source_text": source_dict,
        "target_language": target_lang,
        "glossary": glossary_dict,
        "formatting": formatting,
    }
    return json.dumps(msg, ensure_ascii=False)


def build_audit_sys_text(
    pb: PromptBuilder, target_lang: str, lang_code: str,
    glossary_context: dict | None = None,
) -> str:
    sections = pb.build_audit_prompt_sections(
        target_lang=target_lang,
        target_lang_code=lang_code,
        row_key="description",
        glossary_context=glossary_context or None,
    )
    return "\n\n".join(s["content"] for s in sections if s.get("active") and s.get("content"))


def build_audit_user_text(
    source_dict: dict, output_dict: dict, source_lang: str,
    target_lang: str, lang_code: str, glossary_dict: dict,
) -> str:
    msg = {
        "source":      {"lang": source_lang, "text": source_dict},
        "translation": {"lang": f"{target_lang}/{lang_code}", "text": output_dict},
        "glossary":    glossary_dict,
        "context_key": "description",
    }
    return json.dumps(msg, ensure_ascii=False)


def parse_audit_output_tokens(txt_path: str) -> int:
    """Extract LLM-generated text from audit TXT and count tokens with tiktoken.

    Extracts:
      - [상세 - 역번역] … up to [역번역 비활성화됨] 이전 내용 (활성 시만)
      - [상세 - AI 검수 결과] … up to [상세 - RAG Payload] or next [시트] (LLM 응답 전체)
    """
    with open(txt_path, encoding="utf-8") as f:
        content = f.read()

    llm_text = ""

    # [상세 - AI 검수 결과] 섹션 — [상세 - RAG Payload] 또는 다음 섹션 헤더 전까지
    for m in re.finditer(
        r'\[상세 - AI 검수 결과\](.*?)(?=\[상세 - RAG Payload\]|\[상세 - AI Payload\]|\[시트\]|$)',
        content, re.DOTALL
    ):
        chunk = m.group(1).strip()
        if chunk:
            llm_text += chunk + "\n"

    # [상세 - 역번역] 섹션 — 비활성화 메시지 제외
    for m in re.finditer(
        r'\[상세 - 역번역\](.*?)(?=\[상세 - |$)',
        content, re.DOTALL
    ):
        chunk = m.group(1).strip()
        if chunk and "[역번역 비활성화됨]" not in chunk:
            llm_text += chunk + "\n"

    if not llm_text.strip():
        print(f"  WARNING: no LLM sections found in {txt_path}")
        return 0

    import tiktoken
    enc = tiktoken.get_encoding("o200k_base")
    tok = len(enc.encode(llm_text))
    print(f"  audit output TXT parsed: {len(llm_text):,}자 → {tok:,}tok (tiktoken)")
    return tok


def cost(tok: int, model: str, direction: str) -> float:
    return tok * PRICING[model][direction]


# ---------------------------------------------------------------------------
# Core async calculation
# ---------------------------------------------------------------------------

async def compute_workflow(
    pb: PromptBuilder,
    mh: ModelHandler,
    tc: TranslationChecker,
    story_label: str,
    xlsx_path: str,
    source_sheet: str,
    source_lang_code: str,
    targets: list,          # list of (sheet_name, lang_code, target_lang, source_lang)
    workflow_label: str,
) -> list[dict]:
    """Compute token counts for one workflow (one source → N targets)."""

    source_dict = extract_source(xlsx_path, source_sheet)
    source_lang = targets[0][3]  # same for all rows in a workflow
    source_text_str = " ".join(source_dict.values())

    # Load glossary once for this workflow's source language
    await tc.load_glossary_from_file(GLOSSARY_CSV, source_lang_code)
    glossary_loaded = bool(tc.glossary)
    print(f"    [{workflow_label}] 용어집: {len(tc.glossary)}개 항목 로드됨" if glossary_loaded else
          f"    [{workflow_label}] 용어집 로드 실패")

    # Pre-build per-target glossary dicts (filtered by target lang + source text)
    glossary_per_target = {}
    for _, lcode, _, _ in targets:
        gd = tc._get_glossary_context_as_dict(
            target_lang_code=lcode,
            source_text=source_text_str,
            row_key="description",
        )
        glossary_per_target[lcode] = gd

    # User message differs per target (glossary dict varies by target lang)
    # but source text is the same — compute per-target user_text
    user_texts = {
        lcode: build_translation_user_text(source_dict, tgt_lang, glossary_per_target[lcode])
        for _, lcode, tgt_lang, _ in targets
    }
    # Count user tokens per target (may differ slightly due to glossary size)
    tasks_user = [mh.count_tokens(user_texts[lcode], TRANSLATION_MODEL) for _, lcode, _, _ in targets]

    rows = []
    tasks_sys = [
        mh.count_tokens(
            build_translation_sys_text(pb, tgt_lang, src_lang, lcode, glossary_per_target[lcode] or None),
            TRANSLATION_MODEL,
        )
        for _, lcode, tgt_lang, src_lang in targets
    ]

    # Extract output texts
    output_dicts = []
    for sheet_name, lcode, tgt_lang, _ in targets:
        output_dicts.append(extract_source(xlsx_path, sheet_name))

    tasks_out = [
        mh.count_tokens(json.dumps(od, ensure_ascii=False), TRANSLATION_MODEL)
        for od in output_dicts
    ]

    # Build audit sys texts and user texts
    tasks_au_sys = [
        mh.count_tokens(
            build_audit_sys_text(pb, tgt_lang, lcode, glossary_per_target[lcode] or None),
            AUDIT_MODEL,
        )
        for _, lcode, tgt_lang, _ in targets
    ]
    tasks_au_user = [
        mh.count_tokens(
            build_audit_user_text(source_dict, od, source_lang, tgt_lang, lcode, glossary_per_target[lcode]),
            AUDIT_MODEL,
        )
        for (_, lcode, tgt_lang, _), od in zip(targets, output_dicts)
    ]

    # Run all in parallel
    sys_results      = await asyncio.gather(*tasks_sys)
    user_results     = await asyncio.gather(*tasks_user)
    out_results      = await asyncio.gather(*tasks_out)
    au_sys_results   = await asyncio.gather(*tasks_au_sys)
    au_user_results  = await asyncio.gather(*tasks_au_user)

    for i, (sheet_name, lcode, tgt_lang, src_lang) in enumerate(targets):
        tr_sys_tok, sys_method = sys_results[i]
        tr_user_tok, _         = user_results[i]
        tr_out_tok, _          = out_results[i]
        au_sys_tok, _          = au_sys_results[i]
        au_user_tok, _         = au_user_results[i]

        tr_in_total  = tr_sys_tok + tr_user_tok
        au_in_total  = au_sys_tok + au_user_tok

        tr_cost = cost(tr_in_total, TRANSLATION_MODEL, "input") + cost(tr_out_tok, TRANSLATION_MODEL, "output")
        au_cost = cost(au_in_total, AUDIT_MODEL, "input")  # output added later from TXT

        rows.append({
            "story":         story_label,
            "workflow":      workflow_label,
            "sheet":         sheet_name,
            "lang_code":     lcode,
            "target_lang":   tgt_lang,
            "tr_sys_tok":    tr_sys_tok,
            "tr_user_tok":   tr_user_tok,
            "tr_out_tok":    tr_out_tok,
            "tr_in_total":   tr_in_total,
            "tr_cost":       tr_cost,
            "au_sys_tok":    au_sys_tok,
            "au_user_tok":   au_user_tok,
            "au_out_tok":    0,  # filled in Phase 2
            "au_in_total":   au_in_total,
            "au_cost":       au_cost,
            "sys_method":    sys_method,
        })

    return rows


async def main_async(
    audit_txt_015_a: str | None,
    audit_txt_015_b: str | None,
    audit_txt_006_a: str | None,
    audit_txt_006_b: str | None,
):
    pb = PromptBuilder()
    mh = ModelHandler()
    tc = TranslationChecker()

    all_rows = []

    print("\n=== Story 015 (최소, 436자 KR) ===")
    print("  Workflow A: KR→4")
    rows_015_a = await compute_workflow(pb, mh, tc, "015", STORY_015, "KR(한국)", "ko_KR", KR_TO_4, "A(KR→4)")
    all_rows.extend(rows_015_a)

    print("  Workflow B: EN→20")
    rows_015_b = await compute_workflow(pb, mh, tc, "015", STORY_015, "US(미국)", "en_US", EN_TO_20, "B(EN→20)")
    all_rows.extend(rows_015_b)

    print("\n=== Story 006 (최대, KR 1385자 / EN 3228자) ===")
    print("  Workflow A: KR→4")
    rows_006_a = await compute_workflow(pb, mh, tc, "006", STORY_006, "KR(한국)", "ko_KR", KR_TO_4, "A(KR→4)")
    all_rows.extend(rows_006_a)

    print("  Workflow B: EN→20")
    rows_006_b = await compute_workflow(pb, mh, tc, "006", STORY_006, "US(미국)", "en_US", EN_TO_20, "B(EN→20)")
    all_rows.extend(rows_006_b)

    # ---------------------------------------------------------------------------
    # Phase 2: audit output tokens — per story × workflow
    # TXT에 워크플로우 전체(N개국) 결과가 포함된 경우: 합계 ÷ N = 국가당 평균 적용
    # TXT가 단일 타겟 1개인 경우: N=1로 취급해 그대로 적용
    # ---------------------------------------------------------------------------
    WF_TARGET_COUNTS = {
        "A(KR→4)":  len(KR_TO_4),   # 4
        "B(EN→20)": len(EN_TO_20),  # 20
    }
    phase2_map = [
        ("015", "A(KR→4)",  audit_txt_015_a),
        ("015", "B(EN→20)", audit_txt_015_b),
        ("006", "A(KR→4)",  audit_txt_006_a),
        ("006", "B(EN→20)", audit_txt_006_b),
    ]
    for story_id, wf_label, txt_path in phase2_map:
        if not txt_path:
            continue
        print(f"\n[Phase 2] Story {story_id} {wf_label} audit TXT: {txt_path}")
        au_out_total = parse_audit_output_tokens(txt_path)
        n_targets = WF_TARGET_COUNTS[wf_label]
        au_out_per = round(au_out_total / n_targets)
        print(f"  → 전체 {au_out_total:,}tok ÷ {n_targets}개국 = 평균 {au_out_per:,}tok/타겟")
        for r in all_rows:
            if r["story"] == story_id and r["workflow"] == wf_label:
                r["au_out_tok"] = au_out_per
                r["au_cost"] += cost(au_out_per, AUDIT_MODEL, "output")

    audit_provided = any(p for _, _, p in phase2_map)
    print_report(all_rows, audit_provided)


# ---------------------------------------------------------------------------
# Report printer
# ---------------------------------------------------------------------------

def print_report(rows: list[dict], audit_provided: bool):
    HDR_TR = f"{'타겟':<22} {'sys':>5} {'user':>5} {'out':>5} {'in합계':>6}  {'번역비용($)':>11}"
    HDR_AU = f"{'타겟':<22} {'sys':>5} {'user':>5} {'out':>5} {'in합계':>6}  {'검수비용($)':>11}"
    SEP = "-" * 72

    for story_id in ["015", "006"]:
        story_rows = [r for r in rows if r["story"] == story_id]
        if not story_rows:
            continue

        src_info = "436자 KR" if story_id == "015" else "KR 1,385자 / EN 3,228자"
        print(f"\n{'='*72}")
        print(f"  STORY {story_id}  ({src_info}, 24개국)")
        print(f"{'='*72}")

        for wf_label in sorted(set(r["workflow"] for r in story_rows)):
            wf_rows = [r for r in story_rows if r["workflow"] == wf_label]
            print(f"\n  [번역] {wf_label}  (model: {TRANSLATION_MODEL})")
            print("  " + HDR_TR)
            print("  " + SEP)
            tr_sub = 0.0
            for r in wf_rows:
                c = r["tr_cost"]
                tr_sub += c
                print(f"  {r['lang_code']+'('+r['sheet'].split('(')[0]+')':.<22}"
                      f" {r['tr_sys_tok']:>5,} {r['tr_user_tok']:>5,} {r['tr_out_tok']:>5,} "
                      f"{r['tr_in_total']:>6,}  ${c:>10.6f}")
            print("  " + SEP)
            print(f"  {'소계':>22}                              ${tr_sub:>10.6f}")

            print(f"\n  [검수] {wf_label}  (model: {AUDIT_MODEL})")
            print("  " + HDR_AU)
            print("  " + SEP)
            au_sub = 0.0
            out_note = ""
            for r in wf_rows:
                c = r["au_cost"]
                au_sub += c
                out_str = str(r["au_out_tok"]) if r["au_out_tok"] else "TBD"
                print(f"  {r['lang_code']+'('+r['sheet'].split('(')[0]+')':.<22}"
                      f" {r['au_sys_tok']:>5,} {r['au_user_tok']:>5,} {out_str:>5} "
                      f"{r['au_in_total']:>6,}  ${c:>10.6f}")
            if not audit_provided:
                out_note = "  ※ 검수 output tok: --audit-txt-XXX-a/b 인수로 TXT 제공 시 추가"
            print("  " + SEP)
            print(f"  {'소계':>22}                              ${au_sub:>10.6f}")
            if out_note:
                print(out_note)

        # Story totals
        tr_total = sum(r["tr_cost"] for r in story_rows)
        au_total = sum(r["au_cost"] for r in story_rows)
        print(f"\n  {'─'*40}")
        print(f"  STORY {story_id} 번역 합계:  ${tr_total:.6f}  "
              f"({sum(r['tr_in_total'] for r in story_rows):,} in-tok / "
              f"{sum(r['tr_out_tok'] for r in story_rows):,} out-tok)")
        print(f"  STORY {story_id} 검수 합계:  ${au_total:.6f}  "
              f"({sum(r['au_in_total'] for r in story_rows):,} in-tok / "
              f"{sum(r['au_out_tok'] for r in story_rows):,} out-tok)")
        print(f"  STORY {story_id} 전체 합계:  ${tr_total + au_total:.6f}")

    # Summary
    print(f"\n{'='*72}")
    print("  최종 요약 (1건 스토리 처리 기준)")
    print(f"{'='*72}")
    print(f"  {'스토리':<8} {'번역($)':>12} {'검수($)':>12} {'합계($)':>12}")
    print(f"  {'-'*48}")
    for story_id, src_chars in [("015", 436), ("006", 1385)]:
        story_rows = [r for r in rows if r["story"] == story_id]
        tr = sum(r["tr_cost"] for r in story_rows)
        au = sum(r["au_cost"] for r in story_rows)
        total = tr + au
        n = len(story_rows)
        print(f"  {story_id+'('+str(src_chars)+'자)':.<8} ${tr:>11.6f} ${au:>11.6f} ${total:>11.6f}  ({n}개국)")

    print(f"\n  가격 기준:")
    for model, p in PRICING.items():
        print(f"    {model}: input ${p['input']*1e6:.3f}/M tok, output ${p['output']*1e6:.3f}/M tok")
    print("  ※ 검수 output 포함 여부: --audit-txt-015/006 인수 제공 시 반영됨")


# ---------------------------------------------------------------------------
# Entry
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Token & cost report for Story 015 / 006")
    parser.add_argument("--audit-txt-015-a", default=None,
                        help="Story 015 Workflow A 대표 검수 TXT (e.g. JA sheet)")
    parser.add_argument("--audit-txt-015-b", default=None,
                        help="Story 015 Workflow B 대표 검수 TXT (e.g. FR sheet)")
    parser.add_argument("--audit-txt-006-a", default=None,
                        help="Story 006 Workflow A 대표 검수 TXT (e.g. JA sheet)")
    parser.add_argument("--audit-txt-006-b", default=None,
                        help="Story 006 Workflow B 대표 검수 TXT (e.g. FR sheet)")
    args = parser.parse_args()

    asyncio.run(main_async(
        args.audit_txt_015_a,
        args.audit_txt_015_b,
        args.audit_txt_006_a,
        args.audit_txt_006_b,
    ))


if __name__ == "__main__":
    main()
