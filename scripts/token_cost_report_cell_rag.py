#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Cell-level token & cost report with RAG always enabled.

This mirrors the integrated pipeline more closely than token_cost_report.py:
one translation call and one audit call per source cell x target sheet.
"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import re
import sys
from collections import defaultdict

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

from checker_service import TranslationChecker
from model_handler import ModelHandler
from prompt_builder import PromptBuilder


EXCEL_DIR = os.path.join(ROOT, "@translation_data", "@excel")
STORY_015 = os.path.join(EXCEL_DIR, "(CX Center) SmartThings_2.0_Story_Contents_015_0227.xlsx")
STORY_006 = os.path.join(EXCEL_DIR, "(CX Center) SmartThings_2.0_Story_Contents_006_0319.xlsx")
GLOSSARY_CSV = os.path.join(ROOT, "uploads", "glossary_e2bc0367-f624-4d28-ab50-ce02e330b1e3.csv")

TRANSLATION_MODEL = "gemini-3.5-flash"
AUDIT_MODEL = "gpt-5.4-mini"

KR_TO_4 = [
    ("US(미국)", "en_US", "English_US", "Korean"),
    ("CN(중국)", "zh_CN", "Simplified Chinese", "Korean"),
    ("TW(대만)", "zh_TW", "Traditional Chinese", "Korean"),
    ("JA(일본)", "ja_JP", "Japanese", "Korean"),
]

EN_TO_20 = [
    ("UK(영국)", "en_GB", "English_UK", "English"),
    ("AU(호주)", "en_AU", "English_AU", "English"),
    ("SG(싱가포르)", "en_SG", "English_SG", "English"),
    ("FR(프랑스)", "fr_FR", "French", "English"),
    ("BE(벨기에)", "fr_BE", "French_BE", "English"),
    ("CA(캐나다)", "fr_CA", "French_CA", "English"),
    ("DE(독일)", "de_DE", "German", "English"),
    ("IT(이탈리아)", "it_IT", "Italian", "English"),
    ("ES(스페인)", "es_ES", "Spanish",    "English"),
    ("NL(네덜란드)", "nl_NL", "Dutch", "English"),
    ("SE(스웨덴)", "sv_SE", "Swedish", "English"),
    ("AE(아랍에메리트)", "ar_AE", "Arabic", "English"),
    ("PT(포르투갈)", "pt_PT", "European Portuguese", "English"),
    ("BR(브라질)", "pt_BR", "Brazilian Portuguese", "English"),
    ("RU(러시아)", "ru_RU", "Russian", "English"),
    ("TR(터키)", "tr_TR", "Turkish", "English"),
    ("PL(폴란드)", "pl_PL", "Polish", "English"),
    ("VN(베트남)", "vi_VN", "Vietnamese", "English"),
    ("TH(태국)", "th_TH", "Thai", "English"),
    ("ID(인도네시아)", "id_ID", "Indonesian", "English"),
]

PRICING = {
    "gemini-3.5-flash": {"input": 1.50 / 1_000_000, "output": 9.00 / 1_000_000},
    "gpt-5.4-mini": {"input": 0.75 / 1_000_000, "output": 4.50 / 1_000_000},
}


def cost(tokens: int, model: str, direction: str) -> float:
    return tokens * PRICING[model][direction]


def extract_cells(xlsx_path: str, sheet_name: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    rows = []
    for r in range(7, 29):
        key = ws.cell(r, 2).value
        text = ws.cell(r, 3).value
        if key and text and str(text).strip().lower() != "x":
            rows.append({"row": r, "row_key": str(key), "text": str(text)})
    wb.close()
    return rows


def extract_target_text(xlsx_path: str, sheet_name: str, row: int) -> str:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    value = ws.cell(row, 3).value
    wb.close()
    return str(value or "")


def build_translation_user_text(source_text: str, target_lang: str, glossary_dict: dict | None, row_key: str) -> str:
    pb = PromptBuilder()
    msg = {
        "context_key": row_key,
        "source_text": source_text,
        "target_language": target_lang,
        "glossary": glossary_dict or {},
        "formatting": pb.build_input_formatting(target_lang, row_key=row_key),
    }
    return json.dumps(msg, ensure_ascii=False)


def build_audit_user_text(
    source_text: str,
    target_text: str,
    source_lang: str,
    target_lang: str,
    lang_code: str,
    glossary_dict: dict | None,
    row_key: str,
) -> str:
    msg = {
        "source": {"lang": source_lang, "text": source_text},
        "translation": {"lang": f"{target_lang}/{lang_code}", "text": target_text},
        "glossary": glossary_dict or {},
        "context_key": row_key,
    }
    return "당신은 전문 번역 검수 전문가입니다. 아래 JSON 데이터를 분석하여 상세한 검수 결과를 반환하세요.\n\n[Input Data]\n" + json.dumps(msg, ensure_ascii=False)


def build_translation_system(
    pb: PromptBuilder,
    target_lang: str,
    source_lang: str,
    row_key: str,
    glossary_dict: dict | None,
    rag_context: str,
) -> str:
    return pb.build_translation_prompt(
        target_lang=target_lang,
        source_lang=source_lang,
        bx_style_on=False,
        glossary_context=glossary_dict,
        rag_context=rag_context,
        row_key=row_key,
    )


def build_audit_system(
    pb: PromptBuilder,
    target_lang: str,
    lang_code: str,
    row_key: str,
    glossary_dict: dict | None,
) -> str:
    return pb.build_audit_prompt(
        source_lang="",
        target_lang=target_lang,
        target_lang_code=lang_code,
        glossary_context=glossary_dict,
        row_key=row_key,
    )


def parse_audit_output_tokens(txt_path: str) -> int:
    with open(txt_path, encoding="utf-8") as f:
        content = f.read()
    llm_text = ""
    for m in re.finditer(
        r"\[상세 - AI 검수 결과\](.*?)(?=\[상세 - RAG Payload\]|\[상세 - AI Payload\]|\[시트\]|$)",
        content,
        re.DOTALL,
    ):
        chunk = m.group(1).strip()
        if chunk:
            llm_text += chunk + "\n"
    for m in re.finditer(r"\[상세 - 역번역\](.*?)(?=\[상세 - |$)", content, re.DOTALL):
        chunk = m.group(1).strip()
        if chunk and "[역번역 비활성화됨]" not in chunk:
            llm_text += chunk + "\n"
    if not llm_text.strip():
        return 0
    import tiktoken

    return len(tiktoken.get_encoding("o200k_base").encode(llm_text))


async def count_many(mh: ModelHandler, items: list[tuple[str, str]]) -> list[int]:
    results = await asyncio.gather(*(mh.count_tokens(text, model) for text, model in items))
    return [tok for tok, _ in results]


async def compute_workflow(
    story: str,
    xlsx_path: str,
    source_sheet: str,
    source_lang_code: str,
    targets: list[tuple[str, str, str, str]],
    workflow: str,
    translation_model: str = TRANSLATION_MODEL,
    audit_model: str = AUDIT_MODEL,
) -> list[dict]:
    pb = PromptBuilder()
    mh = ModelHandler()
    checker = TranslationChecker()
    await checker.load_glossary_from_file(GLOSSARY_CSV, source_lang_code)
    cells = extract_cells(xlsx_path, source_sheet)

    rows = []
    token_jobs: list[tuple[str, str]] = []
    row_refs = []

    for sheet_name, lang_code, target_lang, source_lang in targets:
        for cell in cells:
            source_text = cell["text"]
            row_key = cell["row_key"]
            glossary = checker._get_glossary_context_as_dict(
                lang_code,
                source_text=source_text,
                skip_deactivated=True,
                row_key=row_key,
            )
            glossary = glossary or None
            rag_context = ""
            if checker.rag_retriever and checker.rag_retriever.is_available():
                rag_context = checker.rag_retriever.format_for_prompt(
                    source_text,
                    sheet_name,
                    source_lang=source_lang,
                    n_results=2,
                    identity_match_enabled=True,
                )
            target_text = extract_target_text(xlsx_path, sheet_name, cell["row"])

            tr_sys = build_translation_system(pb, target_lang, source_lang, row_key, glossary, rag_context)
            tr_user = build_translation_user_text(source_text, target_lang, glossary, row_key)
            au_sys = build_audit_system(pb, target_lang, lang_code, row_key, glossary)
            au_user = build_audit_user_text(source_text, target_text, source_lang, target_lang, lang_code, glossary, row_key)

            base = {
                "story": story,
                "workflow": workflow,
                "sheet": sheet_name,
                "lang_code": lang_code,
                "row": cell["row"],
                "rag_chars": len(rag_context),
                "rag_on": bool(rag_context),
            }
            row_refs.append(base)
            token_jobs.extend(
                [
                    (tr_sys, translation_model),
                    (tr_user, translation_model),
                    (target_text, translation_model),
                    (au_sys, audit_model),
                    (au_user, audit_model),
                ]
            )

    counts = await count_many(mh, token_jobs)
    for i, base in enumerate(row_refs):
        offset = i * 5
        tr_sys, tr_user, tr_out, au_sys, au_user = counts[offset : offset + 5]
        row = dict(base)
        row.update(
            {
                "tr_sys_tok": tr_sys,
                "tr_user_tok": tr_user,
                "tr_out_tok": tr_out,
                "tr_in_total": tr_sys + tr_user,
                "au_sys_tok": au_sys,
                "au_user_tok": au_user,
                "au_in_total": au_sys + au_user,
                "au_out_tok": 0,
            }
        )
        rows.append(row)

    return rows


def apply_audit_outputs(rows: list[dict], args: argparse.Namespace) -> None:
    phase2 = [
        ("015", "A(KR→4)", args.audit_txt_015_a),
        ("015", "B(EN→20)", args.audit_txt_015_b),
        ("006", "A(KR→4)", args.audit_txt_006_a),
        ("006", "B(EN→20)", args.audit_txt_006_b),
    ]
    for story, workflow, path in phase2:
        if not path:
            continue
        matching = [r for r in rows if r["story"] == story and r["workflow"] == workflow]
        if not matching:
            continue
        total = parse_audit_output_tokens(path)
        per_call = round(total / len(matching))
        for row in matching:
            row["au_out_tok"] = per_call


def summarize(rows: list[dict], translation_model: str = TRANSLATION_MODEL, audit_model: str = AUDIT_MODEL) -> dict:
    summary = {}
    for story in sorted({r["story"] for r in rows}):
        story_rows = [r for r in rows if r["story"] == story]
        tr_in = sum(r["tr_in_total"] for r in story_rows)
        tr_out = sum(r["tr_out_tok"] for r in story_rows)
        au_in = sum(r["au_in_total"] for r in story_rows)
        au_out = sum(r["au_out_tok"] for r in story_rows)
        tr_cost = cost(tr_in, translation_model, "input") + cost(tr_out, translation_model, "output")
        au_cost = cost(au_in, audit_model, "input") + cost(au_out, audit_model, "output")
        summary[story] = {
            "calls": len(story_rows),
            "rag_calls": sum(1 for r in story_rows if r["rag_on"]),
            "rag_chars": sum(r["rag_chars"] for r in story_rows),
            "tr_in": tr_in,
            "tr_out": tr_out,
            "au_in": au_in,
            "au_out": au_out,
            "tr_cost": tr_cost,
            "au_cost": au_cost,
            "total_cost": tr_cost + au_cost,
        }
    return summary


def print_summary(summary: dict, translation_model: str = TRANSLATION_MODEL, audit_model: str = AUDIT_MODEL) -> None:
    krw_rate = 1501
    print("\n=== Cell-level RAG ON Cost Summary ===")
    if translation_model == audit_model:
        print(f"Model: {translation_model} (번역+검수 단일 모델)")
    else:
        print(f"Translation: {translation_model}  |  Audit: {audit_model}")
    print("Rate: 1 USD = 1,501 KRW")
    print()
    header = (
        f"{'Story':<8} {'calls':>6} {'RAG':>7} {'tr_in':>10} {'tr_out':>9} "
        f"{'au_in':>10} {'au_out':>10} {'tr_cost($)':>11} {'au_cost($)':>10} {'Total($)':>10} {'Total(KRW)':>11}"
    )
    print(header)
    print("-" * len(header))
    for story, data in summary.items():
        print(
            f"{story:<8} {data['calls']:>6,} {data['rag_calls']:>7,} "
            f"{data['tr_in']:>10,} {data['tr_out']:>9,} {data['au_in']:>10,} {data['au_out']:>10,} "
            f"{data['tr_cost']:>11.6f} {data['au_cost']:>10.6f} {data['total_cost']:>10.6f} "
            f"{data['total_cost'] * krw_rate:>11,.0f}"
        )


async def main_async(args: argparse.Namespace) -> None:
    tm, am = args.translation_model, args.audit_model
    all_rows = []
    all_rows.extend(await compute_workflow("015", STORY_015, "KR(한국)", "ko_KR", KR_TO_4, "A(KR→4)", tm, am))
    all_rows.extend(await compute_workflow("015", STORY_015, "US(미국)", "en_US", EN_TO_20, "B(EN→20)", tm, am))
    all_rows.extend(await compute_workflow("006", STORY_006, "KR(한국)", "ko_KR", KR_TO_4, "A(KR→4)", tm, am))
    all_rows.extend(await compute_workflow("006", STORY_006, "US(미국)", "en_US", EN_TO_20, "B(EN→20)", tm, am))
    apply_audit_outputs(all_rows, args)
    print_summary(summarize(all_rows, tm, am), tm, am)


def main() -> None:
    parser = argparse.ArgumentParser(description="Cell-level RAG-on token cost report")
    parser.add_argument("--audit-txt-015-a")
    parser.add_argument("--audit-txt-015-b")
    parser.add_argument("--audit-txt-006-a")
    parser.add_argument("--audit-txt-006-b")
    parser.add_argument(
        "--translation-model", default=TRANSLATION_MODEL,
        choices=list(PRICING), dest="translation_model",
    )
    parser.add_argument(
        "--audit-model", default=AUDIT_MODEL,
        choices=list(PRICING), dest="audit_model",
    )
    args = parser.parse_args()
    asyncio.run(main_async(args))


if __name__ == "__main__":
    main()
