# -*- coding: utf-8 -*-
"""
rag_db_builder.py
SmartThings 번역 RAG DB 구축 모듈

사용법:
  python rag_db_builder.py --pilot            # story_001 파일 1개만 빌드 후 보고
  python rag_db_builder.py --build-all        # 전체 파일 빌드 (이미 처리된 파일 skip)
  python rag_db_builder.py --update-story story_001   # 특정 스토리 증분 업데이트
  python rag_db_builder.py --status           # DB 현황 조회
"""

import os
import sys
import sqlite3
import json
import glob
import argparse
import asyncio
from datetime import datetime
from pathlib import Path

import openpyxl
from google import genai
from google.genai import types
from dotenv import load_dotenv
import chromadb

load_dotenv()

# ─── 경로 설정 ────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
EXCEL_DIR = BASE_DIR / "@translation_data" / "@excel"
RAG_DB_DIR = BASE_DIR / "rag_db"
CHROMA_DIR = RAG_DB_DIR / "chroma"
SQLITE_PATH = RAG_DB_DIR / "rag_store.db"

RAG_DB_DIR.mkdir(parents=True, exist_ok=True)
CHROMA_DIR.mkdir(parents=True, exist_ok=True)

# ChromaDB 컬렉션명
COLLECTION_KR = "smartthings_kr_source"  # Group A용 (KR 기준)
COLLECTION_US = "smartthings_us_source"  # Group B용 (US 기준)

# ─── Source 분기 정의 ─────────────────────────────────────────────────────────
# Group A: KR(한국) 시트를 Source로 사용하는 언어
GROUP_A_SOURCE_SHEETS = {"JA(일본)", "CN(중국)", "TW(대만)"}
GROUP_A_SHEETS = GROUP_A_SOURCE_SHEETS
GROUP_A_KEY = "KR(한국)"   # Group A의 Source 시트명
GROUP_B_KEY = "US(미국)"   # Group B의 Source 시트명

# 임베딩 모델
EMBEDDING_MODEL = "gemini-embedding-001"
EMBED_BATCH_SIZE = 50

# 파싱 범위
STORY_ID_CELL = "C5"
CONTENT_ROW_START = 7
CONTENT_ROW_END = 28
SECTION_COL = 2   # B열
CONTENT_COL = 3   # C열

# ─── 정규화 & 매핑 유틸리티 ────────────────────────────────────────────────────────
def normalize_text(text: str) -> str:
    """
    RAG 검색의 정확도를 위해 텍스트를 정규화합니다.
    - Unicode 정규화 (NFKC)
    - 공백 압축 및 양끝 공백 제거
    - 스마트 따옴표를 일반 따옴표로 변환
    - 전각 문자를 반각으로 변환 (NFKC에서 대부분 처리됨)
    """
    if not text:
        return ""
    import unicodedata
    import re
    # 1. Unicode 정규화
    text = unicodedata.normalize('NFKC', text)
    # 2. 스마트 따옴표 공통화
    text = text.replace('“', '"').replace('”', '"').replace('‘', "'").replace('’', "'")
    # 3. 연속된 공백 서포트 (하나로 압축)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def get_collection_name(source_lang: str) -> str:
    """
    source_lang에 따라 검색/저장할 ChromaDB 컬렉션 이름을 반환합니다.
    기본적으로 'Korean'일 경우 KR 컬렉션을, 그 외(English 등)는 US 컬렉션을 사용합니다.
    """
    if not source_lang:
        return COLLECTION_US
    sl = source_lang.lower()
    if "korea" in sl or "ko" == sl:
        return COLLECTION_KR
    return COLLECTION_US


# ─── Gemini 임베딩 클라이언트 ──────────────────────────────────────────────────
def get_gemini_client():
    api_key = os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY")
    if not api_key:
        raise ValueError("❌ GEMINI_API_KEY 또는 GOOGLE_API_KEY 환경변수가 설정되지 않았습니다.")
    return genai.Client(api_key=api_key)


def embed_texts(client, texts: list[str]) -> list[list[float]]:
    """텍스트 리스트를 Gemini 임베딩으로 변환 (50개씩 배치 처리)"""
    if not texts:
        return []
    all_embeddings = []
    for i in range(0, len(texts), EMBED_BATCH_SIZE):
        batch = texts[i:i + EMBED_BATCH_SIZE]
        response = client.models.embed_content(
            model=f"models/{EMBEDDING_MODEL}",
            contents=batch,
            config=types.EmbedContentConfig(task_type="RETRIEVAL_DOCUMENT")
        )
        all_embeddings.extend([e.values for e in response.embeddings])
    return all_embeddings


# ─── ChromaDB 초기화 ──────────────────────────────────────────────────────────
def get_chroma_collections():
    chroma_client = chromadb.PersistentClient(path=str(CHROMA_DIR))
    col_kr = chroma_client.get_or_create_collection(COLLECTION_KR)
    col_us = chroma_client.get_or_create_collection(COLLECTION_US)
    return chroma_client, col_kr, col_us


# ─── SQLite 초기화 ────────────────────────────────────────────────────────────
def get_sqlite_conn():
    conn = sqlite3.connect(str(SQLITE_PATH))
    conn.row_factory = sqlite3.Row
    conn.execute("""
        CREATE TABLE IF NOT EXISTS rag_pairs (
            id TEXT PRIMARY KEY,
            story_id TEXT,
            section_code TEXT,
            source_group TEXT,       -- 'kr' or 'us'
            source_text TEXT,
            target_lang TEXT,
            target_sheet TEXT,
            target_text TEXT,
            original_file TEXT,
            created_at TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS processed_files (
            filename TEXT PRIMARY KEY,
            story_id TEXT,
            mtime TEXT,
            record_count INTEGER,
            processed_at TEXT
        )
    """)
    conn.commit()
    return conn


# ─── 엑셀 파싱 ───────────────────────────────────────────────────────────────
def parse_excel_file(filepath: Path) -> dict:
    """
    엑셀 파일 파싱:
    - C5: story_id
    - B7:B28: section_code
    - C7:C28: content text (각 시트별)
    
    반환 구조:
    {
      "story_id": "story_001",
      "filename": "파일명.xlsx",
      "sheets": {
        "KR(한국)": [{"row": 7, "section_code": "//...", "text": "..."}, ...],
        "US(미국)": [...],
        ...
      }
    }
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  ⚠ 파일 열기 실패 {filepath.name}: {e}")
        return None

    # story_id: 첫 번째 시트의 C5
    first_ws = wb[wb.sheetnames[0]]
    story_id = str(first_ws[STORY_ID_CELL].value or "").strip()
    if not story_id:
        story_id = filepath.stem  # fallback

    result = {
        "story_id": story_id,
        "filename": filepath.name,
        "sheets": {}
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in range(CONTENT_ROW_START, CONTENT_ROW_END + 1):
            b_val = ws.cell(row=row, column=SECTION_COL).value
            c_val = ws.cell(row=row, column=CONTENT_COL).value
            section_code = str(b_val).strip() if b_val else ""
            text = str(c_val).strip() if c_val else ""
            # 빈 행, 'x' placeholder, None 값 스킵
            if text and text.lower() not in ("x", "none", ""):
                rows.append({
                    "row": row,
                    "section_code": section_code,
                    "text": text
                })
        if rows:
            result["sheets"][sheet_name] = rows

    return result


# ─── 단일 파일 DB 저장 ────────────────────────────────────────────────────────
def build_file(
    filepath: Path,
    gemini_client,
    col_kr: chromadb.Collection,
    col_us: chromadb.Collection,
    conn: sqlite3.Connection,
    log_fn=print
) -> int:
    """
    파일 1개를 파싱하여 ChromaDB + SQLite에 저장.
    반환: 저장된 레코드 수
    """
    parsed = parse_excel_file(filepath)
    if not parsed:
        return 0

    story_id = parsed["story_id"]
    filename = parsed["filename"]
    sheets = parsed["sheets"]

    kr_source_rows = sheets.get(GROUP_A_KEY, [])
    us_source_rows = sheets.get(GROUP_B_KEY, [])

    # source row를 dict로 빠른 조회 가능하게 (row번호 기준)
    kr_src_map = {r["row"]: r for r in kr_source_rows}
    us_src_map = {r["row"]: r for r in us_source_rows}

    records = []  # (id, source_group, source_text, target_lang, target_sheet, target_text, section_code)

    for sheet_name, rows in sheets.items():
        # Source 시트 자체는 target으로 저장하지 않음
        if sheet_name in (GROUP_A_KEY, GROUP_B_KEY):
            continue

        # Group 분기
        if sheet_name in GROUP_A_SHEETS:
            src_map = kr_src_map
            source_group = "kr"
        else:
            src_map = us_src_map
            source_group = "us"

        for row_data in rows:
            row_num = row_data["row"]
            target_text = row_data["text"]
            section_code = row_data["section_code"]

            src_row = src_map.get(row_num)
            if not src_row:
                continue  # 매칭되는 source 없으면 skip
            source_text = src_row["text"]
            source_text_norm = normalize_text(source_text)

            rec_id = f"{filename}::{sheet_name}::{row_num}"
            records.append({
                "id": rec_id,
                "story_id": story_id,
                "section_code": section_code,
                "source_group": source_group,
                "source_lang": "Korean" if source_group == "kr" else "English",
                "source_text_raw": source_text,
                "source_text_norm": source_text_norm,
                "target_lang": sheet_name,
                "target_sheet": sheet_name,
                "target_text": target_text,
                "original_file": filename
            })

    if not records:
        log_fn(f"  ⚠ [{story_id}] 저장할 레코드 없음 (시트 구조 확인 필요)")
        return 0

    # ─── ChromaDB 임베딩 & 저장 ────────────────────────────────────────────
    kr_recs = [r for r in records if r["source_group"] == "kr"]
    us_recs = [r for r in records if r["source_group"] == "us"]

    def upsert_to_collection(col, recs):
        if not recs:
            return
        # 임베딩은 원본 텍스트(raw)를 사용거나 정규화된 텍스트를 사용할 수 있음. 
        # 여기서는 좀 더 정제된 정규화 텍스트를 임베딩에 사용하여 유사도를 높임.
        texts = [r["source_text_norm"] for r in recs]
        try:
            embeddings = embed_texts(gemini_client, texts)
        except Exception as e:
            log_fn(f"  ⚠ 임베딩 오류: {e}")
            embeddings = [[0.0] * 768] * len(texts)  # fallback 더미

        col.upsert(
            ids=[r["id"] for r in recs],
            embeddings=embeddings,
            documents=[r["source_text_raw"] for r in recs],
            metadatas=[{
                "story_id": r["story_id"],
                "section_code": r["section_code"],
                "source_lang": r["source_lang"],
                "source_text_norm": r["source_text_norm"],
                "target_lang": r["target_lang"],
                "target_text": r["target_text"],
                "original_file": r["original_file"]
            } for r in recs]
        )

    upsert_to_collection(col_kr, kr_recs)
    upsert_to_collection(col_us, us_recs)

    # ─── SQLite 저장 ──────────────────────────────────────────────────────
    now = datetime.now().isoformat()
    conn.executemany("""
        INSERT OR REPLACE INTO rag_pairs
        (id, story_id, section_code, source_group, source_text, target_lang, target_sheet, target_text, original_file, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, [(
        r["id"], r["story_id"], r["section_code"], r["source_group"],
        r["source_text_raw"], r["target_lang"], r["target_sheet"],
        r["target_text"], r["original_file"], now
    ) for r in records])

    # processed_files 기록
    mtime = str(os.path.getmtime(filepath))
    conn.execute("""
        INSERT OR REPLACE INTO processed_files (filename, story_id, mtime, record_count, processed_at)
        VALUES (?, ?, ?, ?, ?)
    """, (filename, story_id, mtime, len(records), now))
    conn.commit()

    log_fn(f"  ✓ [{story_id}] {filename} → {len(records)}건 저장 (KR:{len(kr_recs)}, US:{len(us_recs)})")
    return len(records)


# ─── 스토리 삭제 (증분 업데이트용) ──────────────────────────────────────────────
def delete_story(story_id: str, col_kr, col_us, conn: sqlite3.Connection, log_fn=print):
    """특정 story_id의 모든 레코드 삭제"""
    # ChromaDB: where 필터로 삭제
    for col in (col_kr, col_us):
        existing = col.get(where={"story_id": story_id})
        if existing["ids"]:
            col.delete(ids=existing["ids"])
            log_fn(f"  ChromaDB에서 {len(existing['ids'])}건 삭제")

    # SQLite 삭제
    cur = conn.execute("DELETE FROM rag_pairs WHERE story_id = ?", (story_id,))
    conn.execute("DELETE FROM processed_files WHERE story_id = ?", (story_id,))
    conn.commit()
    log_fn(f"  SQLite에서 {cur.rowcount}건 삭제")


# ─── 파일 목록 조회 ───────────────────────────────────────────────────────────
def get_excel_files() -> list[Path]:
    pattern = str(EXCEL_DIR / "*.xlsx")
    files = [Path(f) for f in glob.glob(pattern) if not Path(f).name.startswith("~$")]
    return sorted(files)


def get_already_processed(conn: sqlite3.Connection) -> dict:
    """filename → mtime 매핑"""
    rows = conn.execute("SELECT filename, mtime FROM processed_files").fetchall()
    return {r["filename"]: r["mtime"] for r in rows}


# ─── 메인 빌드 함수 ───────────────────────────────────────────────────────────
def run_pilot(log_fn=print) -> dict:
    """파일럿: 첫 번째 엑셀 파일 1개만 빌드"""
    files = get_excel_files()
    if not files:
        log_fn("❌ @translation_data/@excel 폴더에 엑셀 파일이 없습니다.")
        return {}

    target_file = files[0]
    log_fn(f"🚀 [파일럿 빌드] 대상 파일: {target_file.name}")

    gemini_client = get_gemini_client()
    _, col_kr, col_us = get_chroma_collections()
    conn = get_sqlite_conn()

    count = build_file(target_file, gemini_client, col_kr, col_us, conn, log_fn)

    # 결과 요약
    story_id = parse_excel_file(target_file)["story_id"]
    sample = conn.execute(
        "SELECT source_text, target_lang, target_text FROM rag_pairs WHERE story_id = ? LIMIT 3",
        (story_id,)
    ).fetchall()

    summary = {
        "story_id": story_id,
        "file": target_file.name,
        "records_saved": count,
        "samples": [dict(r) for r in sample]
    }
    log_fn(f"\n📊 파일럿 결과 요약:")
    log_fn(f"  - 스토리: {story_id}")
    log_fn(f"  - 저장 레코드: {count}건")
    log_fn(f"  - 샘플 (최대 3건):")
    for s in sample:
        log_fn(f"    [{s['target_lang']}] {s['source_text'][:40]} → {s['target_text'][:40]}")
    return summary


def run_build_all(log_fn=print, force: bool = False) -> int:
    """전체 파일 빌드 (force=False면 mtime 체크하여 skip)"""
    if force:
        log_fn("🚀 [강제 재빌드 모드] 모든 파일을 처음부터 다시 색인합니다.")
    else:
        log_fn("📂 [증분 빌드 모드] 수정된 파일만 색인합니다.")

    files = get_excel_files()
    if not files:
        log_fn("❌ @translation_data/@excel 폴더에 엑셀 파일이 없습니다.")
        return 0

    gemini_client = get_gemini_client()
    _, col_kr, col_us = get_chroma_collections()
    conn = get_sqlite_conn()
    already = get_already_processed(conn) if not force else {}

    total = 0
    skipped = 0
    for i, filepath in enumerate(files, 1):
        fname = filepath.name
        mtime = str(os.path.getmtime(filepath))

        if fname in already and already[fname] == mtime:
            log_fn(f"  [{i}/{len(files)}] ⏭ 이미 처리됨, skip: {fname}")
            skipped += 1
            continue

        log_fn(f"  [{i}/{len(files)}] 처리 중: {fname}")
        cnt = build_file(filepath, gemini_client, col_kr, col_us, conn, log_fn)
        total += cnt

    log_fn(f"\n✅ 전체 빌드 완료: {total}건 저장, {skipped}개 파일 skip")
    return total


def run_update_story(story_id: str, log_fn=print) -> int:
    """특정 스토리 증분 업데이트"""
    _, col_kr, col_us = get_chroma_collections()
    conn = get_sqlite_conn()

    # 기존 레코드 삭제
    log_fn(f"🗑 [{story_id}] 기존 레코드 삭제 중...")
    delete_story(story_id, col_kr, col_us, conn, log_fn)

    # 해당 story_id 파일 찾기
    files = get_excel_files()
    gemini_client = get_gemini_client()

    target_files = []
    for f in files:
        parsed = parse_excel_file(f)
        if parsed and parsed.get("story_id") == story_id:
            target_files.append(f)

    if not target_files:
        log_fn(f"⚠ story_id='{story_id}'에 해당하는 파일을 찾지 못했습니다.")
        return 0

    total = 0
    for f in target_files:
        cnt = build_file(f, gemini_client, col_kr, col_us, conn, log_fn)
        total += cnt

    log_fn(f"✅ [{story_id}] 증분 업데이트 완료: {total}건 재저장")
    return total


def run_status(log_fn=print):
    """DB 현황 출력"""
    conn = get_sqlite_conn()
    _, col_kr, col_us = get_chroma_collections()

    total_pairs = conn.execute("SELECT COUNT(*) FROM rag_pairs").fetchone()[0]
    processed_files = conn.execute("SELECT COUNT(*) FROM processed_files").fetchone()[0]
    story_counts = conn.execute(
        "SELECT story_id, COUNT(*) as cnt FROM rag_pairs GROUP BY story_id ORDER BY story_id"
    ).fetchall()

    log_fn("📊 RAG DB 현황:")
    log_fn(f"  - 처리된 파일: {processed_files}개")
    log_fn(f"  - 총 번역 쌍: {total_pairs}건")
    log_fn(f"  - ChromaDB KR 컬렉션: {col_kr.count()}건")
    log_fn(f"  - ChromaDB US 컬렉션: {col_us.count()}건")
    log_fn(f"  - 스토리별 현황:")
    for row in story_counts:
        log_fn(f"    {row['story_id']}: {row['cnt']}건")

    return {
        "processed_files": processed_files,
        "total_pairs": total_pairs,
        "kr_vectors": col_kr.count(),
        "us_vectors": col_us.count(),
    }


# ─── 비동기 래퍼 (SSE용) ──────────────────────────────────────────────────────
async def build_all_async(log_fn, force: bool = False):
    """FastAPI SSE에서 호출하기 위한 비동기 래퍼"""
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, lambda: run_build_all(log_fn, force=force))


async def update_story_async(story_id: str, log_fn):
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, run_update_story, story_id, log_fn)


# ─── CLI 진입점 ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="SmartThings RAG DB Builder")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--pilot", action="store_true", help="파일럿: 첫 파일 1개만 빌드")
    group.add_argument("--build-all", action="store_true", help="전체 파일 빌드")
    parser.add_argument("--force", action="store_true", help="이미 처리된 파일도 강제 재빌드")
    group.add_argument("--update-story", metavar="STORY_ID", help="특정 스토리 증분 업데이트")
    group.add_argument("--status", action="store_true", help="DB 현황 조회")
    args = parser.parse_args()

    if args.pilot:
        run_pilot()
    elif args.build_all:
        run_build_all(force=args.force)
    elif args.update_story:
        run_update_story(args.update_story)
    elif args.status:
        run_status()
